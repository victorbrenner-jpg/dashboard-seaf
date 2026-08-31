import base64
import datetime
import html
import io
import json
import os
from pathlib import Path
import re
import sqlite3
import urllib.parse
import urllib.request
import unicodedata
import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
import streamlit.components.v1 as components
import textwrap

# 1. CONFIGURAÇÃO DA PÁGINA (Deve ser a primeira linha executável do Streamlit)
st.set_page_config(
    page_title="Painel de Controle Financeiro SEAF - 2026",
    layout="wide",
    initial_sidebar_state="expanded",
)


# -------------------------------------------------------------------------
# GERENCIAMENTO DE ESTADO, MEMÓRIA DE FILTROS E NAVEGAÇÃO ENTRE TELAS
# -------------------------------------------------------------------------
def sincronizar_filtro(chave_memoria, chave_widget):
    if chave_widget in st.session_state:
        st.session_state[chave_memoria] = st.session_state[chave_widget]


# Estado da Tela Ativa
if "tela_atual" not in st.session_state:
    st.session_state["tela_atual"] = "Liquidação (NL)"

# --- MEMÓRIA DA TELA 1 (Pagamentos - OB) ---
if "mem_ob_tipo_data" not in st.session_state:
    st.session_state["mem_ob_tipo_data"] = "Por Mês de Competência"
if "mem_ob_meses" not in st.session_state:
    st.session_state["mem_ob_meses"] = []
if "mem_ob_dt_ini" not in st.session_state:
    st.session_state["mem_ob_dt_ini"] = datetime.date(2026, 1, 1)
if "mem_ob_dt_fim" not in st.session_state:
    st.session_state["mem_ob_dt_fim"] = datetime.date(2026, 12, 31)
if "mem_ob_despesa" not in st.session_state:
    st.session_state["mem_ob_despesa"] = []
elif isinstance(st.session_state["mem_ob_despesa"], str):
    st.session_state["mem_ob_despesa"] = (
        []
        if st.session_state["mem_ob_despesa"] == "Todas as Despesas"
        else [st.session_state["mem_ob_despesa"]]
    )
if "mem_ob_credores" not in st.session_state:
    st.session_state["mem_ob_credores"] = []
if "mem_ob_fontes" not in st.session_state:
    st.session_state["mem_ob_fontes"] = []
if "mem_ob_objetos" not in st.session_state:
    st.session_state["mem_ob_objetos"] = []

# --- MEMÓRIA DA TELA 2 (Liquidação - NL) ---
if "mem_nl_tipo_periodo" not in st.session_state:
    st.session_state["mem_nl_tipo_periodo"] = "Por Mês de Competência"
if "mem_nl_comps" not in st.session_state:
    st.session_state["mem_nl_comps"] = []
if "mem_nl_datas" not in st.session_state:
    st.session_state["mem_nl_datas"] = None
if "mem_nl_dt_ini" not in st.session_state:
    intervalo_nl_anterior = st.session_state.get("mem_nl_datas")
    st.session_state["mem_nl_dt_ini"] = (
        intervalo_nl_anterior[0]
        if isinstance(intervalo_nl_anterior, (tuple, list))
        and len(intervalo_nl_anterior) == 2
        else None
    )
if "mem_nl_dt_fim" not in st.session_state:
    intervalo_nl_anterior = st.session_state.get("mem_nl_datas")
    st.session_state["mem_nl_dt_fim"] = (
        intervalo_nl_anterior[1]
        if isinstance(intervalo_nl_anterior, (tuple, list))
        and len(intervalo_nl_anterior) == 2
        else None
    )
if "mem_nl_grupo" not in st.session_state:
    st.session_state["mem_nl_grupo"] = []
elif isinstance(st.session_state["mem_nl_grupo"], str):
    # Mantém compatibilidade com a seleção única salva em sessões antigas.
    st.session_state["mem_nl_grupo"] = (
        []
        if st.session_state["mem_nl_grupo"] == "Todos"
        else [st.session_state["mem_nl_grupo"]]
    )
if "mem_nl_status" not in st.session_state:
    st.session_state["mem_nl_status"] = "Todos"
if "mem_nl_credores" not in st.session_state:
    st.session_state["mem_nl_credores"] = []
if "mem_nl_fonte" not in st.session_state:
    st.session_state["mem_nl_fonte"] = []
elif isinstance(st.session_state["mem_nl_fonte"], str):
    # Mantém compatibilidade com a seleção única salva em sessões antigas.
    st.session_state["mem_nl_fonte"] = (
        []
        if st.session_state["mem_nl_fonte"] == "Todas as fontes (Exibe tudo)"
        else [st.session_state["mem_nl_fonte"]]
    )
if "mem_nl_objetos" not in st.session_state:
    st.session_state["mem_nl_objetos"] = []

# --- MEMÓRIA DA TELA 3 (Priorização semanal) ---
# Os widgets desta tela não são renderizados quando o usuário navega para NL
# ou OB. Por isso os valores selecionados ficam em chaves independentes.
if "mem_plan_grupos" not in st.session_state:
    st.session_state["mem_plan_grupos"] = []
if "mem_plan_tipos_nl" not in st.session_state:
    st.session_state["mem_plan_tipos_nl"] = []
if "mem_plan_mes_programacao" not in st.session_state:
    # No primeiro acesso, abre no mês corrente. A seleção feita pelo usuário
    # continua preservada durante a navegação entre as telas.
    st.session_state["mem_plan_mes_programacao"] = datetime.date.today().strftime("%m/%Y")

# Estilização CSS
st.markdown(
    """
    <style>
    .titulo-pagina {
        color: #002b49;
        font-size: 2.25rem !important;
        font-weight: 700 !important;
        margin-bottom: 4px !important;
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif;
    }
    .barra-sistema {
        display: flex;
        align-items: center;
        justify-content: space-between;
        background: linear-gradient(90deg, #002b49 0%, #005691 52%, #028090 100%);
        color: #ffffff;
        border-radius: 9px 9px 0 0;
        padding: 10px 18px;
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif;
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        width: 100vw !important;
        z-index: 100002 !important;
        box-shadow: 0 3px 10px rgba(0, 43, 73, 0.25) !important;
    }
    .barra-sistema .marca-sistema {
        font-weight: 800;
        letter-spacing: 0.3px;
        font-size: 15px;
    }
    .barra-sistema .exercicio-sistema {
        font-size: 12px;
        font-weight: 700;
        opacity: 0.95;
    }
    .st-key-seletor_tela_global {
        background: #eef4fa;
        border: 1px solid #cbd5e1;
        border-top: 0;
        border-radius: 0 0 9px 9px;
        padding: 7px 12px 8px 12px;
        margin-bottom: 18px;
        position: fixed !important;
        top: 41px !important;
        left: 0 !important;
        width: 100vw !important;
        z-index: 100001 !important;
        box-shadow: 0 5px 10px rgba(15, 23, 42, 0.14) !important;
    }
    /* Cabeçalho único e fixo: a navegação nunca se separa da barra SEAF. */
    .st-key-topo_navegacao {
        position: static !important;
    }
    .st-key-topo_navegacao .barra-sistema {
        border-radius: 0 !important;
    }
    .st-key-topo_navegacao .st-key-seletor_tela_global {
        margin-bottom: 0 !important;
        padding: 8px 22px 9px !important;
    }
    /* Navegação por módulos: abas corporativas, com módulo ativo evidente. */
    .st-key-seletor_tela_global [data-testid="stSegmentedControl"] {
        background: transparent !important;
    }
    .st-key-seletor_tela_global [data-testid="stSegmentedControl"] [role="radiogroup"],
    .st-key-seletor_tela_global [data-baseweb="button-group"] {
        gap: 7px !important;
        background: transparent !important;
    }
    .st-key-seletor_tela_global [data-testid="stSegmentedControl"] button,
    .st-key-seletor_tela_global button[data-testid*="segmented_control"] {
        min-height: 36px !important;
        padding: 0 17px !important;
        border: 1px solid #c7d5e2 !important;
        border-radius: 7px !important;
        background: #ffffff !important;
        color: #334155 !important;
        font-size: 13px !important;
        font-weight: 650 !important;
        box-shadow: 0 2px 5px rgba(15, 42, 68, 0.07) !important;
        transition: all .18s ease !important;
    }
    .st-key-seletor_tela_global [data-testid="stSegmentedControl"] button:hover,
    .st-key-seletor_tela_global button[data-testid*="segmented_control"]:hover {
        background: #f0fdfa !important;
        border-color: #028090 !important;
        color: #005b63 !important;
        transform: translateY(-1px);
        box-shadow: 0 5px 10px rgba(2, 128, 144, 0.14) !important;
    }
    .st-key-seletor_tela_global [data-testid="stSegmentedControl"] button[aria-pressed="true"],
    .st-key-seletor_tela_global button[data-testid*="segmented_control"][aria-pressed="true"] {
        border-color: #007b84 !important;
        background: linear-gradient(135deg, #005691 0%, #028090 100%) !important;
        color: #ffffff !important;
        box-shadow: 0 5px 12px rgba(0, 86, 145, 0.25) !important;
    }
    .st-key-seletor_tela_global [data-testid="stSegmentedControl"] button[aria-pressed="true"]:hover,
    .st-key-seletor_tela_global button[data-testid*="segmented_control"][aria-pressed="true"]:hover {
        background: linear-gradient(135deg, #004a7c 0%, #01757d 100%) !important;
        color: #ffffff !important;
    }
    [data-testid="stMainBlockContainer"] {
        padding-top: 118px !important;
    }
    /* Mantém o cabeçalho técnico invisível, sem eliminar o controle nativo
       que permite reabrir os filtros quando a barra lateral for recolhida. */
    header[data-testid="stHeader"] {
        background: transparent !important;
        height: 0 !important;
        z-index: 100003 !important;
    }
    header[data-testid="stHeader"] [data-testid="stToolbar"],
    header[data-testid="stHeader"] [data-testid="stDecoration"] {
        display: none !important;
    }
    [data-testid="stSidebarCollapsedControl"] {
        display: none !important;
    }
    [data-testid="stSidebar"] {
        z-index: 100000 !important;
        padding-top: 108px !important;
        /* Aba lateral de filtros: abre apenas ao passar o mouse ou navegar
           pelos seus campos com o teclado. */
        display: block !important;
        visibility: visible !important;
        transform: translateX(0) !important;
        margin-left: 0 !important;
        min-width: 28px !important;
        width: 28px !important;
        max-width: 28px !important;
        overflow: hidden !important;
        transition: width .22s ease .55s, min-width .22s ease .55s, max-width .22s ease .55s !important;
        background: #f1f5f9 !important;
        border-right: 1px solid #d4e0ea !important;
    }
    [data-testid="stSidebar"][aria-expanded="false"] {
        min-width: 28px !important;
        width: 28px !important;
        max-width: 28px !important;
        transform: translateX(0) !important;
        visibility: visible !important;
    }
    [data-testid="stSidebar"]:hover,
    [data-testid="stSidebar"]:focus-within,
    [data-testid="stSidebar"][aria-expanded="false"]:hover,
    [data-testid="stSidebar"][aria-expanded="false"]:focus-within {
        min-width: 312px !important;
        width: 312px !important;
        max-width: 312px !important;
        overflow-y: auto !important;
        transition-delay: 0s !important;
    }
    [data-testid="stSidebar"] > div:first-child {
        min-width: 312px !important;
        width: 312px !important;
    }
    [data-testid="stSidebar"] [data-testid="stSidebarUserContent"] {
        visibility: visible !important;
        opacity: 1 !important;
    }
    [data-testid="stSidebar"]::after {
        content: "FILTROS";
        position: fixed;
        top: 138px;
        left: 5px;
        z-index: 100005;
        writing-mode: vertical-rl;
        transform: rotate(180deg);
        color: #ffffff;
        background: #005691;
        border-radius: 0 6px 6px 0;
        padding: 10px 5px;
        font-size: 10px;
        font-weight: 800;
        letter-spacing: .5px;
        pointer-events: none;
    }
    [data-testid="stSidebar"]:hover::after,
    [data-testid="stSidebar"]:focus-within::after {
        display: none;
    }
    .subtitulo-pagina {
        color: #6c757d;
        font-size: 1.25rem !important;
        font-style: italic;
        margin-top: 0px !important;
        margin-bottom: 2px !important;
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif;
    }
    .metric-card {
        background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
        border-left: 6px solid #002b49;
        padding: 12px 16px;
        border-radius: 8px;
        box-shadow: 0 4px 12px rgba(0, 43, 73, 0.05);
        border: 1px solid #e2e8f0;
    }
    /* Botões: contraste e hierarquia para ações de decisão. */
    div.stButton > button,
    div.stDownloadButton > button {
        min-height: 38px;
        border: 1px solid #b8c9d8;
        border-radius: 7px;
        background: linear-gradient(180deg, #ffffff 0%, #f4f7fa 100%);
        color: #002b49;
        font-weight: 700;
        box-shadow: 0 2px 5px rgba(15, 42, 68, 0.08);
        transition: transform .16s ease, box-shadow .16s ease, border-color .16s ease, background .16s ease;
    }
    div.stButton > button:hover,
    div.stDownloadButton > button:hover {
        border-color: #028090;
        background: #f0fdfa;
        box-shadow: 0 5px 12px rgba(2, 128, 144, 0.16);
        transform: translateY(-1px);
    }
    div.stButton > button:active,
    div.stDownloadButton > button:active {
        transform: translateY(0);
        box-shadow: 0 1px 3px rgba(15, 42, 68, 0.12);
    }
    div.stButton > button:focus-visible,
    div.stDownloadButton > button:focus-visible {
        outline: 3px solid rgba(2, 128, 144, 0.25);
        outline-offset: 2px;
    }
    div.stDownloadButton > button,
    div.st-key-btn_gerar_relatorio_nl button,
    div.st-key-btn_gerar_relatorio_pd button,
    [class*="st-key-adicionar_objeto_"] button {
        border-color: #007b84;
        background: linear-gradient(135deg, #005691 0%, #028090 100%);
        color: #ffffff;
        box-shadow: 0 4px 10px rgba(0, 86, 145, 0.22);
    }
    div.stDownloadButton > button:hover,
    div.st-key-btn_gerar_relatorio_nl button:hover,
    div.st-key-btn_gerar_relatorio_pd button:hover,
    [class*="st-key-adicionar_objeto_"] button:hover {
        border-color: #006f78;
        background: linear-gradient(135deg, #004a7c 0%, #01757d 100%);
        color: #ffffff;
    }
    div.stDownloadButton > button:disabled,
    div.st-key-btn_gerar_relatorio_nl button:disabled,
    div.st-key-btn_gerar_relatorio_pd button:disabled,
    [class*="st-key-adicionar_objeto_"] button:disabled {
        opacity: 1 !important;
        border-color: #cbd5e1 !important;
        background: #e8eef3 !important;
        color: #64748b !important;
        box-shadow: none !important;
        cursor: not-allowed !important;
        transform: none !important;
    }
    div.stDownloadButton > button:disabled:hover,
    div.st-key-btn_gerar_relatorio_nl button:disabled:hover,
    div.st-key-btn_gerar_relatorio_pd button:disabled:hover,
    [class*="st-key-adicionar_objeto_"] button:disabled:hover {
        border-color: #cbd5e1 !important;
        background: #e8eef3 !important;
        color: #64748b !important;
        box-shadow: none !important;
        transform: none !important;
    }
    [class*="st-key-limpar_"] button {
        border-color: #d8b4b4;
        color: #9f2d2d;
        background: #fffafa;
        box-shadow: none;
    }
    [class*="st-key-limpar_"] button:hover {
        border-color: #c65b5b;
        color: #8b2525;
        background: #fff1f1;
        box-shadow: 0 4px 10px rgba(159, 45, 45, 0.12);
    }
    /* Ação de adicionar: mantém o rótulo visível antes, durante e depois do hover. */
    div.stButton[class*="st-key-adicionar_objeto_"] > button,
    [class*="st-key-adicionar_objeto_"] button {
        border-color: #d8b4b4 !important;
        color: #9f2d2d !important;
        background: #fffafa !important;
        box-shadow: none !important;
    }
    div.stButton[class*="st-key-adicionar_objeto_"] > button:hover:not(:disabled),
    [class*="st-key-adicionar_objeto_"] button:hover:not(:disabled) {
        border-color: #c65b5b !important;
        color: #8b2525 !important;
        background: #fff1f1 !important;
        box-shadow: 0 4px 10px rgba(198, 91, 91, 0.14) !important;
        transform: translateY(-1px);
    }
    /* Grades da priorização: cabeçalho institucional e leitura reforçada. */
    [data-testid="stDataFrame"] [role="columnheader"] {
        background: #005691 !important;
        color: #ffffff !important;
        border-color: #004a7c !important;
        font-weight: 700 !important;
    }
    [data-testid="stDataFrame"] [role="columnheader"] * {
        color: #ffffff !important;
        fill: #ffffff !important;
        font-weight: 700 !important;
    }
    [data-testid="stDataFrame"] [role="gridcell"],
    [data-testid="stDataFrame"] [role="gridcell"] * {
        color: #173a5e !important;
    }
    [data-testid="stDataFrame"] [role="gridcell"] {
        border-color: #d8e1ea !important;
    }
    .div-titulo {
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif;
        color: #002b49;
        font-weight: 800;
        letter-spacing: -0.5px;
    }
    .tabela-container {
        width: 100% !important;
        margin-bottom: 30px !important;
        overflow-x: auto !important; 
        overflow-y: hidden !important;
        background-color: #ffffff;
        border-radius: 6px;
        border: 1px solid #e2e8f0;
    }
    .subtitulo-tabela-html {
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif !important;
        font-size: 14.5px !important;
        font-weight: 700 !important;
        padding: 14px 20px !important;
        color: #ffffff !important;
        margin: 0px !important;
        letter-spacing: 0.5px;
    }
    .html-executiva {
        width: 100% !important;
        min-width: 1350px !important; 
        border-collapse: collapse !important;
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif !important;
        font-size: 13px !important;
    }
    .html-executiva th {
        font-weight: 700 !important;
        padding: 14px 12px !important;
        text-align: center !important;
        border-bottom: 2px solid #cbd5e1 !important;
        background-color: #f8fafc !important;
        color: #475569 !important;
        text-transform: uppercase;
        font-size: 11px;
        letter-spacing: 0.7px;
        white-space: nowrap !important; 
    }
    .html-executiva th:first-child {
        text-align: left !important;
        padding-left: 20px !important;
        width: 250px !important; 
    }
    .html-executiva td {
        padding: 12px 12px !important;
        border-bottom: 1px solid #f1f5f9 !important;
        text-align: center !important;
        color: #334155 !important;
        white-space: nowrap !important;
    }
    .html-executiva td:first-child {
        text-align: left !important;
        padding-left: 20px !important;
        font-weight: 600;
        color: #0f172a !important;
        white-space: normal !important;
    }
    .html-executiva th:last-child, 
    .html-executiva td:last-child {
        font-weight: bold !important;
        background-color: #f8fafc !important;
        text-align: right !important;
        padding-right: 20px !important;
        width: 140px !important;
    }
    .gnd-badge {
        display: inline-block;
        width: 9px;
        height: 9px;
        border-radius: 50%;
        margin-right: 10px;
        vertical-align: middle;
    }
    .html-executiva tbody tr:nth-child(even) {
        background-color: #fdfdfd !important;
    }
    .html-executiva tbody tr:hover {
        background-color: #f1f5f9 !important;
    }
    .linha-total-html {
        font-weight: bold !important;
        background-color: #f8fafc !important;
    }
    .linha-total-html td {
        border-top: 2px solid #002b49 !important;
        border-bottom: 3px double #002b49 !important;
        color: #002b49 !important;
        font-size: 13.5px !important;
        font-weight: 700 !important;
    }
    .tabela-dinamica-container {
        width: 100%;
        border: 1px solid #cbd5e1;
        border-radius: 6px;
        background: #ffffff;
        overflow: hidden;
        font-family: 'Segoe UI', sans-serif;
        font-size: 13px;
    }
    .tabela-dinamica-header {
        display: grid;
        grid-template-columns: 2.2fr 1fr 1fr;
        background-color: #092e4d;
        padding: 10px 14px;
        border-bottom: 2px solid #005691;
        font-weight: 700;
        font-size: 11px;
        color: #ffffff;
        letter-spacing: 0.5px;
        text-transform: uppercase;
        align-items: center;
    }
    .tabela-dinamica-container > .tabela-dinamica-header:last-child {
        font-size: 13px;
    }
    details.credor-group {
        border-bottom: 1px solid #e2e8f0;
        background-color: #ffffff;
        transition: background-color 0.2s ease;
    }
    details.credor-group:hover {
        background-color: #f8fafc;
    }
    summary.credor-summary {
        display: grid;
        grid-template-columns: 2.2fr 1fr 1fr;
        padding: 10px 14px;
        cursor: pointer;
        font-weight: 700;
        color: #002b49;
        user-select: none;
        align-items: center;
    }
    summary.credor-summary::-webkit-details-marker {
        display: inline-block;
        margin-right: 6px;
    }
    .subtabela-container {
        padding: 6px 10px 10px 10px;
        background: #f8fafc;
        border-top: 1px solid #e2e8f0;
    }
    .subtabela-detalhe {
        width: 100%;
        border-collapse: collapse;
        font-size: 12px;
        table-layout: fixed;
    }
    .subtabela-detalhe th {
        background-color: #edf2f7;
        color: #475569;
        font-size: 10px;
        text-transform: uppercase;
        padding: 6px 8px;
        border-bottom: 1px solid #cbd5e1;
        vertical-align: middle;
        text-align: center;
    }
    .subtabela-detalhe td {
        padding: 7px 8px;
        border-bottom: 1px solid #e2e8f0;
        color: #334155;
        vertical-align: middle;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        text-align: center;
    }
    .tabela-simples-container {
        width: 100%;
        border: 1px solid #cbd5e1;
        border-radius: 6px;
        background: #ffffff;
        overflow: hidden;
    }
    .tabela-simples {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Segoe UI', sans-serif;
        font-size: 12px;
        table-layout: auto;
        margin: 0 !important;
    }
    .tabela-simples th {
        color: #ffffff;
        font-weight: 700;
        font-size: 12px;
        letter-spacing: 0.3px;
        text-transform: uppercase;
        padding: 9px 12px;
        border-bottom: 2px solid #092e4d;
        background-color: #092e4d;
        vertical-align: middle;
    }
    .tabela-simples td {
        padding: 9px 12px;
        border-bottom: 1px solid #f1f5f9;
        color: #334155;
        vertical-align: middle;
    }
    .tabela-simples tr.total-row td {
        color: #002b49;
        font-weight: 700;
        font-size: 12px;
        border-top: 2px solid #005691;
        border-bottom: none;
        background-color: #f1f5f9;
        padding: 9px 12px !important;
        line-height: 1.2 !important;
        vertical-align: middle;
    }
    .tabela-fontes td:first-child {
        font-size: 12px;
        line-height: 1.25;
        padding-right: 8px;
    }
    .tabela-fontes th:first-child {
        width: 68%;
    }
    .tabela-fontes th:last-child {
        width: 32%;
        white-space: nowrap;
    }
    .progresso-priorizacao {
        margin: 10px 0 18px 0;
    }
    .progresso-priorizacao-texto {
        display: block;
        min-height: 22px;
        line-height: 22px;
        color: #334155;
        font-size: 13px;
        font-weight: 500;
    }
    .progresso-priorizacao-texto strong {
        color: #028090;
        font-weight: 700;
    }
    .progresso-priorizacao-trilha {
        width: 100%;
        height: 8px;
        margin-top: 5px;
        border-radius: 999px;
        overflow: hidden;
        background: #e2e8f0;
    }
    .progresso-priorizacao-preenchimento {
        height: 100%;
        min-width: 0;
        border-radius: inherit;
        background: linear-gradient(90deg, #028090, #0ea5a8);
    }
    /* Grades da Priorização: contraste institucional, sem o cinza apagado
       do componente padrão do Streamlit. */
    .tabela-priorizacao-executiva {
        width: 100%;
        overflow: auto;
        border: 1px solid #b9cce0;
        border-radius: 8px;
        background: #ffffff;
        box-shadow: 0 2px 7px rgba(15, 46, 76, 0.08);
    }
    .tabela-priorizacao-executiva-grade {
        display: grid;
        min-width: 780px;
        font-size: 13px;
        color: #173a5e;
    }
    .tabela-priorizacao-executiva-cabecalho {
        padding: 11px 10px;
        position: sticky;
        top: 0;
        z-index: 3;
        background: linear-gradient(90deg, #315b85, #4d7fa8);
        color: #ffffff !important;
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 0.15px;
        text-align: left;
        border-right: 1px solid rgba(255,255,255,0.28);
        white-space: nowrap;
    }
    .tabela-priorizacao-executiva-celula {
        padding: 10px;
        color: #173a5e !important;
        background: #ffffff;
        border-top: 1px solid #dce6ef;
        border-right: 1px solid #e4ebf2;
        vertical-align: middle;
    }
    .tabela-priorizacao-executiva-celula.centralizado { text-align: center; }
    .tabela-priorizacao-executiva-celula.linha-par { background: #f8fbfe; }
    .tabela-priorizacao-executiva-celula.linha-total {
        background: #e7f0f8 !important;
        color: #002b49 !important;
        font-weight: 700;
        border-top: 2px solid #005691;
    }
    </style>
""",
    unsafe_allow_html=True,
)


def formatar_brl(valor):
    val_str = (
        f"{valor:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
    )
    return f"R$ {val_str}"


def renderizar_tabela_priorizacao(
    df, colunas_centralizadas=None, altura_maxima=None, template_colunas=None,
    largura_minima=780,
):
    """Renderiza tabelas somente-leitura com contraste executivo estável.

    O componente nativo do Streamlit usa canvas em algumas versões, o que pode
    ignorar estilos de cabeçalho. Nesta tela as grades são apenas de consulta;
    por isso a tabela HTML preserva o padrão visual institucional.
    """
    colunas_centralizadas = set(colunas_centralizadas or [])
    dados = df.copy().fillna("")
    if dados.empty:
        st.info("Não há registros para os filtros selecionados.")
        return

    quantidade_colunas = len(dados.columns)
    if template_colunas and len(template_colunas) == quantidade_colunas:
        template_grade = " ".join(template_colunas)
    else:
        template_grade = f"repeat({quantidade_colunas}, minmax(110px, 1fr))"
    itens_grade = [
        f'<div class="tabela-priorizacao-executiva-cabecalho">{html.escape(str(coluna))}</div>'
        for coluna in dados.columns
    ]
    for indice_linha, (_, registro) in enumerate(dados.iterrows()):
        eh_total = any(
            str(valor).strip().upper() == "TOTAL GERAL" for valor in registro.values
        )
        classe_linha = "linha-total" if eh_total else ("linha-par" if indice_linha % 2 else "")
        for coluna, valor in registro.items():
            classes = ["tabela-priorizacao-executiva-celula", classe_linha]
            if coluna in colunas_centralizadas:
                classes.append("centralizado")
            texto = html.escape(str(valor))
            itens_grade.append(f'<div class="{" ".join(classes)}">{texto}</div>')

    estilo_altura = f"max-height:{int(altura_maxima)}px;" if altura_maxima else ""
    st.markdown(
        f"""
        <div class="tabela-priorizacao-executiva" style="{estilo_altura}">
          <div class="tabela-priorizacao-executiva-grade"
               style="grid-template-columns: {template_grade}; min-width: {int(largura_minima)}px;">
            {"".join(itens_grade)}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def renderizar_tabela_credor_pd(df_filtrado):
    """Tabela expansível de credores, seguindo o padrão da tela de Liquidação."""
    if df_filtrado.empty:
        st.info("Não há PDs para os filtros selecionados.")
        return

    totais = (
        df_filtrado.groupby("Nome do Credor", dropna=False)
        .agg(**{"Qtd. PDs": ("Número PD", "nunique"), "Valor total": ("Valor", "sum")})
        .sort_values("Valor total", ascending=False)
    )
    grupos_html = []
    for credor, total in totais.iterrows():
        credor_texto = str(credor or "NÃO INFORMADO")
        dados_credor = (
            df_filtrado[df_filtrado["Nome do Credor"] == credor]
            .sort_values(["Data Emissão", "Número PD"], na_position="last")
        )
        linhas = []
        for _, registro in dados_credor.iterrows():
            data = registro["Data Emissão"]
            data_texto = data.strftime("%d/%m/%Y") if pd.notna(data) else "—"
            linhas.append(
                "<tr>"
                f"<td title='{html.escape(str(registro['Número PD']))}' style='width:22%; font-family:monospace; font-weight:600; color:#005691;'>{html.escape(str(registro['Número PD']))}</td>"
                f"<td style='width:18%;'>{data_texto}</td>"
                f"<td style='width:14%;'>{html.escape(str(registro['Fonte']))}</td>"
                f"<td style='width:22%;'>{html.escape(str(registro['Despesa']))}</td>"
                f"<td style='width:24%; font-weight:700; color:#002b49;'>{formatar_brl(registro['Valor'])}</td>"
                "</tr>"
            )
        grupos_html.append(
            "<details class='credor-group'>"
            "<summary class='credor-summary'>"
            f"<span>{html.escape(credor_texto)}</span>"
            f"<span style='text-align:right; color:#475569;'>{int(total['Qtd. PDs'])} PD(s)</span>"
            f"<span style='text-align:right; color:#002b49;'>{formatar_brl(total['Valor total'])}</span>"
            "</summary>"
            "<div class='subtabela-container'><table class='subtabela-detalhe'><thead><tr>"
            "<th>PD</th><th>EMISSÃO</th><th>FONTE</th><th>DESPESA</th><th>VALOR</th>"
            "</tr></thead><tbody>"
            + "".join(linhas)
            + "</tbody></table></div></details>"
        )

    total_qtd = int(totais["Qtd. PDs"].sum())
    total_valor = float(totais["Valor total"].sum())
    html_final = (
        "<div class='tabela-dinamica-container'>"
        "<div class='tabela-dinamica-header'><div>CREDOR</div>"
        "<div style='text-align:right;'>QTD. PDs</div><div style='text-align:right;'>VALOR TOTAL</div></div>"
        + "".join(grupos_html)
        + "<div class='tabela-dinamica-header' style='background:#f1f5f9; border-top:2px solid #005691; border-bottom:none;'>"
        "<div style='color:#002b49; font-weight:800;'>TOTAL GERAL</div>"
        f"<div style='text-align:right; color:#475569; font-weight:800;'>{total_qtd} PD(s)</div>"
        f"<div style='text-align:right; color:#028090; font-weight:800;'>{formatar_brl(total_valor)}</div>"
        "</div></div>"
    )
    if hasattr(st, "html"):
        st.html(html_final)
    else:
        st.markdown(html_final, unsafe_allow_html=True)


def renderizar_tabela_resumo_pd(df, coluna_descricao, titulo_descricao):
    """Resumo compacto com o mesmo cabeçalho usado nas tabelas da Liquidação."""
    linhas = []
    for _, registro in df.iterrows():
        descricao = html.escape(str(registro[coluna_descricao]))
        valor = formatar_brl(registro["Valor total"])
        linhas.append(
            f"<tr><td>{descricao}</td><td style='text-align:right; font-weight:600; color:#002b49; white-space:nowrap;'>{valor}</td></tr>"
        )
    total = float(df["Valor total"].sum()) if not df.empty else 0.0
    tabela_html = (
        "<div class='tabela-simples-container'><table class='tabela-simples tabela-fontes'>"
        "<thead><tr>"
        f"<th style='text-align:left;'>{html.escape(titulo_descricao)}</th>"
        "<th style='text-align:right;'>VALOR TOTAL</th>"
        "</tr></thead><tbody>"
        + "".join(linhas)
        + "<tr class='total-row'><td>TOTAL GERAL</td>"
        f"<td style='text-align:right; color:#028090; white-space:nowrap;'>{formatar_brl(total)}</td>"
        "</tr></tbody></table></div>"
    )
    st.markdown(tabela_html, unsafe_allow_html=True)


def gerar_relatorio_pd_excel(df_filtrado):
    """Preenche o modelo oficial de PD sem remover as tabelas dinâmicas nativas."""
    if df_filtrado.empty:
        raise ValueError("Não há PDs para exportar com os filtros selecionados.")

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as erro:
        raise ModuleNotFoundError(
            "O pacote openpyxl não está instalado. Execute: pip install -r requirements.txt"
        ) from erro

    # O arquivo modelo possui uma aba-base com sete colunas. Essa base alimenta
    # as tabelas dinâmicas da aba Tabela_Dinâmica, portanto o modelo precisa ser
    # aberto e preenchido em vez de criarmos um novo Excel do zero.
    def normalizar_nome_coluna(valor):
        texto = unicodedata.normalize("NFKD", str(valor or ""))
        texto = "".join(
            caractere for caractere in texto
            if not unicodedata.combining(caractere)
        )
        return re.sub(r"[^a-z0-9]+", "", texto.lower())

    def localizar_coluna(*opcoes):
        colunas_normalizadas = {
            normalizar_nome_coluna(coluna): coluna
            for coluna in df_filtrado.columns
        }
        for opcao in opcoes:
            coluna = colunas_normalizadas.get(normalizar_nome_coluna(opcao))
            if coluna is not None:
                return coluna
        return None

    def serie_texto(coluna, padrao=""):
        if coluna is None:
            return pd.Series(padrao, index=df_filtrado.index, dtype="object")
        return df_filtrado[coluna].fillna(padrao).astype(str).str.strip()

    def serie_numero(coluna):
        if coluna is None:
            return pd.Series(0.0, index=df_filtrado.index, dtype="float64")
        valores = df_filtrado[coluna]
        if pd.api.types.is_numeric_dtype(valores):
            return pd.to_numeric(valores, errors="coerce").fillna(0.0)
        return converter_valor_monetario(valores).fillna(0.0)

    coluna_credor = localizar_coluna(
        "Nome do Credor", "Credor_Tratado", "Credor",
        "Entidade / Credor", "Entidade"
    )
    coluna_objeto = localizar_coluna(
        "Objeto da Despesa", "Objeto Despesa", "Objeto de Despesa", "Objeto"
    )
    coluna_grupo = localizar_coluna(
        "GD", "Grupo", "GND", "Grupo de Despesa"
    )
    coluna_numero = localizar_coluna(
        "Número PD", "Numero PD", "Número", "Numero", "PD"
    )
    coluna_tipo_pd = localizar_coluna(
        "Tipo de PD", "Tipo PD", "Tipo_PD",
        "Tipo de OB", "Tipo OB", "OB"
    )
    coluna_status = localizar_coluna("Status")
    coluna_valor = localizar_coluna(
        "Valor", "Valor Total", "Valor PD", "Valor Programado"
    )

    campos_obrigatorios = {
        "nome do credor": coluna_credor,
        "objeto da despesa": coluna_objeto,
        "grupo": coluna_grupo,
        "número da PD": coluna_numero,
        "valor": coluna_valor,
    }
    campos_ausentes = [
        nome for nome, coluna in campos_obrigatorios.items()
        if coluna is None
    ]
    if campos_ausentes:
        raise ValueError(
            "Não foi possível montar o relatório de PD. Colunas ausentes: "
            + ", ".join(campos_ausentes)
        )

    grupo_relatorio = serie_texto(coluna_grupo, padrao="NÃO INFORMADO").str.upper()
    grupo_relatorio = grupo_relatorio.replace(
        {
            "GD1": "1",
            "GD3": "3",
            "GD4": "4",
            "": "NÃO INFORMADO",
            "NAN": "NÃO INFORMADO",
        }
    )

    base_relatorio_modelo = pd.DataFrame(
        {
            "Nome do Credor": serie_texto(coluna_credor),
            "Objeto Despesa": serie_texto(coluna_objeto),
            "GD": grupo_relatorio,
            "Número": serie_texto(coluna_numero),
            "Tipo de PD": serie_texto(coluna_tipo_pd, padrao="NÃO INFORMADO"),
            "Status": serie_texto(coluna_status, padrao="NÃO INFORMADO"),
            "Valor": serie_numero(coluna_valor),
        }
    ).sort_values(
        ["Objeto Despesa", "Nome do Credor", "Número"], kind="stable"
    )

    caminho_modelo = (
        Path(__file__).resolve().parent
        / "modelos"
        / "modelo_relatorio_programa_desembolso.xlsx"
    )
    if not caminho_modelo.exists():
        raise FileNotFoundError(
            "Modelo não encontrado. Inclua "
            "modelos/modelo_relatorio_programa_desembolso.xlsx no repositório."
        )

    workbook = load_workbook(caminho_modelo, data_only=False)
    cabecalhos_esperados = [
        "Nome do Credor",
        "Objeto Despesa",
        "GD",
        "Número",
        "Tipo de PD",
        "Status",
        "Valor",
    ]

    def normalizar_cabecalho_excel(valor):
        texto = unicodedata.normalize("NFKD", str(valor or ""))
        texto = "".join(
            caractere for caractere in texto
            if not unicodedata.combining(caractere)
        )
        return re.sub(r"[^A-Z0-9]+", "", texto.upper())

    conjunto_esperado = {
        normalizar_cabecalho_excel(cabecalho)
        for cabecalho in cabecalhos_esperados
    }
    aba_base = None
    linha_cabecalho = None
    for aba in workbook.worksheets:
        for linha in range(1, min(aba.max_row, 10) + 1):
            cabecalhos_encontrados = {
                normalizar_cabecalho_excel(aba.cell(linha, coluna).value)
                for coluna in range(1, len(cabecalhos_esperados) + 1)
            }
            if conjunto_esperado.issubset(cabecalhos_encontrados):
                aba_base = aba
                linha_cabecalho = linha
                break
        if aba_base is not None:
            break

    if aba_base is None:
        raise ValueError(
            "A aba-base do modelo de PD não foi localizada. O modelo precisa "
            "conter as colunas: " + ", ".join(cabecalhos_esperados)
        )

    # Mantém os cabeçalhos exatamente como o modelo das tabelas dinâmicas espera.
    for coluna, cabecalho in enumerate(cabecalhos_esperados, start=1):
        aba_base.cell(linha_cabecalho, coluna).value = cabecalho

    primeira_linha_dados = linha_cabecalho + 1

    # Limpa somente os dados antigos da base. A aba Tabela_Dinâmica e seus
    # objetos permanecem intactos.
    for linha in aba_base.iter_rows(
        min_row=primeira_linha_dados,
        max_row=max(aba_base.max_row, primeira_linha_dados),
        min_col=1,
        max_col=len(cabecalhos_esperados),
    ):
        for celula in linha:
            celula.value = None

    for numero_linha, registro in enumerate(
        base_relatorio_modelo.itertuples(index=False, name=None),
        start=primeira_linha_dados,
    ):
        for numero_coluna, valor in enumerate(registro, start=1):
            if numero_coluna == 3:
                grupo = str(valor).replace("GD", "").strip()
                valor = int(grupo) if grupo.isdigit() else grupo
            elif numero_coluna == 7:
                valor = float(valor) if pd.notna(valor) else 0.0
            else:
                valor = "" if pd.isna(valor) else str(valor)

            celula = aba_base.cell(numero_linha, numero_coluna, value=valor)
            if numero_coluna == 7:
                celula.number_format = 'R$ #,##0.00'

    # O modelo usa uma fonte dinâmica para as tabelas dinâmicas. Forçamos a
    # atualização ao abrir o arquivo para que os consolidados reflitam os dados
    # recém-gravados, exatamente como já ocorre no relatório de NL.
    for aba in workbook.worksheets:
        for tabela_dinamica in getattr(aba, "_pivots", []):
            cache = getattr(tabela_dinamica, "cache", None)
            if cache is not None:
                cache.refreshOnLoad = True
                cache.enableRefresh = True

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass

    arquivo = io.BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    return arquivo.getvalue()

def gerar_resumo_gerencial_ob_excel(df_ob, meses_ordem):
    """Gera um resumo mensal único, incluindo valores por tipo de despesa."""

    resumo_mensal = (
        df_ob.groupby("Mes_Extenso", observed=False)
        .agg(**{"Qtd. Docs": ("Valor_Limpo", "count"), "Total Pago": ("Valor_Limpo", "sum")})
        .reindex(meses_ordem, fill_value=0)
        .rename_axis("Mês de Referência")
        .reset_index()
    )

    valores_por_tipo = df_ob.pivot_table(
        index="Mes_Extenso",
        columns="Despesa_Tratada",
        values="Valor_Limpo",
        aggfunc="sum",
        fill_value=0.0,
        observed=False,
    ).reindex(meses_ordem, fill_value=0.0)

    tipos_colunas = [
        ("CORRENTE", "Corrente"),
        ("RP", "Restos a Pagar (RP)"),
        ("DEA", "Exercícios Anteriores (DEA)"),
    ]
    for tipo_origem, titulo_coluna in tipos_colunas:
        resumo_mensal[titulo_coluna] = (
            valores_por_tipo[tipo_origem].to_numpy()
            if tipo_origem in valores_por_tipo.columns
            else 0.0
        )

    arquivo = io.BytesIO()
    with pd.ExcelWriter(arquivo, engine="xlsxwriter") as writer:
        resumo_mensal.to_excel(writer, sheet_name="Resumo Mensal", startrow=2, index=False)

        workbook = writer.book
        formato_titulo = workbook.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#002B49",
            "font_size": 14, "align": "center", "valign": "vcenter",
        })
        formato_subtitulo = workbook.add_format({
            "italic": True, "font_color": "#475569", "font_size": 10,
        })
        formato_cabecalho = workbook.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#315B85",
            "border": 0, "align": "center", "valign": "vcenter",
        })
        formato_texto = workbook.add_format({"border": 0})
        formato_qtd = workbook.add_format({"num_format": "#,##0", "border": 0, "align": "center"})
        formato_moeda = workbook.add_format({"num_format": 'R$ #,##0.00', "border": 0, "align": "right"})
        formato_total_texto = workbook.add_format({
            "bold": True, "bg_color": "#F1F5F9", "top": 2, "top_color": "#002B49",
        })
        formato_total_qtd = workbook.add_format({
            "bold": True, "bg_color": "#F1F5F9", "top": 2, "top_color": "#002B49",
            "num_format": "#,##0", "align": "center",
        })
        formato_total_moeda = workbook.add_format({
            "bold": True, "bg_color": "#F1F5F9", "top": 2, "top_color": "#002B49",
            "num_format": 'R$ #,##0.00', "align": "right",
        })

        aba = writer.sheets["Resumo Mensal"]
        ultima_coluna = len(resumo_mensal.columns) - 1
        aba.hide_gridlines(2)
        aba.merge_range(0, 0, 0, ultima_coluna, "Resumo Gerencial por Mês", formato_titulo)
        aba.merge_range(1, 0, 1, ultima_coluna, "Conforme os filtros selecionados no painel.", formato_subtitulo)
        aba.set_row(0, 24)
        aba.set_row(2, 22)
        aba.set_column(0, 0, 20, formato_texto)
        aba.set_column(1, 1, 14, formato_qtd)
        aba.set_column(2, ultima_coluna, 24, formato_moeda)

        for coluna, titulo_coluna in enumerate(resumo_mensal.columns):
            aba.write(2, coluna, titulo_coluna, formato_cabecalho)

        linha_total = len(resumo_mensal) + 3
        aba.write(linha_total, 0, "TOTAL GERAL", formato_total_texto)
        aba.write_formula(linha_total, 1, f"=SUM(B4:B{linha_total})", formato_total_qtd)
        for coluna in range(2, ultima_coluna + 1):
            letra_coluna = chr(65 + coluna)
            aba.write_formula(
                linha_total,
                coluna,
                f"=SUM({letra_coluna}4:{letra_coluna}{linha_total})",
                formato_total_moeda,
            )
        aba.autofilter(2, 0, len(resumo_mensal) + 2, ultima_coluna)
        aba.freeze_panes(3, 0)

    arquivo.seek(0)
    return arquivo.getvalue()


def converter_valor_monetario(serie):
    """Converte valores do CSV e do padrão brasileiro sem inflar decimais."""
    def converter(valor):
        if pd.isna(valor):
            return 0.0
        if isinstance(valor, (int, float, np.number)):
            return float(valor)

        texto = str(valor).strip().replace("R$", "").replace(" ", "")
        if not texto or texto == "-":
            return 0.0

        negativo = texto.startswith("-") or (
            texto.startswith("(") and texto.endswith(")")
        )
        texto = texto.replace("-", "").replace("(", "").replace(")", "")
        ultimo_ponto = texto.rfind(".")
        ultima_virgula = texto.rfind(",")

        if ultimo_ponto >= 0 and ultima_virgula >= 0:
            # O último separador é o decimal; o outro é de milhar.
            if ultima_virgula > ultimo_ponto:
                texto = texto.replace(".", "").replace(",", ".")
            else:
                texto = texto.replace(",", "")
        elif ultima_virgula >= 0:
            texto = texto.replace(",", ".")
        # Com apenas ponto, trata-se do decimal do CSV publicado pelo Sheets.

        numero = pd.to_numeric(texto, errors="coerce")
        if pd.isna(numero):
            return 0.0
        return -float(numero) if negativo else float(numero)

    return serie.map(converter).astype(float)


def ler_csv_url(url):
    req = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    )
    with urllib.request.urlopen(req) as response:
        csv_data = response.read().decode("utf-8")
    return pd.read_csv(io.StringIO(csv_data))


# -------------------------------------------------------------------------
# DADOS DA TELA DE PLANEJAMENTO
# A tela é somente analítica: usa a mesma origem da NL e nunca grava nela.
# -------------------------------------------------------------------------
LINK_NL_DOCUMENTOS = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTWnQkc7oF-YdXKoVTiYeUPYDHGzaeQaiEGqX6fNmB29mkzcd1kAvZVMujFDf02y7j1X8UJzqglAzTL/pub?gid=1892412645&single=true&output=csv"
LINK_NL_RELACOES = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTDHDK6dPnS9favMwwkNSLYZ5i9yjQrJjCEGxpifdQfD9_8dAsHSVlc8TECEyKJi0hWy3Wi5gEM_tI0/pub?gid=1866981074&single=true&output=csv"

# Planilha exclusiva da priorização. A base de NL permanece somente para leitura.
URL_API_PRIORIZACAO = "https://script.google.com/macros/s/AKfycbxJn-8aq4QgZTYj7QuJg8qgAkZOt3_88hq68UG832QUj-o22wKbmbNEGm5WsxocU9NQ/exec"


def _valor_json_seguro(valor):
    """Converte tipos do pandas/numpy para um valor aceito pelo Apps Script."""
    if valor is None or (isinstance(valor, float) and not np.isfinite(valor)):
        return None
    if isinstance(valor, (pd.Timestamp, datetime.datetime, datetime.date)):
        return valor.isoformat()
    if isinstance(valor, np.generic):
        return valor.item()
    try:
        return None if pd.isna(valor) else valor
    except (TypeError, ValueError):
        return valor


def _registros_json(df):
    if df is None or df.empty:
        return []
    return [
        {str(chave): _valor_json_seguro(valor) for chave, valor in linha.items()}
        for linha in df.to_dict(orient="records")
    ]


def carregar_priorizacao_google():
    """Lê pactuados e programação já salvos no Apps Script da Fonte 500."""
    try:
        # Parâmetro variável evita que uma resposta antiga do endpoint fique
        # reaproveitada por cache intermediário depois de edição manual na planilha.
        url_leitura = f"{URL_API_PRIORIZACAO}?_ts={int(datetime.datetime.now().timestamp() * 1000)}"
        req = urllib.request.Request(
            url_leitura,
            headers={"User-Agent": "Mozilla/5.0", "Cache-Control": "no-cache"},
        )
        # A tela não deve ficar bloqueada por uma inicialização lenta do Apps
        # Script. Em caso de falha, o usuário ainda monta o cenário localmente.
        with urllib.request.urlopen(req, timeout=20) as resposta:
            conteudo = json.loads(resposta.read().decode("utf-8"))
        if not isinstance(conteudo, dict) or conteudo.get("ok") is False:
            raise ValueError("A planilha retornou uma resposta inválida.")
        return conteudo
    except Exception as erro:
        return {"ok": False, "erro": str(erro), "pactuados": [], "programacao": []}


def _post_priorizacao_google(carga, timeout=90):
    """Envia uma alteração pontual à planilha de priorização."""
    dados = json.dumps(carga, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        URL_API_PRIORIZACAO,
        data=dados,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "User-Agent": "Mozilla/5.0",
            "Cache-Control": "no-cache",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resposta:
        retorno = json.loads(resposta.read().decode("utf-8"))
    if not isinstance(retorno, dict) or retorno.get("ok") is False:
        raise ValueError(str(retorno.get("erro", "A planilha não confirmou a gravação.")))
    return retorno


def salvar_pactuado_google(pactuado):
    """Atualiza somente o pactuado do mês, sem regravar a Programação."""
    registros = _registros_json(pactuado)
    if not registros:
        return {"ok": True}
    return _post_priorizacao_google({
        "acao": "salvar_pactuado",
        "pactuado": registros[0],
    }, timeout=45)


def salvar_mes_priorizacao_google(mes, fonte, pactuado, programacao_mes):
    """Substitui somente o recorte do mês/fonte em edição."""
    registros_pactuado = _registros_json(pactuado)
    return _post_priorizacao_google({
        "acao": "salvar_mes",
        "mes": str(mes),
        "fonte": str(fonte),
        "pactuado": registros_pactuado[0] if registros_pactuado else {},
        "programacao": _registros_json(programacao_mes),
    }, timeout=90)




def _normalizar_nome_campo_priorizacao(nome):
    """Normaliza cabeçalhos vindos do Apps Script/Google Sheets."""
    texto = "" if nome is None else str(nome).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^A-Za-z0-9]+", "_", texto).strip("_").upper()
    return texto


def _achar_coluna_priorizacao(df, *aliases):
    """Localiza coluna mesmo com acento, espaço, hífen ou pequenas variações."""
    if df is None or df.empty:
        return None
    mapa = {_normalizar_nome_campo_priorizacao(c): c for c in df.columns}
    for alias in aliases:
        chave = _normalizar_nome_campo_priorizacao(alias)
        if chave in mapa:
            return mapa[chave]
    # fallback: permite prefixos como VALOR_PROGRAMADO / VALOR_PROGRAMA...
    for alias in aliases:
        chave = _normalizar_nome_campo_priorizacao(alias)
        for normalizado, original in mapa.items():
            if normalizado.startswith(chave) or chave.startswith(normalizado):
                return original
    return None


def canonicalizar_programacao_priorizacao(registros):
    """Converte a resposta da aba Programação para um esquema fixo do painel."""
    df = pd.DataFrame(registros or [])
    colunas_saida = [
        "Mes", "Fonte", "Grupo", "Tipo_NL", "Semana", "Objeto",
        "Numero_NL", "Credor", "Valor_Programado", "Observacao", "Atualizado_Em",
    ]
    if df.empty:
        return pd.DataFrame(columns=colunas_saida)

    aliases = {
        "Mes": ("Mes", "Mês", "Mês de programação", "Mes de programacao"),
        "Fonte": ("Fonte",),
        "Grupo": ("Grupo",),
        "Tipo_NL": ("Tipo_NL", "Tipo de NL", "Tipo NL"),
        "Semana": ("Semana",),
        "Objeto": ("Objeto", "Objeto da Despesa"),
        "Numero_NL": ("Numero_NL", "Número NL", "Numero NL", "NL"),
        "Credor": ("Credor", "Nome do Credor"),
        "Valor_Programado": ("Valor_Programado", "Valor Programado", "Valor_Programa", "Valor"),
        "Observacao": ("Observacao", "Observação"),
        "Atualizado_Em": ("Atualizado_Em", "Atualizado Em"),
    }
    saida = pd.DataFrame(index=df.index)
    for destino, nomes in aliases.items():
        origem = _achar_coluna_priorizacao(df, *nomes)
        if origem is None:
            saida[destino] = 0.0 if destino == "Valor_Programado" else ""
        else:
            saida[destino] = df[origem]

    saida["Mes"] = saida["Mes"].apply(normalizar_mes_priorizacao)
    saida["Fonte"] = saida["Fonte"].fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    saida["Grupo"] = saida["Grupo"].fillna("").astype(str).str.strip()
    saida["Tipo_NL"] = saida["Tipo_NL"].fillna("").astype(str).str.strip()
    saida["Semana"] = saida["Semana"].apply(normalizar_semana_priorizacao)
    saida["Objeto"] = saida["Objeto"].fillna("").astype(str).str.strip()
    saida["Numero_NL"] = saida["Numero_NL"].apply(normalizar_numero_nl_priorizacao)
    saida["Credor"] = saida["Credor"].fillna("").astype(str).str.strip()
    saida["Valor_Programado"] = converter_valor_monetario(saida["Valor_Programado"])
    saida["Observacao"] = saida["Observacao"].fillna("").astype(str)
    return saida[colunas_saida]


def canonicalizar_pactuados_priorizacao(registros):
    """Converte Pactuados para o esquema mensal + distribuição semanal."""
    df = pd.DataFrame(registros or [])
    colunas_saida = [
        "Mes", "Fonte", "Grupo", "Valor_Pactuado",
        "Pactuado_Semana_1", "Pactuado_Semana_2",
        "Pactuado_Semana_3", "Pactuado_Semana_4", "Atualizado_Em",
    ]
    if df.empty:
        return pd.DataFrame(columns=colunas_saida)
    aliases = {
        "Mes": ("Mes", "Mês"),
        "Fonte": ("Fonte",),
        "Grupo": ("Grupo",),
        "Valor_Pactuado": ("Valor_Pactuado", "Valor Pactuado", "Pactuado Mensal", "Valor"),
        "Pactuado_Semana_1": ("Pactuado_Semana_1", "Pactuado Semana 1", "Semana 1"),
        "Pactuado_Semana_2": ("Pactuado_Semana_2", "Pactuado Semana 2", "Semana 2"),
        "Pactuado_Semana_3": ("Pactuado_Semana_3", "Pactuado Semana 3", "Semana 3"),
        "Pactuado_Semana_4": ("Pactuado_Semana_4", "Pactuado Semana 4", "Semana 4"),
        "Atualizado_Em": ("Atualizado_Em", "Atualizado Em"),
    }
    saida = pd.DataFrame(index=df.index)
    colunas_valor = {
        "Valor_Pactuado", "Pactuado_Semana_1", "Pactuado_Semana_2",
        "Pactuado_Semana_3", "Pactuado_Semana_4",
    }
    for destino, nomes in aliases.items():
        origem = _achar_coluna_priorizacao(df, *nomes)
        if origem is None:
            saida[destino] = 0.0 if destino in colunas_valor else ""
        else:
            saida[destino] = df[origem]
    saida["Mes"] = saida["Mes"].apply(normalizar_mes_priorizacao)
    saida["Fonte"] = saida["Fonte"].fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    saida["Grupo"] = saida["Grupo"].fillna("").astype(str).str.strip().str.upper()
    for coluna in colunas_valor:
        saida[coluna] = converter_valor_monetario(saida[coluna])
    return saida[colunas_saida]


def normalizar_semana_priorizacao(valor):
    """Aceita Semana 1/01, 1/01 e variações equivalentes da planilha."""
    texto = "" if pd.isna(valor) else str(valor).strip().upper()
    if not texto:
        return ""
    numeros = re.findall(r"\d+", texto)
    if numeros:
        numero = int(numeros[-1])
        if 1 <= numero <= 4:
            return f"Semana {numero}"
    return str(valor).strip()


def normalizar_mes_priorizacao(valor):
    """Normaliza mês da API/planilha para MM/AAAA sem depender do formato digitado.

    O Google Sheets pode exibir 08/2026, mas o Apps Script devolver a célula
    como uma data ISO (ex.: 2026-08-01T03:00:00.000Z). Por isso tratamos
    explicitamente os dois formatos.
    """
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    if isinstance(valor, (pd.Timestamp, datetime.datetime, datetime.date)):
        return f"{valor.month:02d}/{valor.year:04d}"

    texto = str(valor).strip()
    if not texto:
        return ""

    # Formato visual da planilha: 08/2026, 8/2026, 08-2026 etc.
    encontrado = re.search(r"(?<!\d)(\d{1,2})[\-/](\d{4})(?!\d)", texto)
    if encontrado:
        mes, ano = int(encontrado.group(1)), int(encontrado.group(2))
        if 1 <= mes <= 12:
            return f"{mes:02d}/{ano:04d}"

    # Formato ISO devolvido pelo Apps Script/JSON: 2026-08-01T03:00:00.000Z
    encontrado_iso = re.search(r"(?<!\d)(\d{4})-(\d{1,2})-(\d{1,2})(?:[T\s].*)?$", texto)
    if encontrado_iso:
        ano, mes = int(encontrado_iso.group(1)), int(encontrado_iso.group(2))
        if 1 <= mes <= 12:
            return f"{mes:02d}/{ano:04d}"

    # Também tenta interpretar outras strings de data válidas vindas do Sheets.
    # dayfirst=False evita transformar 2026-08-01 em janeiro/agosto por engano.
    try:
        data_convertida = pd.to_datetime(texto, errors="coerce", dayfirst=False)
        if pd.notna(data_convertida):
            return f"{data_convertida.month:02d}/{data_convertida.year:04d}"
    except Exception:
        pass

    # Também aceita rótulos usados manualmente no Google Sheets, como
    # Ago/2026, Agosto/2026, Set/2026 e Setembro/2026.
    mapa_meses_pt = {
        "JAN": 1, "JANEIRO": 1, "FEV": 2, "FEVEREIRO": 2,
        "MAR": 3, "MARCO": 3, "MARÇO": 3, "ABR": 4, "ABRIL": 4,
        "MAI": 5, "MAIO": 5, "JUN": 6, "JUNHO": 6,
        "JUL": 7, "JULHO": 7, "AGO": 8, "AGOSTO": 8,
        "SET": 9, "SETEMBRO": 9, "OUT": 10, "OUTUBRO": 10,
        "NOV": 11, "NOVEMBRO": 11, "DEZ": 12, "DEZEMBRO": 12,
    }
    encontrado_nome = re.search(r"([A-Za-zÀ-ÿ]+)\s*[\-/ ]\s*(\d{4})", texto)
    if encontrado_nome:
        nome_mes = encontrado_nome.group(1).strip().upper().replace(".", "")
        ano = int(encontrado_nome.group(2))
        if nome_mes in mapa_meses_pt:
            return f"{mapa_meses_pt[nome_mes]:02d}/{ano:04d}"
    try:
        data = pd.to_datetime(texto, errors="raise", dayfirst=True)
        return data.strftime("%m/%Y")
    except Exception:
        return texto


def normalizar_numero_nl_priorizacao(valor):
    texto = "" if pd.isna(valor) else str(valor).strip().upper()
    return re.sub(r"\.0$", "", texto)


def assinatura_priorizacao_remota(conteudo):
    """Assinatura estável para detectar edição feita diretamente na planilha."""
    if not isinstance(conteudo, dict):
        return ""
    partes = {}
    for nome in ("pactuados", "programacao"):
        registros = conteudo.get(nome, []) or []
        limpos = []
        for registro in registros:
            if not isinstance(registro, dict):
                continue
            limpo = {
                str(chave): _valor_json_seguro(valor)
                for chave, valor in registro.items()
                if str(chave).strip().lower() not in {"atualizado_em", "atualizado em"}
            }
            limpos.append(limpo)
        limpos.sort(key=lambda item: json.dumps(item, ensure_ascii=False, sort_keys=True, default=str))
        partes[nome] = limpos
    return json.dumps(partes, ensure_ascii=False, sort_keys=True, default=str)


@st.cache_data(ttl=20, show_spinner=False)
def carregar_priorizacao_google_sincronizada(nonce=0):
    """Relê periodicamente a planilha; nonce permite atualização manual imediata."""
    return carregar_priorizacao_google()


def classificar_grupo_planejamento(valor):
    texto = "" if pd.isna(valor) else str(valor).strip().upper()
    compacto = re.sub(r"[^A-Z0-9]", "", texto)
    if re.search(r"\b(GD|GND)1\b", texto) or compacto in {"1", "GD1", "GND1"} or compacto.startswith("1PESSOAL"):
        return "GD1"
    if re.search(r"\b(GD|GND)3\b", texto) or compacto in {"3", "GD3", "GND3"} or compacto.startswith("3OUTRAS"):
        return "GD3"
    if re.search(r"\b(GD|GND)4\b", texto) or compacto in {"4", "GD4", "GND4"} or compacto.startswith("4INVESTIMENTOS"):
        return "GD4"
    return "NÃO INFORMADO"


@st.cache_data(ttl=300)
def carregar_dados_planejamento_nl():
    """Versão enxuta da integração da NL, exclusiva para análise de metas."""
    try:
        docs = ler_csv_url(LINK_NL_DOCUMENTOS)
        relacoes = ler_csv_url(LINK_NL_RELACOES)
    except Exception:
        return pd.DataFrame()
    if docs.empty:
        return pd.DataFrame()

    docs.columns = [str(c).strip() for c in docs.columns]
    relacoes.columns = [str(c).strip() for c in relacoes.columns]
    col_ne_docs = next((c for c in docs.columns if any(x in c.upper() for x in ["DOCUMENTONE", "NE", "EMPENHO"])), docs.columns[0])
    col_valor = next((c for c in docs.columns if c.strip().upper() == "VALOR"), docs.columns[-1])
    col_data = next((c for c in docs.columns if "DATA" in c.upper()), None)
    col_grupo_docs = next((c for c in docs.columns if c.strip().upper() == "GRUPO"), None)
    col_fonte_docs = next((c for c in docs.columns if "FONTE" in c.upper()), None)
    col_status_docs = next((c for c in docs.columns if c.strip().upper() == "STATUS"), None)
    col_tipo_docs = next((c for c in docs.columns if "TIPO" in c.upper() and "NL" in c.upper()), None)

    def chave_ne(serie):
        return serie.fillna("").astype(str).str.strip().str.upper().str.replace(r"\.0$", "", regex=True).str.replace(r"[^A-Z0-9]", "", regex=True)

    def moeda(serie):
        def converter(valor):
            texto = str(valor).strip().replace("R$", "").replace(" ", "")
            if "," in texto and "." in texto:
                texto = texto.replace(".", "").replace(",", ".")
            elif "," in texto:
                texto = texto.replace(",", ".")
            return pd.to_numeric(texto, errors="coerce")
        return serie.map(converter).fillna(0.0)

    docs["NE_Chave"] = chave_ne(docs[col_ne_docs])
    docs["Valor_Executado"] = moeda(docs[col_valor])
    docs["Data_DT"] = pd.to_datetime(docs[col_data], errors="coerce") if col_data else pd.NaT
    docs["Mês"] = docs["Data_DT"].dt.strftime("%m/%Y").fillna("Não informado")
    docs["Grupo_Base"] = docs[col_grupo_docs].apply(classificar_grupo_planejamento) if col_grupo_docs else "NÃO INFORMADO"
    docs["Fonte_Base"] = docs[col_fonte_docs].fillna("NÃO INFORMADA").astype(str).str.strip() if col_fonte_docs else "NÃO INFORMADA"
    docs["Status"] = docs[col_status_docs].fillna("Não informado").astype(str).str.strip() if col_status_docs else "Não informado"
    docs["Tipo de NL"] = docs[col_tipo_docs].fillna("Não informado").astype(str).str.strip() if col_tipo_docs else "Não informado"

    col_ne_rel = next((c for c in relacoes.columns if any(x in c.upper() for x in ["DOCUMENTONE", "NE", "EMPENHO"])), relacoes.columns[0])
    col_fonte = next((c for c in relacoes.columns if "FONTE" in c.upper()), None)
    col_objeto = next((c for c in relacoes.columns if "OBJETO" in c.upper()), None)
    col_grupo_rel = next((c for c in relacoes.columns if c.strip().upper() == "GRUPO"), None)
    relacoes["NE_Chave"] = chave_ne(relacoes[col_ne_rel])
    relacoes["Fonte"] = relacoes[col_fonte].fillna("NÃO INFORMADA").astype(str).str.strip() if col_fonte else "NÃO INFORMADA"
    relacoes["Objeto"] = relacoes[col_objeto].fillna("NÃO INFORMADO").astype(str).str.strip() if col_objeto else "NÃO INFORMADO"
    relacoes["Grupo_Relacao"] = relacoes[col_grupo_rel].apply(classificar_grupo_planejamento) if col_grupo_rel else "NÃO INFORMADO"
    relacoes = relacoes[relacoes["NE_Chave"] != ""].drop_duplicates("NE_Chave")

    dados = docs.merge(relacoes[["NE_Chave", "Fonte", "Objeto", "Grupo_Relacao"]], on="NE_Chave", how="left")
    dados["Fonte"] = dados["Fonte"].fillna("").astype(str).str.strip()
    dados.loc[dados["Fonte"].isin(["", "NAN", "NÃO INFORMADA"]), "Fonte"] = dados.loc[
        dados["Fonte"].isin(["", "NAN", "NÃO INFORMADA"]), "Fonte_Base"
    ]
    dados["Fonte"] = dados["Fonte"].replace("", "NÃO INFORMADA")
    dados["Objeto"] = dados["Objeto"].fillna("NÃO INFORMADO")
    dados["Grupo"] = dados["Grupo_Relacao"].fillna("NÃO INFORMADO")
    sem_grupo = dados["Grupo"] == "NÃO INFORMADO"
    dados.loc[sem_grupo, "Grupo"] = dados.loc[sem_grupo, "Grupo_Base"]
    return dados[["Mês", "Data_DT", "Fonte", "Grupo", "Objeto", "Status", "Tipo de NL", "Valor_Executado"]]


@st.cache_data(ttl=300)
def carregar_base_nl_espelho_planejamento():
    """Replica o mesmo relacionamento Fonte/Objeto/Grupo usado no painel NL."""
    try:
        df1 = ler_csv_url(LINK_NL_DOCUMENTOS)
        df2 = ler_csv_url(LINK_NL_RELACOES)
    except Exception:
        return pd.DataFrame()
    if df1.empty:
        return pd.DataFrame()

    df1.columns = [str(c).strip() for c in df1.columns]
    df2.columns = [str(c).strip() for c in df2.columns]

    def chave_ne(serie):
        return (
            serie.fillna("").astype(str).str.strip().str.upper()
            .str.replace(r"\.0$", "", regex=True)
            .str.replace(r"[^A-Z0-9]", "", regex=True)
        )

    def valor_nl(serie):
        texto = serie.fillna("0").astype(str)
        texto = texto.str.replace(r"[R$\s.]", "", regex=True).str.replace(",", ".")
        return pd.to_numeric(texto, errors="coerce").fillna(0.0)

    col_ne1 = next((c for c in df1.columns if any(p in c.upper() for p in ["NE", "EMPENHO", "DOCUMENTONE"])), df1.columns[0])
    col_valor = "Valor" if "Valor" in df1.columns else df1.columns[-1]
    col_data = next((c for c in df1.columns if "data" in c.lower()), None)
    col_credor = next((c for c in df1.columns if "NOME" in c.upper() and "CREDOR" in c.upper()), None)
    if col_credor is None:
        col_credor = next((c for c in df1.columns if "CREDOR" in c.upper()), None)
    col_numero = next((c for c in df1.columns if c.strip().upper() in ["NÚMERO", "NUMERO"]), None)
    df1["NE_Chave"] = chave_ne(df1[col_ne1])
    # Mantém a mesma interpretação de data utilizada no painel de NL.
    df1["Data_DT"] = pd.to_datetime(df1[col_data], errors="coerce") if col_data else pd.NaT
    df1["Competencia"] = df1["Data_DT"].dt.strftime("%m/%Y").fillna("Não informada")
    df1["Grupo_Filtro"] = df1["Grupo"].fillna("Todos").astype(str).str.strip() if "Grupo" in df1.columns else "Todos"
    df1["Status_Filtro"] = df1["Status"].fillna("Não informado").astype(str).str.strip() if "Status" in df1.columns else "Não informado"
    col_tipo = next((c for c in df1.columns if "TIPO" in c.upper() and "NL" in c.upper()), None)
    df1["Tipo_NL_Filtro"] = df1[col_tipo].fillna("Não informado").astype(str).str.strip() if col_tipo else "Não informado"
    df1["Valor_Total_Limpo"] = valor_nl(df1[col_valor])
    df1["Credor_NL"] = df1[col_credor].fillna("Não informado").astype(str).str.strip() if col_credor else "Não informado"
    df1["Numero_NL"] = df1[col_numero].fillna("").astype(str).str.strip() if col_numero else ""

    col_ne2 = "DocumentoNE" if "DocumentoNE" in df2.columns else next((c for c in df2.columns if "NE" in c.upper()), df2.columns[0])
    col_fonte2 = "Fonte" if "Fonte" in df2.columns else next((c for c in df2.columns if "FONTE" in c.upper()), None)
    col_objeto2 = "Objeto da Despesa" if "Objeto da Despesa" in df2.columns else next((c for c in df2.columns if "OBJETO" in c.upper()), None)
    col_grupo2 = "Grupo" if "Grupo" in df2.columns else next((c for c in df2.columns if "GRUPO" in c.upper()), None)
    df2["NE_Chave"] = chave_ne(df2[col_ne2])
    df2["Fonte_Relacao"] = df2[col_fonte2].fillna("NÃO INFORMADA").astype(str).str.strip() if col_fonte2 else "NÃO INFORMADA"
    df2["Objeto_Relacao"] = df2[col_objeto2].fillna("NÃO INFORMADO").astype(str).str.strip() if col_objeto2 else "NÃO INFORMADO"
    df2["Grupo_Relacao"] = df2[col_grupo2].fillna("NÃO INFORMADO").astype(str).str.strip() if col_grupo2 else "NÃO INFORMADO"
    df2 = df2[df2["NE_Chave"] != ""].drop_duplicates(subset=["NE_Chave"])

    base = pd.merge(
        df1,
        df2[["NE_Chave", "Fonte_Relacao", "Objeto_Relacao", "Grupo_Relacao"]],
        on="NE_Chave",
        how="left",
    )
    base["Fonte_Relacao"] = base["Fonte_Relacao"].fillna("NÃO INFORMADA")
    base["Objeto_Relacao"] = base["Objeto_Relacao"].fillna("NÃO INFORMADO")
    base["Grupo_Relacao"] = base["Grupo_Relacao"].fillna("NÃO INFORMADO")
    base["Grupo_Classificado"] = base["Grupo_Relacao"].apply(classificar_grupo_planejamento)
    sem_grupo = base["Grupo_Classificado"] == "NÃO INFORMADO"
    base.loc[sem_grupo, "Grupo_Classificado"] = base.loc[sem_grupo, "Grupo_Filtro"].apply(classificar_grupo_planejamento)
    return base


@st.cache_data(ttl=300)
def carregar_historico_ob_fonte_500():
    """Histórico pago da mesma BASE consolidada exibida no painel de OB."""
    link_base_ob = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTD3b7L6byArEDgkVKOXXlc7RK0M2QKXLov83OydCaks3rDISWYWfgGNi6vG6pwy8t5Ul3Fd2wArhtT/pub?gid=1786485134&single=true&output=csv"
    try:
        df = ler_csv_url(link_base_ob)
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return pd.DataFrame()

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df.columns = [str(c).strip() for c in df.columns]
    for coluna in ["Número", "NE", "Data Emissão", "Valor", "Fonte", "Nome do Credor", "Tipo de OB", "GRUPO", "Despesa"]:
        if coluna not in df.columns:
            df[coluna] = None
    df = df.dropna(subset=["Valor"]).copy()

    # Mesma chave de segurança usada no painel de Pagamentos (OB).
    chaves = []
    for coluna in ["Número", "NE", "Data Emissão", "Valor", "Fonte", "Nome do Credor", "Tipo de OB"]:
        chave = f"__historico_{coluna}"
        if coluna == "Valor":
            df[chave] = converter_valor_monetario(df[coluna]).round(2)
        elif coluna == "Data Emissão":
            data = pd.to_datetime(df[coluna], errors="coerce", dayfirst=True)
            df[chave] = data.dt.strftime("%Y-%m-%d").fillna(df[coluna].fillna("").astype(str).str.strip().str.upper())
        else:
            df[chave] = df[coluna].fillna("").astype(str).str.strip().str.upper().str.replace(r"\s+", " ", regex=True)
        chaves.append(chave)
    df = df.drop_duplicates(subset=chaves, keep="first").copy()

    df["Valor_Pago"] = converter_valor_monetario(df["Valor"])
    datas = pd.to_datetime(df["Data Emissão"], errors="coerce", dayfirst=True)
    mapa_meses = {
        1: "Jan/2026", 2: "Fev/2026", 3: "Mar/2026", 4: "Abr/2026",
        5: "Mai/2026", 6: "Jun/2026", 7: "Jul/2026", 8: "Ago/2026",
        9: "Set/2026", 10: "Out/2026", 11: "Nov/2026", 12: "Dez/2026",
    }
    df["Mês"] = datas.dt.month.map(mapa_meses).fillna("Não identificado")
    grupo_texto = df["GRUPO"].fillna("").astype(str).str.upper()
    df["Grupo"] = np.where(
        grupo_texto.str.contains("INVEST|4", regex=True),
        "4 - INVESTIMENTOS",
        "3 - OUTRAS DESPESAS CORRENTES",
    )
    despesa_texto = df["Despesa"].fillna("").astype(str).str.upper()
    df["Tipo de Despesa"] = np.select(
        [
            despesa_texto.str.contains("DEA|EXERC|ANTERIOR|RECONHECIMENTO", regex=True),
            despesa_texto.str.contains("RP|RESTO|PAGAR", regex=True),
        ],
        ["DEA", "RP"],
        default="CORRENTE",
    )
    fonte_500 = df["Fonte"].fillna("").astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False)
    return df.loc[fonte_500, ["Número", "Mês", "Grupo", "Tipo de Despesa", "Valor_Pago"]]


# -------------------------------------------------------------------------
# CONEXÃO EXCLUSIVA DA TELA RELATÓRIO 009717
# -------------------------------------------------------------------------
URL_API_RELATORIO_009717_PADRAO = "https://script.google.com/macros/s/AKfycbywfyRrszPy3wqbSsrLsFBgd5rTw3d4tNKl3zBmJBknhjAv2bI0qAvzZv3Tk35KkTwI/exec"

# Base exclusiva do Programa de Desembolso. A origem deve ser a aba
# PD_ABA_TRATADA, publicada como CSV para que o Streamlit Cloud possa lê-la.
LINK_PD_ABA_TRATADA = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRsMrqzxYHgTRv_tBJnDU_Rg1OpFmh_FCCo55w671Kna-IE8FIPD4rhL7O-bDwCsNMQW4Qj7UZGaBFP/"
    "pub?gid=27472472&single=true&output=csv"
)


def _normalizar_nome_pd(nome):
    """Normaliza cabeçalhos da base de PD sem depender de acentos ou espaços."""
    texto = unicodedata.normalize("NFKD", str(nome or ""))
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return re.sub(r"[^a-z0-9]+", "", texto.lower())


@st.cache_data(ttl=300, show_spinner=False)
def carregar_dados_pd():
    """Lê a PD_ABA_TRATADA e replica a relação por NE usada na tela de NL."""
    try:
        bruto = ler_csv_url(LINK_PD_ABA_TRATADA)
        # Mesma base de relacionamento utilizada pela Liquidação (NL),
        # correspondente à Base_de_dados_QLIKVIEW publicada como CSV.
        relacoes = ler_csv_url(LINK_NL_RELACOES)
    except Exception as erro:
        return pd.DataFrame(), str(erro)

    if bruto.empty:
        return pd.DataFrame(), "A aba PD_ABA_TRATADA não possui registros para exibir."
    if relacoes.empty:
        return pd.DataFrame(), "A Base_de_dados_QLIKVIEW não possui registros para relacionar os PDs."

    bruto = bruto.loc[:, ~bruto.columns.duplicated()].copy()
    relacoes = relacoes.loc[:, ~relacoes.columns.duplicated()].copy()
    bruto.columns = [str(coluna).strip() for coluna in bruto.columns]
    relacoes.columns = [str(coluna).strip() for coluna in relacoes.columns]

    colunas_normalizadas = {
        _normalizar_nome_pd(coluna): coluna for coluna in bruto.columns
    }
    colunas_rel_normalizadas = {
        _normalizar_nome_pd(coluna): coluna for coluna in relacoes.columns
    }

    def localizar_coluna(*aliases):
        # Correspondência exata após normalização para não confundir, por exemplo,
        # "Objeto da Despesa" com a coluna genérica "Despesa".
        for alias in aliases:
            chave = _normalizar_nome_pd(alias)
            if chave in colunas_normalizadas:
                return colunas_normalizadas[chave]
        return None

    def localizar_coluna_relacao(*aliases):
        for alias in aliases:
            chave = _normalizar_nome_pd(alias)
            if chave in colunas_rel_normalizadas:
                return colunas_rel_normalizadas[chave]
        return None

    def serie_texto(*aliases, padrao="NÃO INFORMADO"):
        coluna = localizar_coluna(*aliases)
        if coluna is None:
            return pd.Series(padrao, index=bruto.index, dtype="object")
        return bruto[coluna].fillna("").astype(str).str.strip().replace("", padrao)

    def chave_ne(serie):
        # É exatamente a mesma regra aplicada na tela de NL: remove máscara,
        # pontuação, espaços e o sufixo .0 antes de fazer o relacionamento.
        return (
            serie.fillna("").astype(str).str.strip().str.upper()
            .str.replace(r"\.0$", "", regex=True)
            .str.replace(r"[^A-Z0-9]", "", regex=True)
        )

    def classificar_gd(valor):
        texto = str(valor or "").strip().upper()
        compacto = re.sub(r"[^A-Z0-9]", "", texto)
        if (
            re.search(r"\bGD1\b", texto)
            or re.search(r"\bGND1\b", texto)
            or compacto.startswith("GD1")
            or compacto.startswith("GND1")
            or compacto == "1"
            or compacto.startswith("1PESSOAL")
        ):
            return "GD1"
        if (
            re.search(r"\bGD3\b", texto)
            or re.search(r"\bGND3\b", texto)
            or compacto.startswith("GD3")
            or compacto.startswith("GND3")
            or compacto == "3"
            or compacto.startswith("3OUTRAS")
        ):
            return "GD3"
        if (
            re.search(r"\bGD4\b", texto)
            or re.search(r"\bGND4\b", texto)
            or compacto.startswith("GD4")
            or compacto.startswith("GND4")
            or compacto == "4"
            or compacto.startswith("4INVESTIMENTOS")
        ):
            return "GD4"
        return "NÃO INFORMADO"

    coluna_numero = localizar_coluna("Número", "Numero", "Número PD", "Numero PD", "PD")
    coluna_valor = localizar_coluna("Valor", "Valor Total", "Valor PD", "Valor Programado")
    coluna_ne_pd = localizar_coluna(
        "DocumentoNE", "Documento NE", "NE", "Número NE", "Numero NE",
        "Empenho", "Nota de Empenho", "Documento Empenho"
    )

    if coluna_numero is None or coluna_valor is None or coluna_ne_pd is None:
        faltantes = []
        if coluna_numero is None:
            faltantes.append("Número do PD")
        if coluna_valor is None:
            faltantes.append("Valor")
        if coluna_ne_pd is None:
            faltantes.append("identificador NE")
        return pd.DataFrame(), (
            "A aba PD_ABA_TRATADA precisa conter: " + ", ".join(faltantes)
            + ". Cabeçalhos encontrados: " + ", ".join(bruto.columns)
        )

    # ---------------- RELAÇÃO COM Base_de_dados_QLIKVIEW ----------------
    col_ne_rel = localizar_coluna_relacao(
        "DocumentoNE", "Documento NE", "NE", "Número NE", "Numero NE",
        "Empenho", "Nota de Empenho"
    )
    col_fonte_rel = localizar_coluna_relacao("Fonte", "Fonte de Recurso")
    col_objeto_rel = localizar_coluna_relacao(
        "Objeto da Despesa", "Objeto de Despesa", "Objeto Despesa", "Objeto"
    )
    col_grupo_rel = localizar_coluna_relacao("Grupo", "GD", "GND", "Grupo de Despesa")

    faltantes_relacao = []
    if col_ne_rel is None:
        faltantes_relacao.append("DocumentoNE/NE")
    if col_objeto_rel is None:
        faltantes_relacao.append("Objeto da Despesa")
    if col_grupo_rel is None:
        faltantes_relacao.append("Grupo")
    if faltantes_relacao:
        return pd.DataFrame(), (
            "A Base_de_dados_QLIKVIEW não possui as colunas necessárias para a relação: "
            + ", ".join(faltantes_relacao)
            + ". Cabeçalhos encontrados: " + ", ".join(relacoes.columns)
        )

    bruto["NE_Chave"] = chave_ne(bruto[coluna_ne_pd])
    relacoes["NE_Chave"] = chave_ne(relacoes[col_ne_rel])
    relacoes["Fonte_Relacao"] = (
        relacoes[col_fonte_rel].fillna("NÃO INFORMADA").astype(str).str.strip()
        if col_fonte_rel else "NÃO INFORMADA"
    )
    relacoes["Objeto_Relacao"] = (
        relacoes[col_objeto_rel].fillna("NÃO INFORMADO").astype(str).str.strip()
    )
    relacoes["Grupo_Relacao"] = (
        relacoes[col_grupo_rel].fillna("NÃO INFORMADO").astype(str).str.strip()
    )

    # Mesmo comportamento da tela de NL: uma linha de relacionamento por NE.
    relacoes_dedup = relacoes[relacoes["NE_Chave"] != ""].drop_duplicates(
        subset=["NE_Chave"], keep="first"
    )

    bruto = bruto.merge(
        relacoes_dedup[["NE_Chave", "Fonte_Relacao", "Objeto_Relacao", "Grupo_Relacao"]],
        on="NE_Chave",
        how="left",
    )

    natureza = serie_texto("Natureza", "Natureza da Despesa", padrao="")
    grupo_informado = serie_texto("GD", "Grupo", "GND", padrao="")
    grupo_derivado = natureza.astype(str).str.extract(r"^\s*([134])", expand=False)
    grupo_fallback = grupo_informado.where(grupo_informado.astype(str).str.strip() != "", "")
    grupo_fallback = grupo_fallback.replace({"NÃO INFORMADO": ""})
    grupo_fallback = grupo_fallback.where(
        grupo_fallback != "", "GD" + grupo_derivado.fillna("NÃO INFORMADO")
    )
    grupo_fallback = grupo_fallback.astype(str).str.upper().replace(
        {"3": "GD3", "4": "GD4", "1": "GD1"}
    )

    grupo_relacionado = bruto["Grupo_Relacao"].fillna("NÃO INFORMADO").apply(classificar_gd)
    grupo = grupo_relacionado.where(grupo_relacionado != "NÃO INFORMADO", grupo_fallback)

    # Na tela de NL, Objeto e Fonte são determinados pela Base_de_dados_QLIKVIEW
    # após a combinação pela NE. Se a NE não tiver correspondência, o correto é
    # sinalizar como não informado em vez de aproveitar uma coluna inadequada da PD.
    objeto_final = (
        bruto["Objeto_Relacao"].fillna("NÃO INFORMADO").astype(str).str.strip()
        .replace({"": "NÃO INFORMADO", "NAN": "NÃO INFORMADO", "NAO INFORMADO": "NÃO INFORMADO"})
    )
    fonte_final = (
        bruto["Fonte_Relacao"].fillna("NÃO INFORMADA").astype(str).str.strip()
        .replace({"": "NÃO INFORMADA", "NAN": "NÃO INFORMADA", "NAO INFORMADA": "NÃO INFORMADA"})
    )

    dados = pd.DataFrame(
        {
            "Número PD": bruto[coluna_numero].fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True),
            "NE_Chave": bruto["NE_Chave"],
            "Data Emissão": pd.to_datetime(
                serie_texto("Data Emissão", "Data Emissao", "Data", padrao=""),
                errors="coerce",
                dayfirst=True,
            ),
            "UG Pagadora": serie_texto("UG Pagadora", "UG_Pagadora", "UG"),
            # Fonte, Objeto e GD passam a vir prioritariamente da mesma relação
            # por NE utilizada pela tela de NL.
            "Fonte": fonte_final,
            "GD": grupo,
            "Despesa": natureza.replace("", "NÃO INFORMADA"),
            "Objeto da Despesa": objeto_final,
            "Nome do Credor": serie_texto("Nome do Credor", "Credor", "Entidade / Credor", "Entidade"),
            "Tipo de PD": serie_texto("Tipo de PD", "Tipo PD", "Tipo_PD", "Tipo de OB", "OB", "Tipo OB"),
            "Tipo de OB": serie_texto("Tipo de PD", "Tipo PD", "Tipo_PD", "Tipo de OB", "OB", "Tipo OB"),
            "Status": serie_texto("Status"),
            "Valor": converter_valor_monetario(bruto[coluna_valor]),
        }
    )
    dados = dados[dados["Número PD"].ne("")].copy()
    dados["UG Pagadora"] = dados["UG Pagadora"].str.replace(r"\.0$", "", regex=True)
    dados["Fonte"] = dados["Fonte"].astype(str).str.replace(r"\.0$", "", regex=True)
    dados["Valor"] = dados["Valor"].fillna(0.0)
    return dados, ""


def chamar_api_relatorio_009717(url_api, acao="status", timeout=120):
    """Conversa via POST com o Web App usado exclusivamente pelo Relatório 009717.

    O parâmetro _ts impede que proxies/CDNs reutilizem uma resposta anterior.
    """
    url_api = (url_api or "").strip()
    if not url_api:
        raise ValueError("Informe a URL do Web App do Apps Script.")

    separador = "&" if "?" in url_api else "?"
    url_chamada = (
        f"{url_api}{separador}_ts="
        f"{int(datetime.datetime.now().timestamp() * 1000)}"
    )

    payload = json.dumps(
        {
            "acao": str(acao),
            "_ts": int(datetime.datetime.now().timestamp() * 1000),
        },
        ensure_ascii=False,
    ).encode("utf-8")

    req = urllib.request.Request(
        url_chamada,
        data=payload,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "User-Agent": "Mozilla/5.0",
            "Cache-Control": "no-cache, no-store, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
        method="POST",
    )

    with urllib.request.urlopen(req, timeout=timeout) as resposta:
        corpo = resposta.read().decode("utf-8", errors="replace").strip()

    if not corpo:
        raise ValueError("O Apps Script retornou uma resposta vazia.")

    try:
        conteudo = json.loads(corpo)
    except json.JSONDecodeError as erro_json:
        inicio_resposta = re.sub(r"\s+", " ", corpo[:220]).strip()
        raise ValueError(
            "O Apps Script não retornou JSON. "
            f"Início da resposta recebida: {inicio_resposta!r}"
        ) from erro_json

    if not isinstance(conteudo, dict):
        raise ValueError("O Apps Script retornou uma resposta inválida.")

    if conteudo.get("ok") is False:
        raise ValueError(str(conteudo.get("erro", "Falha ao acessar o relatório.")))

    return conteudo


# -------------------------------------------------------------------------
# NAVEGAÇÃO PRINCIPAL ENTRE AS TELAS
# -------------------------------------------------------------------------
# Mantém a tela selecionada durante a atualização do nome exibido no menu.
if st.session_state.get("tela_atual") == "Planejamento NL":
    st.session_state["tela_atual"] = "Planejar Priorização"

opcoes_tela = [
    "Liquidação (NL)",
    "Programa de Desembolso (PD)",
    "Pagamentos (OB)",
    "Planejar Priorização",
    "Relatório 009717",
]
if (
    "seletor_tela_global" not in st.session_state
    or st.session_state["seletor_tela_global"] not in opcoes_tela
):
    st.session_state["seletor_tela_global"] = st.session_state["tela_atual"]

with st.container(key="topo_navegacao"):
    st.markdown(
        """
        <div class='barra-sistema'>
            <span class='marca-sistema'>🏛️ SEAF | Painel de Controle Financeiro</span>
            <span class='exercicio-sistema'>EXERCÍCIO 2026</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    tela_selecionada = st.segmented_control(
        "Módulos do sistema",
        options=opcoes_tela,
        format_func=lambda opcao: {
            "Pagamentos (OB)": "💳 Pagamentos (OB)",
            "Liquidação (NL)": "📑 Liquidação (NL)",
            "Programa de Desembolso (PD)": "📅 Programa de Desembolso (PD)",
            "Planejar Priorização": "🎯 Planejar Priorização",
            "Relatório 009717": "📊 Relatório 009717",
        }[opcao],
        selection_mode="single",
        key="seletor_tela_global",
        label_visibility="collapsed",
    )
st.session_state["tela_atual"] = tela_selecionada or st.session_state["tela_atual"]

st.sidebar.markdown("---")

# -------------------------------------------------------------------------
# RENDERIZAÇÃO CONDICIONAL DAS TELAS
# -------------------------------------------------------------------------
if st.session_state["tela_atual"] == "Pagamentos (OB)":

    def atualizar_banco_via_csv():
        caminho_csv = r"C:\Users\victor.brenner\Desktop\Pagamentos_2026\02_Bases_Novas_Fontes\base_2026.csv"
        caminho_db = "pagamentos2026.db"

        if not os.path.exists(caminho_csv):
            st.sidebar.error(
                "Arquivo CSV não encontrado no caminho especificado."
            )
            return False

        try:
            df_novo = None
            for enc in ["utf-8-sig", "latin-1", "cp1252", "utf-8"]:
                for sep_tentativa in [";", ","]:
                    try:
                        df_novo = pd.read_csv(
                            caminho_csv, sep=sep_tentativa, encoding=enc
                        )
                        if df_novo is not None and len(df_novo.columns) > 1:
                            break
                    except:
                        continue
                if df_novo is not None and len(df_novo.columns) > 1:
                    break

            if df_novo is None:
                st.sidebar.error("Não foi possível ler o arquivo CSV.")
                return False

            df_novo.columns = [str(c).strip() for c in df_novo.columns]

            col_data = next(
                (
                    c
                    for c in df_novo.columns
                    if c.lower()
                    in ["data emissão", "data emissao", "data", "dt_emissao"]
                ),
                df_novo.columns[0],
            )
            col_ob = next(
                (
                    c
                    for c in df_novo.columns
                    if c == "DocumentoGD"
                    or "OB" in c.upper()
                    or "NÚMERO" in c.upper()
                    or "NUMERO" in c.upper()
                ),
                df_novo.columns[0],
            )
            col_valor = next(
                (
                    c
                    for c in df_novo.columns
                    if "VALOR" in c.upper() or "PAGAMENTO" in c.upper()
                ),
                df_novo.columns[-1],
            )

            df_novo = df_novo.dropna(subset=[col_data, col_ob, col_valor])
            df_novo["id_controle"] = (
                df_novo[col_data].astype(str).str.strip()
                + "_"
                + df_novo[col_valor].astype(str).str.strip()
                + "_"
                + df_novo.index.astype(str)
            )
            df_novo = df_novo.drop_duplicates(subset=["id_controle"])

            conn = sqlite3.connect(caminho_db)
            cursor = conn.cursor()

            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='pagamentos'"
            )
            if not cursor.fetchone():
                df_novo.to_sql(
                    "pagamentos", conn, if_exists="replace", index=False
                )
                cursor.execute(
                    "CREATE UNIQUE INDEX IF NOT EXISTS idx_pag_controle ON pagamentos(id_controle)"
                )
                conn.commit()
                st.sidebar.success(
                    f"Banco inicializado com {len(df_novo)} registros."
                )
            else:
                cursor.execute("PRAGMA table_info(pagamentos)")
                colunas_bd = [info[1] for info in cursor.fetchall()]
                if "id_controle" not in colunas_bd:
                    cursor.execute(
                        "ALTER TABLE pagamentos ADD COLUMN id_controle TEXT"
                    )
                    conn.commit()

                ids_existentes = (
                    pd.read_sql_query(
                        "SELECT id_controle FROM pagamentos", conn
                    )["id_controle"]
                    .dropna()
                    .tolist()
                )
                df_inserir = df_novo[
                    ~df_novo["id_controle"].isin(ids_existentes)
                ].copy()

                if not df_inserir.empty:
                    for col in colunas_bd:
                        if col not in df_inserir.columns:
                            df_inserir[col] = None
                    df_inserir = df_inserir[colunas_bd]
                    df_inserir.to_sql(
                        "pagamentos", conn, if_exists="append", index=False
                    )
                    st.sidebar.success(
                        f"Sucesso! {len(df_inserir)} novos pagamentos adicionados."
                    )
                else:
                    st.sidebar.info("Nenhum registro novo detectado no CSV.")

            conn.close()
            st.cache_data.clear()
            return True
        except Exception as e:
            st.sidebar.error(f"Erro ao processar atualização: {e}")
            return False

    @st.cache_data(ttl=60)
    def carregar_dados_auditoria():
        # Aba BASE: consolidado da BASE_HISTORICA com o mês atual tratado.
        LINK_PUBLICADO = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTD3b7L6byArEDgkVKOXXlc7RK0M2QKXLov83OydCaks3rDISWYWfgGNi6vG6pwy8t5Ul3Fd2wArhtT/pub?gid=1786485134&single=true&output=csv"
        try:
            df = ler_csv_url(LINK_PUBLICADO)
        except Exception as e:
            st.error(f"Erro ao conectar com a aba 'BASE' do Google Sheets: {e}")
            return pd.DataFrame()

        if df.empty:
            return df

        df = df.loc[:, ~df.columns.duplicated()]
        df.columns = [str(c).strip() for c in df.columns]

        colunas_obrigatorias = [
            "Número",
            "UG Emitente",
            "UG Pagadora",
            "Data Emissão",
            "Status",
            "Tipo de OB",
            "NE",
            "Credor",
            "Nome do Credor",
            "Valor",
            "Fonte",
            "Natureza",
            "Status de Envio",
            "RE",
            "PD",
            "GRUPO",
            "Elemento",
            "Despesa",
            "OBJETO",
            "DocumentoGD",
        ]
        for col in colunas_obrigatorias:
            if col not in df.columns:
                df[col] = None

        df = df.dropna(subset=["Valor"])

        # TRAVA DE DUPLICIDADE DA BASE OB
        # A planilha consolidada pode trazer a mesma OB duas vezes, com o CPF/CNPJ
        # apresentado de forma diferente (por exemplo, com zero à esquerda). Antes
        # de qualquer filtro ou soma, conservamos somente uma linha para a mesma
        # identificação financeira. Assim, atualizações futuras da base não voltam
        # a inflar os valores exibidos pelo painel.
        colunas_chave_ob = [
            coluna
            for coluna in [
                "Número",
                "NE",
                "Data Emissão",
                "Valor",
                "Fonte",
                "Nome do Credor",
                "Tipo de OB",
            ]
            if coluna in df.columns
        ]

        if colunas_chave_ob:
            for coluna in colunas_chave_ob:
                chave_coluna = f"__chave_ob_{coluna}"
                if coluna == "Valor":
                    df[chave_coluna] = converter_valor_monetario(
                        df[coluna]
                    ).round(2)
                elif coluna == "Data Emissão":
                    data_chave = pd.to_datetime(
                        df[coluna], errors="coerce", dayfirst=True
                    )
                    df[chave_coluna] = data_chave.dt.strftime("%Y-%m-%d").fillna(
                        df[coluna].fillna("").astype(str).str.strip().str.upper()
                    )
                else:
                    df[chave_coluna] = (
                        df[coluna]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                        .str.upper()
                        .str.replace(r"\s+", " ", regex=True)
                    )

            chaves_auxiliares = [f"__chave_ob_{coluna}" for coluna in colunas_chave_ob]
            df = df.drop_duplicates(subset=chaves_auxiliares, keep="first").copy()
            df.drop(columns=chaves_auxiliares, inplace=True)

        df["Despesa_Tratada"] = "CORRENTE"
        if "Despesa" in df.columns:
            serie_despesa = df["Despesa"]
            if isinstance(serie_despesa, pd.DataFrame):
                serie_despesa = serie_despesa.iloc[:, 0]
            amostra = serie_despesa.fillna("").astype(str).str.upper()

            def classificar_texto(val):
                v = str(val).upper().strip()
                if (
                    "DEA" in v
                    or "EXERC" in v
                    or "ANTERIOR" in v
                    or "RECONHECIMENTO" in v
                ):
                    return "DEA"
                elif "RP" in v or "RESTO" in v or "PAGAR" in v:
                    return "RP"
                else:
                    return "CORRENTE"

            df["Despesa_Tratada"] = amostra.apply(classificar_texto)

        df["Grupo_Tratado"] = "3 - OUTRAS DESPESAS CORRENTES"
        if "GRUPO" in df.columns:
            serie_grupo = df["GRUPO"]
            if isinstance(serie_grupo, pd.DataFrame):
                serie_grupo = serie_grupo.iloc[:, 0]
            amostra_grupo = (
                serie_grupo.fillna("").astype(str).str.strip().str.upper()
            )
            df["Grupo_Tratado"] = amostra_grupo.apply(
                lambda v: (
                    "4 - INVESTIMENTOS"
                    if "INVEST" in v or "4" in v
                    else "3 - OUTRAS DESPESAS CORRENTES"
                )
            )

        if "Valor" in df.columns:
            serie_valor = df["Valor"]
            if isinstance(serie_valor, pd.DataFrame):
                serie_valor = serie_valor.iloc[:, 0]
            df["Valor_Limpo"] = converter_valor_monetario(serie_valor)
        else:
            df["Valor_Limpo"] = 0.0

        if "Data Emissão" in df.columns:
            serie_data = df["Data Emissão"]
            if isinstance(serie_data, pd.DataFrame):
                serie_data = serie_data.iloc[:, 0]
            datas_convertidas = pd.to_datetime(
                serie_data, errors="coerce", dayfirst=True
            )
            df["Mes_Num"] = (
                datas_convertidas.dt.month.fillna(0)
                .astype(int)
                .astype(str)
                .str.zfill(2)
            )
            mapa_meses = {
                "01": "Jan/2026",
                "02": "Fev/2026",
                "03": "Mar/2026",
                "04": "Abr/2026",
                "05": "Mai/2026",
                "06": "Jun/2026",
                "07": "Jul/2026",
                "08": "Ago/2026",
                "09": "Set/2026",
                "10": "Out/2026",
                "11": "Nov/2026",
                "12": "Dez/2026",
            }
            df["Mes_Extenso"] = (
                df["Mes_Num"].map(mapa_meses).fillna("Não Identificado")
            )
        else:
            df["Mes_Extenso"] = "Não Identificado"

        df["Credor_Nome_Tratado"] = (
            df["Nome do Credor"]
            .fillna(df["Credor"])
            .fillna("NÃO IDENTIFICADO")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        df["Fonte_Tratada"] = (
            df["Fonte"].fillna("NÃO INFORMADA").astype(str).str.strip()
        )
        df["DocumentoGD_Tratado"] = (
            df["DocumentoGD"].fillna("NÃO CONSTA").astype(str).str.strip()
        )

        return df

    try:
        df_base = carregar_dados_auditoria()
    except Exception as e:
        st.error(f"Erro ao carregar colunas do banco: {e}")
        st.stop()

    ordem_meses_ano = [
        "Jan/2026",
        "Fev/2026",
        "Mar/2026",
        "Abr/2026",
        "Mai/2026",
        "Jun/2026",
        "Jul/2026",
        "Ago/2026",
        "Set/2026",
        "Out/2026",
        "Nov/2026",
        "Dez/2026",
    ]

    st.sidebar.markdown("#### 🎯 Filtros — Pagamentos (OB)")

    tipo_filtro_data = st.sidebar.radio(
        "Como deseja filtrar o período?",
        options=["Por Mês de Competência", "Por Intervalo de Datas"],
        index=(
            0
            if st.session_state["mem_ob_tipo_data"] == "Por Mês de Competência"
            else 1
        ),
        key="w_ob_tipo_data",
        on_change=sincronizar_filtro,
        args=("mem_ob_tipo_data", "w_ob_tipo_data"),
    )

    coluna_data = "Data Emissão"
    coluna_objeto = "OBJETO"

    # --- FUNÇÃO AUXILIAR DE FILTRAGEM DINÂMICA DADOS - TELA OB ---
    def filtrar_df_ob(df, ign_mes=False, ign_dsp=False, ign_cred=False, ign_fnt=False, ign_obj=False):
        d = df.copy()
        if d.empty:
            return d

        # Filtro Data/Mês
        if st.session_state["mem_ob_tipo_data"] == "Por Mês de Competência":
            if not ign_mes and st.session_state["mem_ob_meses"]:
                d = d[d["Mes_Extenso"].isin(st.session_state["mem_ob_meses"])]
        else:
            if not ign_mes:
                dt_i = st.session_state["mem_ob_dt_ini"]
                dt_f = st.session_state["mem_ob_dt_fim"]
                if dt_i and dt_f:
                    s_dt = pd.to_datetime(d[coluna_data], format="mixed", dayfirst=True, errors="coerce").dt.date
                    d = d[(s_dt >= dt_i) & (s_dt <= dt_f)]

        # Filtro Despesa
        if not ign_dsp and st.session_state["mem_ob_despesa"]:
            mapa_despesas = {
                "CORRENTE (Dotação do Ano)": "CORRENTE",
                "RP (Restos a Pagar)": "RP",
                "DEA (Exercícios Anteriores)": "DEA",
            }
            tipos_selecionados = [
                mapa_despesas[tipo]
                for tipo in st.session_state["mem_ob_despesa"]
                if tipo in mapa_despesas
            ]
            if tipos_selecionados:
                d = d[d["Despesa_Tratada"].isin(tipos_selecionados)]

        # Filtro Credores
        if not ign_cred and st.session_state["mem_ob_credores"]:
            d = d[d["Credor_Nome_Tratado"].isin(st.session_state["mem_ob_credores"])]

        # Filtro Fontes
        if not ign_fnt and st.session_state["mem_ob_fontes"]:
            d = d[d["Fonte_Tratada"].isin(st.session_state["mem_ob_fontes"])]

        # Filtro Objetos
        if not ign_obj and st.session_state["mem_ob_objetos"] and coluna_objeto in d.columns:
            d = d[d[coluna_objeto].isin(st.session_state["mem_ob_objetos"])]

        return d

    # CALCULAR OPÇÕES DINÂMICAS PARA CADA WIDGET TELA OB
    df_para_meses = filtrar_df_ob(df_base, ign_mes=True)
    lista_meses_fixa = [m for m in ordem_meses_ano if m in df_para_meses["Mes_Extenso"].unique()] if not df_para_meses.empty else []

    meses_selecionados = []
    data_inicio = None
    data_fim = None

    if tipo_filtro_data == "Por Mês de Competência":
        validos_m_ob = [m for m in st.session_state["mem_ob_meses"] if m in lista_meses_fixa]
        meses_selecionados = st.sidebar.multiselect(
            "Filtrar Período de Competência:",
            options=lista_meses_fixa,
            default=validos_m_ob,
            key="w_ob_meses",
            on_change=sincronizar_filtro,
            args=("mem_ob_meses", "w_ob_meses"),
        )
    else:
        if not df_base.empty and coluna_data in df_base.columns:
            datas_convertidas = (
                pd.to_datetime(
                    df_base[coluna_data],
                    format="mixed",
                    dayfirst=True,
                    errors="coerce",
                )
                .dt.date.dropna()
            )
            data_min = (
                datas_convertidas.min()
                if not datas_convertidas.empty
                else datetime.date(2026, 1, 1)
            )
            data_max = (
                datas_convertidas.max()
                if not datas_convertidas.empty
                else datetime.date(2026, 12, 31)
            )
        else:
            data_min = datetime.date(2026, 1, 1)
            data_max = datetime.date(2026, 12, 31)

        val_ini = (
            st.session_state["mem_ob_dt_ini"]
            if st.session_state["mem_ob_dt_ini"]
            else data_min
        )
        val_fim = (
            st.session_state["mem_ob_dt_fim"]
            if st.session_state["mem_ob_dt_fim"]
            else data_max
        )

        col_dt1, col_dt2 = st.sidebar.columns(2)
        with col_dt1:
            data_inicio = st.date_input(
                "Data Inicial:",
                value=val_ini,
                format="DD/MM/YYYY",
                key="w_ob_dt_ini",
                on_change=sincronizar_filtro,
                args=("mem_ob_dt_ini", "w_ob_dt_ini"),
            )
        with col_dt2:
            data_fim = st.date_input(
                "Data Final:",
                value=val_fim,
                format="DD/MM/YYYY",
                key="w_ob_dt_fim",
                on_change=sincronizar_filtro,
                args=("mem_ob_dt_fim", "w_ob_dt_fim"),
            )

    st.sidebar.markdown("---")

    # FILTRO TIPO DE DESPESA (Opções estáticas, seleção afeta os demais)
    opcoes_despesa_ob = [
        "CORRENTE (Dotação do Ano)",
        "RP (Restos a Pagar)",
        "DEA (Exercícios Anteriores)",
    ]
    despesas_validas_ob = [
        despesa
        for despesa in st.session_state["mem_ob_despesa"]
        if despesa in opcoes_despesa_ob
    ]

    despesas_selecionadas = st.sidebar.multiselect(
        "Filtrar por Tipo de Despesa:",
        options=opcoes_despesa_ob,
        default=despesas_validas_ob,
        placeholder="Todas as despesas",
        key="w_ob_despesa",
        on_change=sincronizar_filtro,
        args=("mem_ob_despesa", "w_ob_despesa"),
    )

    st.sidebar.markdown("---")

    # CREDORES DINÂMICOS
    df_para_credores = filtrar_df_ob(df_base, ign_cred=True)
    nomes_disponiveis = (
        sorted(
            [
                str(n).strip()
                for n in df_para_credores["Credor_Nome_Tratado"].unique()
                if n and str(n).lower() != "nan"
            ]
        )
        if not df_para_credores.empty
        else []
    )
    validos_c_ob = [
        c for c in st.session_state["mem_ob_credores"] if c in nomes_disponiveis
    ]

    nomes_selecionados = st.sidebar.multiselect(
        "Filtrar por Entidade / Credor:",
        options=nomes_disponiveis,
        default=validos_c_ob,
        key="w_ob_credores",
        on_change=sincronizar_filtro,
        args=("mem_ob_credores", "w_ob_credores"),
    )

    st.sidebar.markdown("---")
    # FONTES DINÂMICAS
    df_para_fontes = filtrar_df_ob(df_base, ign_fnt=True)
    lista_fontes = (
        sorted(
            [
                str(f).strip()
                for f in df_para_fontes["Fonte_Tratada"].unique()
                if f and str(f).lower() != "nan"
            ]
        )
        if not df_para_fontes.empty
        else []
    )
    validos_f_ob = [
        f for f in st.session_state["mem_ob_fontes"] if f in lista_fontes
    ]

    fontes_selecionadas = st.sidebar.multiselect(
        "Filtrar por Fonte de Recurso:",
        options=lista_fontes,
        default=validos_f_ob,
        placeholder="Todas as fontes (Exibe tudo)",
        key="w_ob_fontes",
        on_change=sincronizar_filtro,
        args=("mem_ob_fontes", "w_ob_fontes"),
    )

    st.sidebar.markdown("---")
    # OBJETOS DINÂMICOS
    df_para_objetos = filtrar_df_ob(df_base, ign_obj=True)

    if not df_para_objetos.empty and coluna_objeto in df_para_objetos.columns:
        lista_objetos = sorted(
            [
                str(obj).strip()
                for obj in df_para_objetos[coluna_objeto].dropna().unique()
                if str(obj).lower() != "nan"
            ]
        )
        validos_o_ob = [
            o for o in st.session_state["mem_ob_objetos"] if o in lista_objetos
        ]
        objeto_selecionado = st.sidebar.multiselect(
            "Filtrar por Objeto de Despesa:",
            options=lista_objetos,
            default=validos_o_ob,
            placeholder="Todos os objetos",
            key="w_ob_objetos",
            on_change=sincronizar_filtro,
            args=("mem_ob_objetos", "w_ob_objetos"),
        )
    else:
        objeto_selecionado = []

    st.sidebar.markdown("---")

    # DATAFRAME FINALMENTE FILTRADO TELA OB
    df_filtrado = filtrar_df_ob(df_base)

    st.sidebar.markdown("### 🔄 Atualizar Dados do Painel")
    if st.sidebar.button("🔄 Incorporar Novos Pagamentos do CSV", key="btn_csv_ob"):
        atualizar_banco_via_csv()
        st.rerun()

    # CABEÇALHO TELA 1
    st.markdown(
        "<h2 class='titulo-pagina'>📊 Painel de Controle de Pagamentos — Exercício 2026</h2>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Secretaria Executiva de Administração e Finanças (SEAF)</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Gerência Financeira (GFIN)</p>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)

    if not df_filtrado.empty:
        total_real_calculado = float(df_filtrado["Valor_Limpo"].sum())
        total_corrente = float(
            df_filtrado[df_filtrado["Despesa_Tratada"] == "CORRENTE"][
                "Valor_Limpo"
            ].sum()
        )
        total_rp = float(
            df_filtrado[df_filtrado["Despesa_Tratada"] == "RP"][
                "Valor_Limpo"
            ].sum()
        )
        total_dea = float(
            df_filtrado[df_filtrado["Despesa_Tratada"] == "DEA"][
                "Valor_Limpo"
            ].sum()
        )
        qtd_registros = int(df_filtrado.shape[0])
    else:
        total_real_calculado = total_corrente = total_rp = total_dea = 0.0
        qtd_registros = 0

    with col_kpi1:
        qtd_formatada_br = f"{qtd_registros:,}".replace(",", ".")
        st.markdown(
            f"<div class='metric-card'><p style='color: #6c757d; font-size: 11px; font-weight: bold; margin:0;'>VALOR TOTAL PAGO</p><h3 style='color: #002b49; margin: 5px 0;'>{formatar_brl(total_real_calculado)}</h3><p style='color: #28a745; font-size: 11px; margin:0;'>📋 Registros: {qtd_formatada_br}</p></div>",
            unsafe_allow_html=True,
        )
    with col_kpi2:
        st.markdown(
            f"<div class='metric-card'><p style='color: #6c757d; font-size: 11px; font-weight: bold; margin:0;'>CORRENTE</p><h3 style='color: #028090; margin: 5px 0;'>{formatar_brl(total_corrente)}</h3><p style='color: #6c757d; font-size: 11px; margin:0;'>Dotação do Ano</p></div>",
            unsafe_allow_html=True,
        )
    with col_kpi3:
        st.markdown(
            f"<div class='metric-card'><p style='color: #f77f00; font-size: 11px; font-weight: bold; margin:0;'>RESTOS A PAGAR (RP)</p><h3 style='color: #f77f00; margin: 5px 0;'>{formatar_brl(total_rp)}</h3><p style='color: #6c757d; font-size: 11px; margin:0;'>Exercícios Anteriores</p></div>",
            unsafe_allow_html=True,
        )
    with col_kpi4:
        st.markdown(
            f"<div class='metric-card'><p style='color: #d62828; font-size: 11px; font-weight: bold; margin:0;'>EXERC. ANTERIORES (DEA)</p><h3 style='color: #d62828; margin: 5px 0;'>{formatar_brl(total_dea)}</h3><p style='color: #6c757d; font-size: 11px; margin:0;'>Reconhecimento de Passivo</p></div>",
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("### 📋 1. Demonstrativo Analítico por tipo de Despesa")

    if not df_filtrado.empty:
        if (
            tipo_filtro_data == "Por Mês de Competência"
            and meses_selecionados
        ):
            meses_exibicao = [
                m for m in lista_meses_fixa if m in meses_selecionados
            ]
        else:
            meses_presentes = df_filtrado["Mes_Extenso"].unique()
            meses_exibicao = [
                m for m in lista_meses_fixa if m in meses_presentes
            ]

        if not meses_exibicao:
            meses_exibicao = lista_meses_fixa

        df_matriz = df_filtrado.pivot_table(
            index=["Despesa_Tratada", "Grupo_Tratado"],
            columns="Mes_Extenso",
            values="Valor_Limpo",
            aggfunc="sum",
            fill_value=0.0,
        ).reset_index()

        for m in meses_exibicao:
            if m not in df_matriz.columns:
                df_matriz[m] = 0.0

        df_matriz["Total Geral"] = df_matriz[meses_exibicao].sum(axis=1)

        def renderizar_tabela_simetrica_html(
            df_origem,
            chave_natureza,
            grupos_obrigatorios,
            titulo_bloco,
            cor_hexa,
        ):
            linhas = []
            for gnd in grupos_obrigatorios:
                match = df_origem[
                    (df_origem["Despesa_Tratada"] == chave_natureza)
                    & (df_origem["Grupo_Tratado"] == gnd)
                ]
                if not match.empty:
                    linhas.append(match.iloc[0].to_dict())
                else:
                    nova_linha = {
                        "Despesa_Tratada": chave_natureza,
                        "Grupo_Tratado": gnd,
                    }
                    for col_m in meses_exibicao + ["Total Geral"]:
                        nova_linha[col_m] = 0.0
                    linhas.append(nova_linha)

            linhas_corpo_html = ""
            totais_colunas = {m: 0.0 for m in meses_exibicao + ["Total Geral"]}

            for row in linhas:
                colunas_valores = ""
                for m in meses_exibicao + ["Total Geral"]:
                    val = float(row.get(m, 0.0))
                    totais_colunas[m] += val
                    colunas_valores += f"<td>{formatar_brl(val)}</td>"

                linhas_corpo_html += f"<tr><td><span class='gnd-badge' style='background-color: {cor_hexa};'></span>{row['Grupo_Tratado']}</td>{colunas_valores}</tr>"

            valores_totais_gnd = ""
            for m in meses_exibicao + ["Total Geral"]:
                valores_totais_gnd += (
                    f"<td>{formatar_brl(totais_colunas[m])}</td>"
                )

            cabecalhos_meses_html = "".join(
                [f"<th>{mes}</th>" for mes in meses_exibicao]
            )

            html_completo = (
                f"<div class='tabela-container'>"
                f"<div class='subtitulo-tabela-html' style='background: linear-gradient(90deg, {cor_hexa} 0%, #002b49 100%);'>{titulo_bloco}</div>"
                f"<table class='html-executiva'>"
                f"<thead><tr>"
                f"<th style='width: 30%;'>GRUPO DO GASTO (GND)</th>"
                f"{cabecalhos_meses_html}"
                f"<th>Total Geral</th>"
                f"</tr></thead>"
                f"<tbody>{linhas_corpo_html}"
                f"<tr class='linha-total-html'><td>📊 TOTAL GERAL DA NATUREZA</td>{valores_totais_gnd}</tr>"
                f"</tbody></table></div>"
            )
            st.markdown(html_completo, unsafe_allow_html=True)

        renderizar_tabela_simetrica_html(
            df_matriz,
            "CORRENTE",
            ["3 - OUTRAS DESPESAS CORRENTES", "4 - INVESTIMENTOS"],
            "🔵 DESPESAS CORRENTES (Dotação Ordinária do Ano)",
            "#028090",
        )
        renderizar_tabela_simetrica_html(
            df_matriz,
            "RP",
            ["3 - OUTRAS DESPESAS CORRENTES", "4 - INVESTIMENTOS"],
            "🟠 RESTOS A PAGAR - RP (Compromissos de Anos Anteriores)",
            "#f77f00",
        )
        renderizar_tabela_simetrica_html(
            df_matriz,
            "DEA",
            ["3 - OUTRAS DESPESAS CORRENTES", "4 - INVESTIMENTOS"],
            "🔴 DESPESAS DE EXERCÍCIOS ANTERIORES - DEA (Reconhecimento de Passivo)",
            "#d62828",
        )

    else:
        st.info("Nenhum registro financeiro localizado. Verifique os filtros.")

    st.markdown("---")

    st.markdown(
        "### 🏦 2. Distribuição Mensal Consolidada por Fonte de Recurso"
    )

    if not df_filtrado.empty:
        df_matriz_fonte = df_filtrado.pivot_table(
            index="Fonte_Tratada",
            columns="Mes_Extenso",
            values="Valor_Limpo",
            aggfunc="sum",
            fill_value=0.0,
        ).reset_index()

        for m in lista_meses_fixa:
            if m not in df_matriz_fonte.columns:
                df_matriz_fonte[m] = 0.0

        df_matriz_fonte["Total Geral"] = df_matriz_fonte[lista_meses_fixa].sum(
            axis=1
        )
        df_matriz_fonte = df_matriz_fonte.sort_values(
            by="Total Geral", ascending=False
        )

        linhas_fonte_html = ""
        totais_meses_fonte = {
            m: 0.0 for m in lista_meses_fixa + ["Total Geral"]
        }

        for _, row in df_matriz_fonte.iterrows():
            colunas_valores = ""
            for m in lista_meses_fixa + ["Total Geral"]:
                val = float(row[m])
                totais_meses_fonte[m] += val
                colunas_valores += f"<td>{formatar_brl(val)}</td>"
            linhas_fonte_html += (
                f"<tr><td>🏛️ {row['Fonte_Tratada']}</td>{colunas_valores}</tr>"
            )

        valores_totais_fonte = ""
        for m in lista_meses_fixa + ["Total Geral"]:
            valores_totais_fonte += (
                f"<td>{formatar_brl(totais_meses_fonte[m])}</td>"
            )

        cabecalhos_meses_fonte = "".join(
            [f"<th style='width: 10%;'>{mes}</th>" for mes in lista_meses_fixa]
        )

        html_fontes_resumo = (
            f"<div class='tabela-container'>"
            f"<div class='subtitulo-tabela-html' style='background: linear-gradient(90deg, #1d3557 0%, #002b49 100%);'>💰 Origem dos Recursos e Fluxo de Saída Monetária</div>"
            f"<table class='html-executiva'>"
            f"<thead><tr>"
            f"<th style='width: 30%;'>FONTE DE RECURSO</th>"
            f"{cabecalhos_meses_fonte}"
            f"<th style='width: 10%;'>Total Geral</th>"
            f"</tr></thead>"
            f"<tbody>{linhas_fonte_html}"
            f"<tr class='linha-total-html'><td>💰 TOTAL CONSOLIDADO POR FONTE</td>{valores_totais_fonte}</tr>"
            f"</tbody></table></div>"
        )
        st.markdown(html_fontes_resumo, unsafe_allow_html=True)

    st.markdown("---")

    st.markdown("### 📊 3. Análise Temporal e Desembolso Mensal")

    if not df_filtrado.empty:
        df_agrupado_mes = (
            df_filtrado.groupby("Mes_Extenso")
            .agg(
                Qtd_Docs=("Valor_Limpo", "count"),
                Total_Liq=("Valor_Limpo", "sum"),
            )
            .reset_index()
        )

        df_agrupado_mes = df_agrupado_mes.rename(
            columns={"Mes_Extenso": "Mês de Referência"}
        )
        df_agrupado_mes["Mês de Referência"] = pd.Categorical(
            df_agrupado_mes["Mês de Referência"],
            categories=lista_meses_fixa,
            ordered=True,
        )
        df_agrupado_mes = df_agrupado_mes.sort_values(
            "Mês de Referência"
        ).fillna(0)

        for m in lista_meses_fixa:
            if m not in df_agrupado_mes["Mês de Referência"].values:
                nova_linha_vazia = pd.DataFrame(
                    [{"Mês de Referência": m, "Qtd_Docs": 0, "Total_Liq": 0.0}]
                )
                df_agrupado_mes = pd.concat(
                    [df_agrupado_mes, nova_linha_vazia], ignore_index=True
                )

        df_agrupado_mes["Mês de Referência"] = pd.Categorical(
            df_agrupado_mes["Mês de Referência"],
            categories=lista_meses_fixa,
            ordered=True,
        )
        df_agrupado_mes = df_agrupado_mes.sort_values("Mês de Referência")

        with st.container(border=True):
            col_grafico, col_tabela = st.columns([1.1, 0.9], gap="large")

            with col_grafico:
                st.markdown(
                    "<p style='font-weight: 700; color: #002b49; margin-bottom: 5px; font-family: sans-serif;'>Curva Crítica de Desembolso Mensal</p>",
                    unsafe_allow_html=True,
                )

                def formatar_dinamico_br(x):
                    if x >= 1_000_000:
                        valor_m = x / 1_000_000
                        return (
                            f"R$ {valor_m:,.1f}".replace(",", "X")
                            .replace(".", ",")
                            .replace("X", ".")
                            + "M"
                        )
                    elif x >= 1_000:
                        valor_k = int(x / 1_000)
                        return f"R$ {valor_k:,}".replace(",", ".") + "K"
                    elif x > 0:
                        return (
                            f"R$ {x:,.2f}".replace(",", "X")
                            .replace(".", ",")
                            .replace("X", ".")
                        )
                    return "R$ 0"

                fig = px.line(
                    df_agrupado_mes,
                    x="Mês de Referência",
                    y="Total_Liq",
                    markers=True,
                    text=df_agrupado_mes["Total_Liq"].apply(
                        formatar_dinamico_br
                    ),
                    hover_data={"Qtd_Docs": True, "Total_Liq": ":,.2f"},
                )

                fig.update_traces(
                    line=dict(color="#028090", width=4),
                    marker=dict(size=10, color="#f77f00", symbol="circle"),
                    textposition="top center",
                )

                fig.update_layout(
                    xaxis=dict(
                        title=None,
                        showgrid=False,
                        tickfont=dict(size=11, color="#475569"),
                    ),
                    yaxis=dict(
                        title=None,
                        showgrid=True,
                        gridcolor="rgba(218, 224, 233, 0.6)",
                        tickfont=dict(size=11, color="#475569"),
                        tickformat=",.0f",
                    ),
                    margin=dict(l=10, r=10, t=30, b=10),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    height=320,
                )

                st.plotly_chart(
                    fig,
                    use_container_width=True,
                    config={"displayModeBar": False},
                )

            with col_tabela:
                titulo_resumo, acao_resumo = st.columns([1.15, 1], gap="small")
                with titulo_resumo:
                    st.markdown(
                        "<p style='font-weight: 700; color: #002b49; margin: 6px 0 12px; font-family: sans-serif;'>Resumo Gerencial por Mês</p>",
                        unsafe_allow_html=True,
                    )
                with acao_resumo:
                    resumo_ob_excel = gerar_resumo_gerencial_ob_excel(
                        df_filtrado, lista_meses_fixa
                    )
                    st.download_button(
                        "📥 Exportar .xlsx",
                        data=resumo_ob_excel,
                        file_name=(
                            "Resumo_Gerencial_Pagamentos_"
                            f"{datetime.date.today().strftime('%d-%m-%Y')}.xlsx"
                        ),
                        mime=(
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet"
                        ),
                        key="baixar_resumo_gerencial_ob",
                        use_container_width=True,
                    )

                total_documentos = int(df_agrupado_mes["Qtd_Docs"].sum())
                total_financeiro = float(df_agrupado_mes["Total_Liq"].sum())

                total_docs_formatado_br = f"{total_documentos:,}".replace(
                    ",", "."
                )

                linhas_tabela_html = ""
                for _, row in df_agrupado_mes.iterrows():
                    valor_formatado = formatar_brl(row["Total_Liq"])
                    qtd_docs_br = f"{int(row['Qtd_Docs']):,}".replace(",", ".")
                    linhas_tabela_html += f'<tr style="border-bottom: 1px solid #f1f5f9;"><td style="padding: 10px 15px; text-align: left; color: #334155; font-family: sans-serif; font-size: 13px;">{row["Mês de Referência"]}</td><td style="padding: 10px 15px; text-align: center; color: #334155; font-family: sans-serif; font-size: 13px;">{qtd_docs_br}</td><td style="padding: 10px 15px; text-align: right; color: #0f172a; font-family: sans-serif; font-size: 13px; font-weight: 600;">{valor_formatado}</td></tr>'

                html_tabela_gerencial = (
                    f'<div style="border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; background-color: #ffffff; width: 100%;">'
                    f'<table style="width: 100%; border-collapse: collapse; text-align: left; margin: 0; padding: 0;">'
                    f"<thead>"
                    f'<tr style="background-color: #f8fafc; border-bottom: 1px solid #e2e8f0;">'
                    f'<th style="padding: 12px 15px; font-family: sans-serif; font-size: 11px; font-weight: 700; color: #475569; text-align: left;">MÊS DE REFERÊNCIA</th>'
                    f'<th style="padding: 12px 15px; font-family: sans-serif; font-size: 11px; font-weight: 700; color: #475569; text-align: center;">QTD. DOCS</th>'
                    f'<th style="padding: 12px 15px; font-family: sans-serif; font-size: 11px; font-weight: 700; color: #475569; text-align: right;">TOTAL PAGO</th>'
                    f"</tr>"
                    f"</thead>"
                    f"<tbody>{linhas_tabela_html}</tbody>"
                    f"<tfoot>"
                    f'<tr style="background-color: #f8fafc; border-top: 2px solid #002b49; font-weight: 700;">'
                    f'<td style="padding: 12px 15px; font-family: sans-serif; font-size: 13px; color: #002b49; text-align: left;">📊 TOTAL GERAL</td>'
                    f'<td style="padding: 12px 15px; font-family: sans-serif; font-size: 13px; color: #002b49; text-align: center;">{total_docs_formatado_br}</td>'
                    f'<td style="padding: 12px 15px; font-family: sans-serif; font-size: 13px; color: #002b49; text-align: right;">{formatar_brl(total_financeiro)}</td>'
                    f"</tr>"
                    f"</tfoot>"
                    f"</table>"
                    f"</div>"
                )
                st.html(html_tabela_gerencial)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### 🏢 4. Detalhamento de Pagamentos por Credor")

        df_matriz_credor = df_filtrado.pivot_table(
            index="Credor_Nome_Tratado",
            columns="Mes_Extenso",
            values="Valor_Limpo",
            aggfunc="sum",
            fill_value=0.0,
        ).reset_index()

        for m in lista_meses_fixa:
            if m not in df_matriz_credor.columns:
                df_matriz_credor[m] = 0.0

        df_matriz_credor["Total Geral"] = df_matriz_credor[
            lista_meses_fixa
        ].sum(axis=1)
        df_matriz_credor = df_matriz_credor.sort_values(
            by="Total Geral", ascending=False
        )

        linhas_credor_html = ""
        totais_meses_credor = {
            m: 0.0 for m in lista_meses_fixa + ["Total Geral"]
        }

        for _, row in df_matriz_credor.iterrows():
            colunas_valores = ""
            for m in lista_meses_fixa + ["Total Geral"]:
                val = float(row[m])
                totais_meses_credor[m] += val
                colunas_valores += f"<td>{formatar_brl(val)}</td>"
            linhas_credor_html += f"<tr><td>{row['Credor_Nome_Tratado']}</td>{colunas_valores}</tr>"

        valores_totais_credor = ""
        for m in lista_meses_fixa + ["Total Geral"]:
            valores_totais_credor += (
                f"<td>{formatar_brl(totais_meses_credor[m])}</td>"
            )

        cabecalhos_meses_credor = "".join(
            [f"<th style='width: 10%;'>{mes}</th>" for mes in lista_meses_fixa]
        )

        html_credores = (
            f"<div class='tabela-container'>"
            f"<div class='subtitulo-tabela-html' style='background: linear-gradient(90deg, #3a537d 0%, #002b49 100%);'>🏢 Distribuição Mensal de Recursos por Fornecedor / Prestador de Serviço</div>"
            f"<table class='html-executiva'>"
            f"<thead><tr>"
            f"<th style='width: 30%;'>RAZÃO SOCIAL / CREDOR</th>"
            f"{cabecalhos_meses_credor}"
            f"<th style='width: 10%;'>Total Geral</th>"
            f"</tr></thead>"
            f"<tbody>{linhas_credor_html}"
            f"<tr class='linha-total-html'><td>🏢 TOTAL CONSOLIDADO DO FILTRO</td>{valores_totais_credor}</tr>"
            f"</tbody></table></div>"
        )
        st.markdown(html_credores, unsafe_allow_html=True)

elif st.session_state["tela_atual"] == "Liquidação (NL)":
    LINK_PLANILHA_1 = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTWnQkc7oF-YdXKoVTiYeUPYDHGzaeQaiEGqX6fNmB29mkzcd1kAvZVMujFDf02y7j1X8UJzqglAzTL/pub?gid=1892412645&single=true&output=csv"
    LINK_PLANILHA_2 = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTDHDK6dPnS9favMwwkNSLYZ5i9yjQrJjCEGxpifdQfD9_8dAsHSVlc8TECEyKJi0hWy3Wi5gEM_tI0/pub?gid=1866981074&single=true&output=csv"

    def converter_para_float(coluna):
        valores_str = coluna.fillna("0").astype(str)
        valores_str = valores_str.str.replace(
            r"[R$\s.]", "", regex=True
        ).str.replace(",", ".")
        return pd.to_numeric(valores_str, errors="coerce").fillna(0.0)

    def limpar_chave_ne(serie):
        return (
            serie.fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .str.replace(r"\.0$", "", regex=True)
            .str.replace(r"[^A-Z0-9]", "", regex=True)
        )

    def classificar_grupo_nl(valor):
        texto = "" if pd.isna(valor) else str(valor).strip().upper()
        if not texto or texto in {"NÃO INFORMADO", "NAO INFORMADO", "NAN", "NONE", "TODOS"}:
            return "NÃO INFORMADO"

        compacto = re.sub(r"[^A-Z0-9]", "", texto)

        if (
            re.search(r"\bGD1\b", texto)
            or re.search(r"\bGND1\b", texto)
            or compacto.startswith("GD1")
            or compacto.startswith("GND1")
            or compacto == "1"
            or compacto.startswith("1PESSOAL")
        ):
            return "GD1"

        if (
            re.search(r"\bGD3\b", texto)
            or re.search(r"\bGND3\b", texto)
            or compacto.startswith("GD3")
            or compacto.startswith("GND3")
            or compacto == "3"
            or compacto.startswith("3OUTRAS")
        ):
            return "GD3"

        if (
            re.search(r"\bGD4\b", texto)
            or re.search(r"\bGND4\b", texto)
            or compacto.startswith("GD4")
            or compacto.startswith("GND4")
            or compacto == "4"
            or compacto.startswith("4INVESTIMENTOS")
        ):
            return "GD4"

        return "NÃO INFORMADO"

    @st.cache_data(ttl=60)
    def carregar_dados_integrados():
        try:
            df1 = ler_csv_url(LINK_PLANILHA_1)
            df2 = ler_csv_url(LINK_PLANILHA_2)
        except Exception as e:
            st.error(f"Erro no carregamento das planilhas: {e}")
            return pd.DataFrame()

        if df1.empty:
            return df1

        df1.columns = [str(c).strip() for c in df1.columns]
        df2.columns = [str(c).strip() for c in df2.columns]

        col_nl = "Número" if "Número" in df1.columns else df1.columns[0]
        col_ne1 = next(
            (
                c
                for c in df1.columns
                if any(p in c.upper() for p in ["NE", "EMPENHO", "DOCUMENTONE"])
            ),
            df1.columns[0],
        )
        col_credor = (
            "Nome do Credor" if "Nome do Credor" in df1.columns else df1.columns[0]
        )
        col_retido = next(
            (c for c in df1.columns if "retido" in c.lower()), None
        )
        col_valor = "Valor" if "Valor" in df1.columns else df1.columns[-1]

        df1["NL_Numero"] = (
            df1[col_nl].fillna("NL NÃO INFORMADA").astype(str).str.strip()
        )
        df1["NE_Chave"] = limpar_chave_ne(df1[col_ne1])
        df1["Credor_Tratado"] = (
            df1[col_credor]
            .fillna("NÃO IDENTIFICADO")
            .astype(str)
            .str.strip()
            .str.upper()
        )

        col_data = next((c for c in df1.columns if "data" in c.lower()), None)
        if col_data:
            df1["Data_DT"] = pd.to_datetime(df1[col_data], errors="coerce")
            df1["Competencia"] = df1["Data_DT"].dt.strftime("%m/%Y")
        else:
            df1["Data_DT"] = pd.NaT
            df1["Competencia"] = "Não informada"

        df1["Grupo_Filtro"] = (
            df1["Grupo"].fillna("Todos").astype(str).str.strip()
            if "Grupo" in df1.columns
            else "Todos"
        )
        df1["Status_Filtro"] = (
            df1["Status"].fillna("Todos").astype(str).str.strip()
            if "Status" in df1.columns
            else "Todos"
        )

        df1["Valor_Total_Limpo"] = converter_para_float(df1[col_valor])
        df1["Valor_Retido_Limpo"] = (
            converter_para_float(df1[col_retido]) if col_retido else 0.0
        )

        col_ne2 = (
            "DocumentoNE"
            if "DocumentoNE" in df2.columns
            else next(
                (c for c in df2.columns if "NE" in c.upper()), df2.columns[0]
            )
        )
        col_fonte2 = (
            "Fonte"
            if "Fonte" in df2.columns
            else next(
                (c for c in df2.columns if "FONTE" in c.upper()), None
            )
        )
        col_objeto2 = (
            "Objeto da Despesa"
            if "Objeto da Despesa" in df2.columns
            else next(
                (c for c in df2.columns if "OBJETO" in c.upper()), None
            )
        )
        col_grupo2 = (
            "Grupo"
            if "Grupo" in df2.columns
            else next(
                (c for c in df2.columns if "GRUPO" in c.upper()), None
            )
        )

        df2["NE_Chave"] = limpar_chave_ne(df2[col_ne2])

        df2["Fonte_Relacao"] = (
            df2[col_fonte2].fillna("NÃO INFORMADA").astype(str).str.strip()
            if col_fonte2
            else "NÃO INFORMADA"
        )
        df2["Objeto_Relacao"] = (
            df2[col_objeto2].fillna("NÃO INFORMADO").astype(str).str.strip()
            if col_objeto2
            else "NÃO INFORMADO"
        )
        df2["Grupo_Relacao"] = (
            df2[col_grupo2].fillna("NÃO INFORMADO").astype(str).str.strip()
            if col_grupo2
            else "NÃO INFORMADO"
        )

        df2_dedup = df2[df2["NE_Chave"] != ""].drop_duplicates(
            subset=["NE_Chave"]
        )

        df_merged = pd.merge(
            df1,
            df2_dedup[
                ["NE_Chave", "Fonte_Relacao", "Objeto_Relacao", "Grupo_Relacao"]
            ],
            on="NE_Chave",
            how="left",
        )

        df_merged["Fonte_Relacao"] = df_merged["Fonte_Relacao"].fillna(
            "NÃO INFORMADA"
        )
        df_merged["Objeto_Relacao"] = df_merged["Objeto_Relacao"].fillna(
            "NÃO INFORMADO"
        )
        df_merged["Grupo_Relacao"] = df_merged["Grupo_Relacao"].fillna(
            "NÃO INFORMADO"
        )

        df_merged["Grupo_Classificado"] = df_merged["Grupo_Relacao"].apply(
            classificar_grupo_nl
        )
        mask_grupo_indisponivel = df_merged["Grupo_Classificado"] == "NÃO INFORMADO"
        df_merged.loc[mask_grupo_indisponivel, "Grupo_Classificado"] = (
            df_merged.loc[mask_grupo_indisponivel, "Grupo_Filtro"]
            .apply(classificar_grupo_nl)
        )

        return df_merged

    def renderizar_tabela_credor_dinamica(df_filtrado):
        totais_credor = (
            df_filtrado.groupby("Credor_Tratado")
            .agg({"Valor_Retido_Limpo": "sum", "Valor_Total_Limpo": "sum"})
            .sort_values(by="Valor_Total_Limpo", ascending=False)
        )

        tot_retido_geral = totais_credor["Valor_Retido_Limpo"].sum()
        tot_valor_geral = totais_credor["Valor_Total_Limpo"].sum()

        html_grupos = ""

        for credor, row_tot in totais_credor.iterrows():
            v_ret_credor = row_tot["Valor_Retido_Limpo"]
            v_tot_credor = row_tot["Valor_Total_Limpo"]

            df_cred = df_filtrado[df_filtrado["Credor_Tratado"] == credor]
            df_sub = (
                df_cred.groupby(["NL_Numero", "NE_Chave"], as_index=False)
                .agg({"Valor_Retido_Limpo": "sum", "Valor_Total_Limpo": "sum"})
                .sort_values(by="NL_Numero", ascending=True)
            )

            linhas_subtabela = ""
            for _, r in df_sub.iterrows():
                linhas_subtabela += f"""
<tr>
    <td style='width: 25%; text-align: center; font-family: monospace; font-weight: 600; color: #005691;'>{r['NL_Numero']}</td>
    <td style='width: 25%; text-align: center; font-family: monospace; color: #475569;'>{r['NE_Chave']}</td>
    <td style='width: 25%; text-align: center; color: #d97706;'>{formatar_brl(r['Valor_Retido_Limpo'])}</td>
    <td style='width: 25%; text-align: center; font-weight: 600; color: #002b49;'>{formatar_brl(r['Valor_Total_Limpo'])}</td>
</tr>"""

            html_grupos += f"""
<details class='credor-group'>
    <summary class='credor-summary'>
        <span>{credor}</span>
        <span style='text-align: right; color: #d97706;'>{formatar_brl(v_ret_credor)}</span>
        <span style='text-align: right; color: #002b49;'>{formatar_brl(v_tot_credor)}</span>
    </summary>
    <div class='subtabela-container'>
        <table class='subtabela-detalhe'>
            <thead>
                <tr>
                    <th style='width: 25%; text-align: center;'>NL</th>
                    <th style='width: 25%; text-align: center;'>NE</th>
                    <th style='width: 25%; text-align: center;'>VAL. RET.</th>
                    <th style='width: 25%; text-align: center;'>VALOR TOTAL</th>
                </tr>
            </thead>
            <tbody>
                {linhas_subtabela}
            </tbody>
        </table>
    </div>
</details>"""

        html_final = f"""
<div class='tabela-dinamica-container'>
    <div class='tabela-dinamica-header'>
        <div>CREDOR</div>
        <div style='text-align: right;'>VAL. RET.</div>
        <div style='text-align: right;'>VALOR TOTAL</div>
    </div>
    {html_grupos}
    <div class='tabela-dinamica-header' style='background: #f1f5f9; border-top: 2px solid #005691; border-bottom: none;'>
        <div style='color: #002b49; font-weight: 800; text-transform: uppercase; letter-spacing: 0.5px;'>TOTAL GERAL</div>
        <div style='text-align: right; color: #f77f00; font-weight: 800;'>{formatar_brl(tot_retido_geral)}</div>
        <div style='text-align: right; color: #028090; font-weight: 800;'>{formatar_brl(tot_valor_geral)}</div>
    </div>
</div>"""

        if hasattr(st, "html"):
            st.html(html_final)
        else:
            st.markdown(html_final, unsafe_allow_html=True)

    def gerar_relatorio_nl_excel(df_filtrado):
        """Preenche o modelo oficial sem remover as tabelas dinâmicas nativas."""
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as erro:
            raise ModuleNotFoundError(
                "O pacote openpyxl não está instalado. Execute: pip install -r requirements.txt"
            ) from erro

        # A aba-base do modelo tem exatamente estas sete colunas, que alimentam
        # as tabelas dinâmicas nativas existentes na segunda aba do arquivo.
        # A exportação não deve depender de funções definidas em outra seção
        # da tela. Assim ela continua funcionando também no Streamlit Cloud.
        def normalizar_nome_coluna(valor):
            texto = unicodedata.normalize("NFKD", str(valor or ""))
            texto = "".join(
                caractere for caractere in texto
                if not unicodedata.combining(caractere)
            )
            return re.sub(r"[^a-z0-9]+", "", texto.lower())

        def localizar_coluna(*opcoes):
            """Localiza uma coluna respeitando a ordem de prioridade informada."""
            colunas_normalizadas = {
                normalizar_nome_coluna(coluna): coluna
                for coluna in df_filtrado.columns
            }
            for opcao in opcoes:
                coluna = colunas_normalizadas.get(normalizar_nome_coluna(opcao))
                if coluna is not None:
                    return coluna
            return None

        def serie_texto(coluna, padrao=""):
            if coluna is None:
                return pd.Series(padrao, index=df_filtrado.index, dtype="object")
            return df_filtrado[coluna].fillna(padrao).astype(str).str.strip()

        def serie_numero(coluna):
            if coluna is None:
                return pd.Series(0.0, index=df_filtrado.index, dtype="float64")
            valores = df_filtrado[coluna]
            if pd.api.types.is_numeric_dtype(valores):
                return pd.to_numeric(valores, errors="coerce").fillna(0.0)

            def converter_valor(valor):
                if pd.isna(valor):
                    return 0.0
                if isinstance(valor, (int, float)):
                    return float(valor)

                texto = str(valor).strip()
                if texto.lower() in {"", "nan", "none", "null", "-"}:
                    return 0.0

                negativo = texto.startswith("(") and texto.endswith(")")
                texto = texto.replace("R$", "").replace("\xa0", "").replace(" ", "")
                texto = re.sub(r"[^0-9,.\-]", "", texto)

                if "," in texto:
                    texto = texto.replace(".", "").replace(",", ".")
                elif texto.count(".") > 1:
                    texto = texto.replace(".", "")

                try:
                    numero = float(texto)
                except (TypeError, ValueError):
                    return 0.0
                return -abs(numero) if negativo else numero

            return valores.map(converter_valor).astype(float)

        coluna_credor = localizar_coluna(
            "Credor_Tratado", "Nome do Credor", "Credor_NL",
            "Entidade / Credor", "Entidade", "Credor"
        )
        coluna_objeto = localizar_coluna(
            "Objeto_Relacao", "Objeto Despesa", "Objeto da Despesa",
            "Objeto de Despesa", "Objeto"
        )
        coluna_grupo = localizar_coluna(
            "Grupo_Classificado", "GD", "Grupo_Relacao",
            "Grupo", "Grupo de Despesa"
        )
        coluna_numero = localizar_coluna(
            "NL_Numero", "Numero_NL", "Número NL", "Numero NL",
            "Número", "Numero", "NL"
        )
        coluna_status = localizar_coluna(
            "Status_Filtro", "Status", "Status Comp.", "Status Comp"
        )
        coluna_valor = localizar_coluna(
            "Valor_Total_Limpo", "Valor", "Valor Total",
            "Valor Pago", "Valor Liquidado", "Valor da NL"
        )
        coluna_tipo_nl_modelo = localizar_coluna(
            "Tipo_NL_Filtro", "Tipo_NL", "Tipo de NL", "Tipo NL"
        )

        campos_obrigatorios = {
            "credor tratado": coluna_credor,
            "objeto da despesa": coluna_objeto,
            "grupo": coluna_grupo,
            "número da NL": coluna_numero,
            "valor": coluna_valor,
        }
        campos_ausentes = [
            nome for nome, coluna in campos_obrigatorios.items()
            if coluna is None
        ]
        if campos_ausentes:
            raise ValueError(
                "Não foi possível montar o relatório. Colunas ausentes: "
                + ", ".join(campos_ausentes)
            )

        grupo_relatorio = serie_texto(coluna_grupo, padrao="NÃO INFORMADO").str.upper()
        grupo_relatorio = grupo_relatorio.replace(
            {
                "GD1": "1",
                "GD3": "3",
                "GD4": "4",
                "": "NÃO INFORMADO",
                "NAN": "NÃO INFORMADO",
            }
        )

        base_relatorio_modelo = pd.DataFrame(
            {
                "Nome do Credor": serie_texto(coluna_credor),
                "Objeto Despesa": serie_texto(coluna_objeto),
                "GD": grupo_relatorio,
                "Número": serie_texto(coluna_numero),
                "Tipo de NL": serie_texto(
                    coluna_tipo_nl_modelo, padrao="NÃO INFORMADO"
                ),
                "Status": serie_texto(coluna_status),
                "Valor": serie_numero(coluna_valor),
            }
        ).sort_values(
            ["Objeto Despesa", "Nome do Credor", "Número"], kind="stable"
        )

        caminho_modelo = (
            Path(__file__).resolve().parent
            / "modelos"
            / "modelo_relatorio_liquidacao.xlsx"
        )
        if not caminho_modelo.exists():
            raise FileNotFoundError(
                "Modelo não encontrado. Inclua modelos/modelo_relatorio_liquidacao.xlsx no repositório."
            )

        workbook = load_workbook(caminho_modelo, data_only=False)
        cabecalhos_esperados = [
            "Nome do Credor",
            "Objeto Despesa",
            "GD",
            "Número",
            "Tipo de NL",
            "Status",
            "Valor",
        ]

        def normalizar_cabecalho_excel(valor):
            texto = unicodedata.normalize("NFKD", str(valor or ""))
            texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
            return re.sub(r"[^A-Z0-9]+", "", texto.upper())

        conjunto_esperado = {normalizar_cabecalho_excel(cabecalho) for cabecalho in cabecalhos_esperados}
        aba_base = None
        linha_cabecalho = None
        for aba in workbook.worksheets:
            for linha in range(1, min(aba.max_row, 10) + 1):
                cabecalhos_encontrados = {
                    normalizar_cabecalho_excel(aba.cell(linha, coluna).value)
                    for coluna in range(1, len(cabecalhos_esperados) + 1)
                }
                if conjunto_esperado.issubset(cabecalhos_encontrados):
                    aba_base = aba
                    linha_cabecalho = linha
                    break
            if aba_base is not None:
                break

        if aba_base is None:
            raise ValueError(
                "A aba-base do modelo não foi localizada. O modelo precisa conter as colunas de relatório."
            )

        for coluna, cabecalho in enumerate(cabecalhos_esperados, start=1):
            aba_base.cell(linha_cabecalho, coluna).value = cabecalho

        primeira_linha_dados = linha_cabecalho + 1
        for linha in aba_base.iter_rows(
            min_row=primeira_linha_dados,
            max_row=max(aba_base.max_row, primeira_linha_dados),
            min_col=1,
            max_col=len(cabecalhos_esperados),
        ):
            for celula in linha:
                celula.value = None

        for numero_linha, registro in enumerate(
            base_relatorio_modelo.itertuples(index=False, name=None),
            start=primeira_linha_dados,
        ):
            for numero_coluna, valor in enumerate(registro, start=1):
                if numero_coluna == 3:
                    grupo = str(valor).replace("GD", "").strip()
                    valor = int(grupo) if grupo.isdigit() else grupo
                elif numero_coluna == 7:
                    valor = float(valor) if pd.notna(valor) else 0.0
                else:
                    valor = "" if pd.isna(valor) else str(valor)

                celula = aba_base.cell(numero_linha, numero_coluna, value=valor)
                if numero_coluna == 7:
                    celula.number_format = 'R$ #,##0.00'

        # Faz o Excel atualizar os painéis dinâmicos quando o chefe abrir o arquivo.
        for aba in workbook.worksheets:
            for tabela_dinamica in getattr(aba, "_pivots", []):
                cache = getattr(tabela_dinamica, "cache", None)
                if cache is not None:
                    cache.refreshOnLoad = True
                    cache.enableRefresh = True

        try:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
        except AttributeError:
            pass

        arquivo = io.BytesIO()
        workbook.save(arquivo)
        arquivo.seek(0)
        return arquivo.getvalue()

        """Implementação anterior mantida apenas como referência histórica."""
        coluna_tipo_nl = next(
            (
                coluna
                for coluna in df_filtrado.columns
                if "TIPO" in str(coluna).upper()
                and "NL" in str(coluna).upper()
            ),
            None,
        )
        base_relatorio = pd.DataFrame(
            {
                "Nome do Credor": df_filtrado["Credor_Tratado"],
                "Objeto Despesa": df_filtrado["Objeto_Relacao"],
                "GD": df_filtrado["Grupo_Classificado"].replace(
                    {"GD1": "1", "GD3": "3", "GD4": "4"}
                ),
                "Número": df_filtrado["NL_Numero"],
                "Tipo de NL": (
                    df_filtrado[coluna_tipo_nl].fillna("NÃO INFORMADO").astype(str).str.strip()
                    if coluna_tipo_nl else "NÃO INFORMADO"
                ),
                "Status": df_filtrado["Status_Filtro"],
                "Valor": df_filtrado["Valor_Total_Limpo"],
            }
        )
        def adicionar_total(df_resumo):
            return pd.concat(
                [
                    df_resumo,
                    pd.DataFrame(
                        [{df_resumo.columns[0]: "Total Geral", "Valor": df_resumo["Valor"].sum()}]
                    ),
                ],
                ignore_index=True,
            )

        resumo_objeto = adicionar_total(
            base_relatorio.groupby("Objeto Despesa", as_index=False)["Valor"].sum().sort_values("Valor", ascending=False)
        )
        resumo_tipo = adicionar_total(
            base_relatorio.groupby("Tipo de NL", as_index=False)["Valor"].sum().sort_values("Valor", ascending=False)
        )
        resumo_status = adicionar_total(
            base_relatorio.groupby("Status", as_index=False)["Valor"].sum().sort_values("Valor", ascending=False)
        )
        resumo_gd = adicionar_total(
            base_relatorio.groupby("GD", as_index=False)["Valor"].sum().sort_values("GD")
        )

        arquivo = io.BytesIO()
        with pd.ExcelWriter(arquivo, engine="xlsxwriter") as writer:
            base_relatorio.to_excel(writer, sheet_name="Relatório Geral de NL", index=False)
            resumo_objeto.to_excel(writer, sheet_name="Tabela_Dinâmica", startrow=2, startcol=1, index=False)
            resumo_tipo.to_excel(writer, sheet_name="Tabela_Dinâmica", startrow=2, startcol=4, index=False)
            linha_status = 5 + len(resumo_tipo)
            resumo_status.to_excel(writer, sheet_name="Tabela_Dinâmica", startrow=linha_status, startcol=4, index=False)
            linha_gd = linha_status + 3 + len(resumo_status)
            resumo_gd.to_excel(writer, sheet_name="Tabela_Dinâmica", startrow=linha_gd, startcol=4, index=False)

            workbook = writer.book
            aba_base = writer.sheets["Relatório Geral de NL"]
            aba_painel = writer.sheets["Tabela_Dinâmica"]
            formato_titulo = workbook.add_format({"bold": True, "font_size": 14, "font_color": "#FFFFFF", "bg_color": "#002B49", "align": "center", "valign": "vcenter"})
            formato_subtitulo = workbook.add_format({"italic": True, "font_color": "#475569"})
            formato_cabecalho = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#092E4D", "align": "center", "valign": "vcenter"})
            formato_cabecalho_painel = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#4F81BD", "align": "left", "valign": "vcenter"})
            formato_cabecalho_painel_direita = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#4F81BD", "align": "right", "valign": "vcenter"})
            formato_moeda = workbook.add_format({"num_format": "R$ #,##0.00", "align": "right"})
            formato_texto_branco = workbook.add_format({"bg_color": "#FFFFFF"})
            formato_moeda_branco = workbook.add_format({"bg_color": "#FFFFFF", "num_format": "R$ #,##0.00", "align": "right"})
            formato_total_rotulo = workbook.add_format({"bold": True, "bg_color": "#D9EAF7"})
            formato_total_moeda = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "num_format": "R$ #,##0.00", "align": "right"})

            aba_base.freeze_panes(1, 0)
            aba_base.autofilter(0, 0, len(base_relatorio), len(base_relatorio.columns) - 1)
            aba_base.hide_gridlines(0)
            aba_base.set_column("A:A", 42)
            aba_base.set_column("B:B", 55)
            aba_base.set_column("C:C", 8)
            aba_base.set_column("D:F", 18)
            aba_base.set_column("G:G", 18, formato_moeda)
            aba_base.set_row(0, 24)
            for coluna, cabecalho in enumerate(base_relatorio.columns):
                aba_base.write(0, coluna, cabecalho, formato_cabecalho)
            for linha, valores in enumerate(
                base_relatorio.itertuples(index=False, name=None), start=1
            ):
                for coluna, valor in enumerate(valores[:-1]):
                    aba_base.write(linha, coluna, valor, formato_texto_branco)
                aba_base.write_number(linha, 6, float(valores[-1]), formato_moeda_branco)
            aba_base.set_tab_color("#028090")

            aba_painel.set_column("A:Z", 10, formato_texto_branco)
            aba_painel.set_column("A:A", 3, formato_texto_branco)
            aba_painel.merge_range("B1:F1", "Relatório Geral de Liquidações", formato_titulo)
            aba_painel.merge_range("B2:F2", "Painel consolidado conforme os filtros selecionados no sistema.", formato_subtitulo)
            aba_painel.set_column("B:B", 56, formato_texto_branco)
            aba_painel.set_column("C:C", 18, formato_moeda_branco)
            aba_painel.set_column("D:D", 4, formato_texto_branco)
            aba_painel.set_column("E:E", 24, formato_texto_branco)
            aba_painel.set_column("F:F", 18, formato_moeda_branco)
            aba_painel.freeze_panes(2, 1)
            aba_painel.hide_gridlines(0)
            aba_painel.set_tab_color("#002B49")

            for linha, coluna, dados, nome in [
                (2, 1, resumo_objeto, "ResumoObjeto"),
                (2, 4, resumo_tipo, "ResumoTipoNL"),
                (linha_status, 4, resumo_status, "ResumoStatus"),
                (linha_gd, 4, resumo_gd, "ResumoGD"),
            ]:
                aba_painel.add_table(
                    linha, coluna, linha + len(dados), coluna + 1,
                    {
                        "name": nome,
                        "columns": [
                            {
                                "header": dados.columns[0],
                                "header_format": formato_cabecalho_painel,
                            },
                            {
                                "header": "Soma de Valor",
                                "header_format": formato_cabecalho_painel_direita,
                                "format": formato_moeda,
                            },
                        ],
                        "style": None,
                        "banded_rows": False,
                    },
                )
                for indice, valores in enumerate(dados.iloc[:-1].itertuples(index=False, name=None)):
                    aba_painel.write(
                        linha + 1 + indice, coluna, valores[0], formato_texto_branco
                    )
                    aba_painel.write_number(
                        linha + 1 + indice, coluna + 1, float(valores[1]), formato_moeda_branco
                    )
                linha_total = linha + len(dados)
                aba_painel.write(
                    linha_total, coluna, dados.iloc[-1, 0], formato_total_rotulo
                )
                aba_painel.write(
                    linha_total, coluna + 1, dados.iloc[-1, 1], formato_total_moeda
                )

        arquivo.seek(0)
        return arquivo.getvalue()

    def renderizar_tabela_resumida(df_filtrado, coluna_grupo, titulo_coluna):
        df_agrupado = (
            df_filtrado.groupby(coluna_grupo)["Valor_Total_Limpo"]
            .sum()
            .reset_index()
        )
        df_agrupado = df_agrupado.sort_values(
            by="Valor_Total_Limpo", ascending=False
        )

        total_geral = df_agrupado["Valor_Total_Limpo"].sum()

        linhas_html = ""
        for _, row in df_agrupado.iterrows():
            item = row[coluna_grupo]
            valor = row["Valor_Total_Limpo"]
            linhas_html += f"""
<tr>
<td style='text-align: left; font-weight: 500;'>{item}</td>
<td style='text-align: right; font-weight: 600; color: #002b49; white-space: nowrap;'>{formatar_brl(valor)}</td>
</tr>"""

        classe_tabela = (
            "tabela-simples tabela-fontes"
            if titulo_coluna == "FONTE"
            else "tabela-simples"
        )
        html = f"""<div class='tabela-simples-container'>
<table class='{classe_tabela}'>
<thead>
<tr>
<th style='text-align: left;'>{titulo_coluna}</th>
<th style='text-align: right;'>VALOR TOTAL</th>
</tr>
</thead>
<tbody>
{linhas_html}
<tr class='total-row'>
<td style='text-align: left;'>TOTAL GERAL</td>
<td style='text-align: right; color: #028090; white-space: nowrap;'>{formatar_brl(total_geral)}</td>
</tr>
</tbody>
</table>
</div>"""

        st.markdown(html, unsafe_allow_html=True)

    # CABEÇALHO TELA 2
    st.markdown(
        "<h2 class='titulo-pagina'>📊 Painel de Controle Liquidação — Exercício 2026</h2>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Secretaria Executiva de Administração e Finanças (SEAF)</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Gerência Financeira (GFIN)</p>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    _, coluna_relatorio = st.columns([5, 1])
    with coluna_relatorio:
        botao_relatorio = st.empty()

    df_base = carregar_dados_integrados()

    if not df_base.empty:
        st.sidebar.markdown("#### 📑 Filtros — Liquidação (NL)")

        tipo_periodo = st.sidebar.radio(
            "Como deseja filtrar o período?",
            options=["Por Mês de Competência", "Por Intervalo de Datas"],
            index=(
                0
                if st.session_state["mem_nl_tipo_periodo"]
                == "Por Mês de Competência"
                else 1
            ),
            key="w_nl_tipo_periodo",
            on_change=sincronizar_filtro,
            args=("mem_nl_tipo_periodo", "w_nl_tipo_periodo"),
        )

        # --- FUNÇÃO AUXILIAR DE FILTRAGEM DINÂMICA DADOS - TELA NL ---
        def filtrar_df_nl(df, ign_per=False, ign_grp=False, ign_sts=False, ign_cred=False, ign_fnt=False, ign_obj=False):
            d = df.copy()
            if d.empty:
                return d

            # Período
            if not ign_per:
                if st.session_state["mem_nl_tipo_periodo"] == "Por Mês de Competência":
                    if st.session_state["mem_nl_comps"]:
                        d = d[d["Competencia"].astype(str).isin(st.session_state["mem_nl_comps"])]
                else:
                    data_inicio = st.session_state.get("mem_nl_dt_ini")
                    data_fim = st.session_state.get("mem_nl_dt_fim")
                    if data_inicio and data_fim:
                        inicio = pd.Timestamp(data_inicio).normalize()
                        fim = pd.Timestamp(data_fim).normalize()
                        if inicio <= fim:
                            datas_nl = pd.to_datetime(
                                d["Data_DT"], errors="coerce"
                            ).dt.normalize()
                            d = d[(datas_nl >= inicio) & (datas_nl <= fim)]

            # Grupo
            if not ign_grp and st.session_state["mem_nl_grupo"]:
                d = d[
                    d["Grupo_Classificado"].isin(st.session_state["mem_nl_grupo"])
                ]

            # Status
            if not ign_sts and st.session_state["mem_nl_status"] != "Todos":
                d = d[d["Status_Filtro"] == st.session_state["mem_nl_status"]]

            # Credor
            if not ign_cred and st.session_state["mem_nl_credores"]:
                d = d[d["Credor_Tratado"].isin(st.session_state["mem_nl_credores"])]

            # Fonte
            if not ign_fnt and st.session_state["mem_nl_fonte"]:
                d = d[
                    d["Fonte_Relacao"].isin(st.session_state["mem_nl_fonte"])
                ]

            # Objeto
            if not ign_obj and st.session_state["mem_nl_objetos"]:
                d = d[d["Objeto_Relacao"].isin(st.session_state["mem_nl_objetos"])]

            return d

        # CALCULAR OPÇÕES DINÂMICAS TELA NL
        if tipo_periodo == "Por Mês de Competência":
            df_para_per = filtrar_df_nl(df_base, ign_per=True)
            comps_unicas = df_para_per["Competencia"].dropna().astype(str).unique() if not df_para_per.empty else []
            comps = sorted(
                [
                    c
                    for c in comps_unicas
                    if c.strip() not in ["Não informada", "nan", "None", ""]
                ],
                key=lambda competencia: pd.to_datetime(
                    f"01/{competencia}", format="%d/%m/%Y", errors="coerce"
                ),
            )

            def exibir_competencia_nl(competencia):
                data_competencia = pd.to_datetime(
                    f"01/{competencia}", format="%d/%m/%Y", errors="coerce"
                )
                if pd.isna(data_competencia):
                    return competencia
                nomes_meses = [
                    "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                    "Jul", "Ago", "Set", "Out", "Nov", "Dez",
                ]
                return f"{nomes_meses[data_competencia.month - 1]}/{data_competencia.year}"

            validos_comp_nl = [
                c for c in st.session_state["mem_nl_comps"] if c in comps
            ]

            comp_sel = st.sidebar.multiselect(
                "Filtrar Período de Competência:",
                options=comps,
                default=validos_comp_nl,
                format_func=exibir_competencia_nl,
                placeholder="Selecione as opções",
                key="w_nl_comps",
                on_change=sincronizar_filtro,
                args=("mem_nl_comps", "w_nl_comps"),
            )
        else:
            datas_nl_validas = pd.to_datetime(
                df_base["Data_DT"], errors="coerce"
            ).dropna()

            if datas_nl_validas.empty:
                st.sidebar.info("Não há datas válidas para filtrar.")
            else:
                data_min_nl = datas_nl_validas.min().date()
                data_max_nl = datas_nl_validas.max().date()

                def ajustar_data_nl(valor, padrao):
                    data = pd.to_datetime(valor, errors="coerce")
                    if pd.isna(data):
                        return padrao
                    data = data.date()
                    return min(max(data, data_min_nl), data_max_nl)

                data_ini_padrao = ajustar_data_nl(
                    st.session_state.get("mem_nl_dt_ini"), data_min_nl
                )
                data_fim_padrao = ajustar_data_nl(
                    st.session_state.get("mem_nl_dt_fim"), data_max_nl
                )
                if data_ini_padrao > data_fim_padrao:
                    data_ini_padrao = data_min_nl
                    data_fim_padrao = data_max_nl

                col_ini_nl, col_fim_nl = st.sidebar.columns(2)
                with col_ini_nl:
                    data_ini_nl = st.date_input(
                        "Data Inicial:",
                        value=data_ini_padrao,
                        min_value=data_min_nl,
                        max_value=data_max_nl,
                        format="DD/MM/YYYY",
                        key="w_nl_dt_ini",
                        on_change=sincronizar_filtro,
                        args=("mem_nl_dt_ini", "w_nl_dt_ini"),
                    )
                with col_fim_nl:
                    data_fim_nl = st.date_input(
                        "Data Final:",
                        value=data_fim_padrao,
                        min_value=data_min_nl,
                        max_value=data_max_nl,
                        format="DD/MM/YYYY",
                        key="w_nl_dt_fim",
                        on_change=sincronizar_filtro,
                        args=("mem_nl_dt_fim", "w_nl_dt_fim"),
                    )

                st.session_state["mem_nl_datas"] = (data_ini_nl, data_fim_nl)
                if data_ini_nl > data_fim_nl:
                    st.sidebar.error(
                        "A data inicial deve ser anterior ou igual à data final."
                    )

        st.sidebar.divider()

        # FILTRO DE GRUPO
        opcoes_grp_nl = ["GD1", "GD3", "GD4"]
        grupos_validos_nl = [
            grupo
            for grupo in st.session_state["mem_nl_grupo"]
            if grupo in opcoes_grp_nl
        ]
        st.sidebar.multiselect(
            "Filtrar por Grupo:",
            options=opcoes_grp_nl,
            default=grupos_validos_nl,
            placeholder="Todos os grupos (Exibe tudo)",
            key="w_nl_grupo",
            on_change=sincronizar_filtro,
            args=("mem_nl_grupo", "w_nl_grupo"),
        )

        # STATUS DINÂMICO
        df_para_sts = filtrar_df_nl(df_base, ign_sts=True)
        statuses = sorted(
            [s for s in df_para_sts["Status_Filtro"].unique() if s != "Todos"]
        ) if not df_para_sts.empty else []
        opcoes_sts_nl = ["Todos"] + statuses
        idx_sts_nl = (
            opcoes_sts_nl.index(st.session_state["mem_nl_status"])
            if st.session_state["mem_nl_status"] in opcoes_sts_nl
            else 0
        )
        status_sel = st.sidebar.selectbox(
            "Filtrar por Status:",
            options=opcoes_sts_nl,
            index=idx_sts_nl,
            key="w_nl_status",
            on_change=sincronizar_filtro,
            args=("mem_nl_status", "w_nl_status"),
        )

        st.sidebar.divider()

        # CREDORES DINÂMICOS
        df_para_cred = filtrar_df_nl(df_base, ign_cred=True)
        credores = sorted(df_para_cred["Credor_Tratado"].unique()) if not df_para_cred.empty else []
        validos_c_nl = [
            c for c in st.session_state["mem_nl_credores"] if c in credores
        ]
        credor_sel = st.sidebar.multiselect(
            "Filtrar por Entidade / Credor:",
            options=credores,
            default=validos_c_nl,
            placeholder="Selecione as opções",
            key="w_nl_credores",
            on_change=sincronizar_filtro,
            args=("mem_nl_credores", "w_nl_credores"),
        )

        st.sidebar.divider()

        # FONTES DINÂMICAS
        df_para_fnt = filtrar_df_nl(df_base, ign_fnt=True)
        fontes = sorted(
            [
                f
                for f in df_para_fnt["Fonte_Relacao"].unique()
                if f != "NÃO INFORMADA"
            ]
        ) if not df_para_fnt.empty else []
        validas_f_nl = [
            fonte
            for fonte in st.session_state["mem_nl_fonte"]
            if fonte in fontes
        ]
        fontes_selecionadas_nl = st.sidebar.multiselect(
            "Filtrar por Fonte de Recurso:",
            options=fontes,
            default=validas_f_nl,
            placeholder="Todas as fontes (Exibe tudo)",
            key="w_nl_fonte",
            on_change=sincronizar_filtro,
            args=("mem_nl_fonte", "w_nl_fonte"),
        )

        st.sidebar.divider()

        # OBJETOS DINÂMICOS
        df_para_obj = filtrar_df_nl(df_base, ign_obj=True)
        objetos = sorted(
            [
                o
                for o in df_para_obj["Objeto_Relacao"].unique()
                if o != "NÃO INFORMADO"
            ]
        ) if not df_para_obj.empty else []
        
        validos_obj_nl = [
            o for o in st.session_state["mem_nl_objetos"] if o in objetos
        ]
        objeto_sel = st.sidebar.multiselect(
            "Filtrar por Objeto de Despesa:",
            options=objetos,
            default=validos_obj_nl,
            placeholder="Todos os objetos",
            key="w_nl_objetos",
            on_change=sincronizar_filtro,
            args=("mem_nl_objetos", "w_nl_objetos"),
        )

        st.sidebar.divider()

        st.sidebar.markdown("### ⚙️ Atualizar Dados do Painel")
        if st.sidebar.button(
            "🔄 Recarregar Dados das Liquidações",
            use_container_width=True,
            key="btn_recarregar_nl",
        ):
            st.cache_data.clear()
            st.rerun()

        # DATAFRAME FINALMENTE FILTRADO TELA NL
        df_filtrado = filtrar_df_nl(df_base)

        @st.dialog("Gerar Relatório de Liquidações")
        def janela_relatorio_nl():
            st.caption(
                "Escolha o Grupo e a Fonte para o arquivo. Os demais filtros "
                "selecionados no painel também serão mantidos."
            )

            if "relatorio_nl_grupos" not in st.session_state:
                st.session_state["relatorio_nl_grupos"] = list(
                    st.session_state["mem_nl_grupo"]
                )
            if "relatorio_nl_fontes" not in st.session_state:
                st.session_state["relatorio_nl_fontes"] = list(
                    st.session_state["mem_nl_fonte"]
                )

            grupos_relatorio = st.multiselect(
                "Grupo (vazio = todos)",
                ["GD1", "GD3", "GD4"],
                key="relatorio_nl_grupos",
            )
            df_opcoes_relatorio = filtrar_df_nl(
                df_base, ign_grp=True, ign_fnt=True
            )
            if grupos_relatorio:
                df_opcoes_relatorio = df_opcoes_relatorio[
                    df_opcoes_relatorio["Grupo_Classificado"].isin(grupos_relatorio)
                ]
            fontes_relatorio = sorted(
                df_opcoes_relatorio["Fonte_Relacao"].dropna().astype(str).unique()
            )
            fontes_selecionadas = st.multiselect(
                "Fonte de recurso (vazio = todas)",
                fontes_relatorio,
                key="relatorio_nl_fontes",
            )

            df_exportacao = df_opcoes_relatorio.copy()
            if fontes_selecionadas:
                df_exportacao = df_exportacao[
                    df_exportacao["Fonte_Relacao"].isin(fontes_selecionadas)
                ]

            if df_exportacao.empty:
                st.warning("Não há liquidações para os filtros escolhidos.")
                return

            try:
                relatorio_excel = gerar_relatorio_nl_excel(df_exportacao)
                st.download_button(
                    "Baixar relatório",
                    data=relatorio_excel,
                    file_name=(
                        "Relatório Geral de Liquidação "
                        f"{datetime.date.today().strftime('%d.%m.%Y')}.xlsx"
                    ),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="btn_baixar_relatorio_nl",
                )
            except (ModuleNotFoundError, FileNotFoundError, RuntimeError, ValueError) as erro:
                st.error(f"Não foi possível gerar o relatório pelo modelo: {erro}")

        if botao_relatorio.button(
            "📊 Gerar Relatório",
            use_container_width=True,
            key="btn_gerar_relatorio_nl",
        ):
            janela_relatorio_nl()

        qtd_liquidada = len(df_filtrado)
        valor_total = df_filtrado["Valor_Total_Limpo"].sum()

        mask_gd3 = df_filtrado["Grupo_Classificado"] == "GD3"
        mask_gd4 = df_filtrado["Grupo_Classificado"] == "GD4"

        valor_gd3 = df_filtrado.loc[mask_gd3, "Valor_Total_Limpo"].sum()
        valor_gd4 = df_filtrado.loc[mask_gd4, "Valor_Total_Limpo"].sum()

        col_k1, col_k2, col_k3, col_k4 = st.columns(4)

        with col_k1:
            qtd_formatada_br = f"{qtd_liquidada:,}".replace(",", ".")
            st.markdown(
                f"""
                <div class='metric-card'>
                    <p style='color: #6c757d; font-size: 11px; font-weight: bold; margin:0;'>QTD DE LIQUIDAÇÕES</p>
                    <h3 style='color: #002b49; margin: 5px 0;'>{qtd_formatada_br}</h3>
                    <p style='color: #28a745; font-size: 11px; margin:0;'>📋 Documentos NL</p>
                </div>
            """,
                unsafe_allow_html=True,
            )

        with col_k2:
            st.markdown(
                f"""
                <div class='metric-card'>
                    <p style='color: #6c757d; font-size: 11px; font-weight: bold; margin:0;'>VALOR TOTAL</p>
                    <h3 style='color: #028090; margin: 5px 0;'>{formatar_brl(valor_total)}</h3>
                    <p style='color: #6c757d; font-size: 11px; margin:0;'>Total Liquidado</p>
                </div>
            """,
                unsafe_allow_html=True,
            )

        with col_k3:
            st.markdown(
                f"""
                <div class='metric-card'>
                    <p style='color: #6c757d; font-size: 11px; font-weight: bold; margin:0;'>VALOR TOTAL GD3</p>
                    <h3 style='color: #f77f00; margin: 5px 0;'>{formatar_brl(valor_gd3)}</h3>
                    <p style='color: #6c757d; font-size: 11px; margin:0;'>Grupo GD3</p>
                </div>
            """,
                unsafe_allow_html=True,
            )

        with col_k4:
            st.markdown(
                f"""
                <div class='metric-card'>
                    <p style='color: #6c757d; font-size: 11px; font-weight: bold; margin:0;'>VALOR TOTAL GD4</p>
                    <h3 style='color: #2563eb; margin: 5px 0;'>{formatar_brl(valor_gd4)}</h3>
                    <p style='color: #6c757d; font-size: 11px; margin:0;'>Grupo GD4</p>
                </div>
            """,
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)

        col_credor, col_resumos = st.columns([1.15, 0.85])

        with col_credor:
            renderizar_tabela_credor_dinamica(df_filtrado)

        with col_resumos:
            renderizar_tabela_resumida(df_filtrado, "Fonte_Relacao", "FONTE")
            st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
            renderizar_tabela_resumida(df_filtrado, "Objeto_Relacao", "OBJETO")
    else:
        st.warning("Aguardando carregamento e relacionamento das planilhas...")

elif st.session_state["tela_atual"] == "Programa de Desembolso (PD)":
    # ---------------------------------------------------------------------
    # TELA PD — Programa de Desembolso
    # ---------------------------------------------------------------------
    st.markdown(
        "<h2 class='titulo-pagina'>📅 Programa de Desembolso — Exercício 2026</h2>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Secretaria Executiva de Administração e Finanças (SEAF)</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Gerência Financeira (GFIN)</p>",
        unsafe_allow_html=True,
    )
    st.markdown("---")
    _, coluna_relatorio_pd = st.columns([5, 1])
    with coluna_relatorio_pd:
        botao_relatorio_pd = st.empty()

    dados_pd, erro_pd = carregar_dados_pd()
    if erro_pd:
        st.error(
            "Não foi possível carregar a base de PD. "
            "Confirme que a aba PD_ABA_TRATADA está publicada na web como CSV "
            "ou liberada para leitura pública. Detalhe técnico: " + erro_pd
        )
    elif dados_pd.empty:
        st.info("Ainda não há PDs tratados para os filtros selecionados.")
    else:
        st.sidebar.markdown("#### 📅 Filtros — Programa de Desembolso (PD)")
        data_min_pd = dados_pd["Data Emissão"].dropna().min()
        data_max_pd = dados_pd["Data Emissão"].dropna().max()

        if pd.notna(data_min_pd) and pd.notna(data_max_pd):
            intervalo_padrao_pd = (data_min_pd.date(), data_max_pd.date())
            intervalo_pd = st.sidebar.date_input(
                "Período de emissão",
                value=intervalo_padrao_pd,
                format="DD/MM/YYYY",
                key="filtro_pd_periodo",
            )
        else:
            intervalo_pd = None

        def opcoes_pd(df, coluna):
            return sorted(
                valor for valor in df[coluna].dropna().astype(str).unique()
                if valor and valor.upper() != "NÃO INFORMADO"
            )

        filtro_ug_pd = st.sidebar.multiselect(
            "UG pagadora", opcoes_pd(dados_pd, "UG Pagadora"),
            key="filtro_pd_ug",
        )
        filtro_fonte_pd = st.sidebar.multiselect(
            "Fonte de recurso", opcoes_pd(dados_pd, "Fonte"),
            key="filtro_pd_fonte",
        )
        filtro_gd_pd = st.sidebar.multiselect(
            "GD", opcoes_pd(dados_pd, "GD"), key="filtro_pd_gd",
        )
        filtro_despesa_pd = st.sidebar.multiselect(
            "Natureza da despesa", opcoes_pd(dados_pd, "Despesa"),
            key="filtro_pd_despesa",
        )
        filtro_objeto_pd = st.sidebar.multiselect(
            "Objeto da despesa", opcoes_pd(dados_pd, "Objeto da Despesa"),
            key="filtro_pd_objeto",
        )
        filtro_credor_pd = st.sidebar.multiselect(
            "Credor", opcoes_pd(dados_pd, "Nome do Credor"),
            key="filtro_pd_credor",
        )
        filtro_status_pd = st.sidebar.multiselect(
            "Status", opcoes_pd(dados_pd, "Status"), key="filtro_pd_status",
        )
        if st.sidebar.button("🔄 Atualizar dados de PD", key="atualizar_pd"):
            carregar_dados_pd.clear()
            st.rerun()

        filtrado_pd = dados_pd.copy()
        if isinstance(intervalo_pd, (tuple, list)) and len(intervalo_pd) == 2:
            inicio_pd, fim_pd = intervalo_pd
            datas_pd = filtrado_pd["Data Emissão"].dt.date
            filtrado_pd = filtrado_pd[(datas_pd >= inicio_pd) & (datas_pd <= fim_pd)]
        for coluna, selecionados in {
            "UG Pagadora": filtro_ug_pd,
            "Fonte": filtro_fonte_pd,
            "GD": filtro_gd_pd,
            "Despesa": filtro_despesa_pd,
            "Objeto da Despesa": filtro_objeto_pd,
            "Nome do Credor": filtro_credor_pd,
            "Status": filtro_status_pd,
        }.items():
            if selecionados:
                filtrado_pd = filtrado_pd[filtrado_pd[coluna].isin(selecionados)]

        @st.dialog("Gerar Relatório de Programa de Desembolso")
        def janela_relatorio_pd():
            st.caption(
                "Escolha o GD e a fonte para o arquivo. Os demais filtros "
                "selecionados nesta tela serão mantidos."
            )
            opcoes_gd_relatorio = opcoes_pd(filtrado_pd, "GD")
            grupos_relatorio_pd = st.multiselect(
                "GD (vazio = todos)",
                opcoes_gd_relatorio,
                key="relatorio_pd_grupos",
            )
            base_fontes_relatorio_pd = filtrado_pd.copy()
            if grupos_relatorio_pd:
                base_fontes_relatorio_pd = base_fontes_relatorio_pd[
                    base_fontes_relatorio_pd["GD"].isin(grupos_relatorio_pd)
                ]
            fontes_relatorio_pd = st.multiselect(
                "Fonte de recurso (vazio = todas)",
                opcoes_pd(base_fontes_relatorio_pd, "Fonte"),
                key="relatorio_pd_fontes",
            )
            exportacao_pd = base_fontes_relatorio_pd.copy()
            if fontes_relatorio_pd:
                exportacao_pd = exportacao_pd[
                    exportacao_pd["Fonte"].isin(fontes_relatorio_pd)
                ]
            if exportacao_pd.empty:
                st.warning("Não há PDs para os filtros escolhidos.")
                return
            try:
                relatorio_pd_excel = gerar_relatorio_pd_excel(exportacao_pd)
                st.download_button(
                    "Baixar relatório",
                    data=relatorio_pd_excel,
                    file_name=(
                        "Relatório Programa de Desembolso "
                        f"{datetime.date.today().strftime('%d.%m.%Y')}.xlsx"
                    ),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="btn_baixar_relatorio_pd",
                )
            except (ModuleNotFoundError, RuntimeError, ValueError) as erro:
                st.error(f"Não foi possível gerar o relatório de PD: {erro}")

        if botao_relatorio_pd.button(
            "📊 Gerar Relatório",
            use_container_width=True,
            key="btn_gerar_relatorio_pd",
            type="primary",
        ):
            janela_relatorio_pd()

        qtd_pds = len(filtrado_pd)
        valor_pd = float(filtrado_pd["Valor"].sum())
        valor_gd3_pd = float(filtrado_pd.loc[filtrado_pd["GD"] == "GD3", "Valor"].sum())
        valor_gd4_pd = float(filtrado_pd.loc[filtrado_pd["GD"] == "GD4", "Valor"].sum())
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        with kpi1:
            st.markdown(
                f"<div class='metric-card'><p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>QTD. DE PDs</p><h3 style='color:#002b49; margin:5px 0;'>{qtd_pds:,}</h3><p style='color:#28a745; font-size:11px; margin:0;'>📋 Documentos PD</p></div>".replace(",", "."),
                unsafe_allow_html=True,
            )
        with kpi2:
            st.markdown(
                f"<div class='metric-card'><p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>VALOR TOTAL</p><h3 style='color:#028090; margin:5px 0;'>{formatar_brl(valor_pd)}</h3><p style='color:#6c757d; font-size:11px; margin:0;'>Total programado</p></div>",
                unsafe_allow_html=True,
            )
        with kpi3:
            st.markdown(
                f"<div class='metric-card'><p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>VALOR TOTAL GD3</p><h3 style='color:#f77f00; margin:5px 0;'>{formatar_brl(valor_gd3_pd)}</h3><p style='color:#6c757d; font-size:11px; margin:0;'>Grupo GD3</p></div>",
                unsafe_allow_html=True,
            )
        with kpi4:
            st.markdown(
                f"<div class='metric-card'><p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>VALOR TOTAL GD4</p><h3 style='color:#2563eb; margin:5px 0;'>{formatar_brl(valor_gd4_pd)}</h3><p style='color:#6c757d; font-size:11px; margin:0;'>Grupo GD4</p></div>",
                unsafe_allow_html=True,
            )
        st.markdown("<br>", unsafe_allow_html=True)

        def resumo_pd(coluna, titulo, limite=10):
            resumo = (
                filtrado_pd.groupby(coluna, dropna=False, as_index=False)["Valor"].sum()
                .sort_values("Valor", ascending=False)
                .head(limite)
            )
            resumo.columns = [titulo, "Valor total"]
            return resumo

        coluna_credor_pd, coluna_resumos_pd = st.columns([1.25, 0.75])
        with coluna_credor_pd:
            renderizar_tabela_credor_pd(filtrado_pd)
        with coluna_resumos_pd:
            renderizar_tabela_resumo_pd(
                resumo_pd("UG Pagadora", "UG pagadora"),
                "UG pagadora",
                "UG pagadora",
            )
            st.markdown("<div style='height: 14px;'></div>", unsafe_allow_html=True)
            renderizar_tabela_resumo_pd(
                resumo_pd("Fonte", "Fonte"),
                "Fonte",
                "Fonte",
            )

elif st.session_state["tela_atual"] == "Relatório 009717":
    # ---------------------------------------------------------------------
    # TELA EXCLUSIVA — RELATÓRIO 009717
    # IMPORTANTE: todas as alterações deste módulo ficam isoladas aqui.
    # ---------------------------------------------------------------------
    st.markdown(
        "<h2 class='titulo-pagina'>📊 Relatório 009717</h2>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='subtitulo-pagina'>Relatório de Pagamento — acompanhamento e conferência executiva</p>",
        unsafe_allow_html=True,
    )

    if "url_api_relatorio_009717" not in st.session_state:
        st.session_state["url_api_relatorio_009717"] = URL_API_RELATORIO_009717_PADRAO
    if "dados_relatorio_009717" not in st.session_state:
        st.session_state["dados_relatorio_009717"] = None
    if "pdf_relatorio_009717" not in st.session_state:
        st.session_state["pdf_relatorio_009717"] = None
    if "pdf_relatorio_009717_nome" not in st.session_state:
        st.session_state["pdf_relatorio_009717_nome"] = "Relatorio_009717.pdf"

    url_api_009717 = st.session_state["url_api_relatorio_009717"].strip()

    if not url_api_009717:
        st.warning(
            "A conexão está pronta, mas ainda falta informar a URL de implantação "
            "do Web App do Apps Script."
        )
        url_digitada_009717 = st.text_input(
            "URL do Web App do Relatório 009717",
            placeholder="https://script.google.com/macros/s/.../exec",
            key="campo_url_api_relatorio_009717",
        ).strip()
        if url_digitada_009717:
            st.session_state["url_api_relatorio_009717"] = url_digitada_009717
            url_api_009717 = url_digitada_009717

    # Ações compactas do módulo — três botões lado a lado.
    col_conexao_009717, col_atualizar_009717, col_pdf_topo_009717, _espaco_acoes_009717 = st.columns(
        [0.72, 0.86, 1.02, 2.40]
    )

    with col_conexao_009717:
        testar_conexao_009717 = st.button(
            "🔗 Testar conexão",
            use_container_width=True,
            key="btn_testar_conexao_009717",
            disabled=not bool(url_api_009717),
        )

    with col_atualizar_009717:
        atualizar_relatorio_009717 = st.button(
            "🔄 Atualizar relatório",
            use_container_width=True,
            key="btn_atualizar_relatorio_009717",
            disabled=not bool(url_api_009717),
            type="primary",
        )

    with col_pdf_topo_009717:
        gerar_pdf_topo_009717 = st.button(
            "📄 Exportar relatório em PDF",
            use_container_width=True,
            key="btn_exportar_pdf_009717",
            disabled=not bool(url_api_009717),
        )

    if testar_conexao_009717:
        try:
            with st.spinner("Testando conexão com o Apps Script..."):
                retorno = chamar_api_relatorio_009717(
                    url_api_009717,
                    "status",
                    timeout=30,
                )
            st.success(
                f"Conexão realizada com sucesso — "
                f"{retorno.get('planilha', 'planilha localizada')}."
            )
        except Exception as erro:
            st.error(f"Não foi possível conectar ao Apps Script: {erro}")

    if atualizar_relatorio_009717:
        try:
            with st.spinner(
                "Processando cronologia, OB, retenção e recalculando o Relatório 009717..."
            ):
                retorno_processamento = chamar_api_relatorio_009717(
                    url_api_009717,
                    "processar",
                    timeout=180,
                )

                # Depois de processar, relê a aba FOI_PAGO diretamente.
                # Assim o painel não fica preso ao conteúdo retornado pela
                # própria ação "processar", que pode representar o layout anterior.
                retorno = chamar_api_relatorio_009717(
                    url_api_009717,
                    "ler",
                    timeout=60,
                )

            st.session_state["dados_relatorio_009717"] = retorno
            st.session_state["pdf_relatorio_009717"] = None
            st.success(
                retorno_processamento.get(
                    "mensagem",
                    "Relatório atualizado com sucesso."
                )
            )
        except Exception as erro:
            st.error(f"Falha ao atualizar o Relatório 009717: {erro}")

    # Ao abrir a aplicação novamente, carrega automaticamente o último relatório
    # calculado no Google Sheets. Isso substitui o antigo botão "Carregar atual".
    if (
        st.session_state.get("dados_relatorio_009717") is None
        and url_api_009717
    ):
        try:
            with st.spinner("Carregando o último Relatório 009717..."):
                retorno_inicial_009717 = chamar_api_relatorio_009717(
                    url_api_009717,
                    "ler",
                    timeout=60,
                )
            st.session_state["dados_relatorio_009717"] = retorno_inicial_009717
        except Exception as erro:
            st.warning(
                "Não foi possível carregar automaticamente o último relatório. "
                f"Use 'Atualizar Relatório'. Detalhe: {erro}"
            )

    dados_009717 = st.session_state.get("dados_relatorio_009717")

    if isinstance(dados_009717, dict):
        foi_pago_009717 = dados_009717.get("foi_pago", {}) or {}
        cabecalho_009717 = foi_pago_009717.get("cabecalho", []) or []
        linhas_009717 = foi_pago_009717.get("linhas", []) or []

        st.caption(
            f"Última atualização: "
            f"{dados_009717.get('atualizado_em', 'não informada')}"
        )

        if cabecalho_009717 and linhas_009717:
            df_009717 = pd.DataFrame(
                linhas_009717,
                columns=cabecalho_009717,
            )

            def _localizar_coluna_009717(df, *nomes):
                mapa = {
                    unicodedata.normalize("NFKD", str(coluna))
                    .encode("ascii", "ignore")
                    .decode("ascii")
                    .strip()
                    .upper()
                    .replace("_", " "): coluna
                    for coluna in df.columns
                }
                for nome in nomes:
                    chave = (
                        unicodedata.normalize("NFKD", str(nome))
                        .encode("ascii", "ignore")
                        .decode("ascii")
                        .strip()
                        .upper()
                        .replace("_", " ")
                    )
                    if chave in mapa:
                        return mapa[chave]
                return None

            col_fonte_009717 = _localizar_coluna_009717(
                df_009717,
                "Fonte",
            )
            col_credor_009717 = _localizar_coluna_009717(
                df_009717,
                "Nome do Credor",
                "Credor",
            )
            col_natureza_009717 = _localizar_coluna_009717(
                df_009717,
                "Natureza Despesa",
                "Natureza da Despesa",
                "Natureza",
            )
            col_valor_009717 = _localizar_coluna_009717(
                df_009717,
                "Valor",
            )
            col_ob_009717 = _localizar_coluna_009717(
                df_009717,
                "OB_NUMERO",
                "OB Numero",
                "OB",
            )
            col_status_009717 = _localizar_coluna_009717(
                df_009717,
                "STATUS_OB",
                "Status OB",
                "Status",
            )

            colunas_obrigatorias_009717 = {
                "Fonte": col_fonte_009717,
                "Credor": col_credor_009717,
                "Natureza": col_natureza_009717,
                "Valor": col_valor_009717,
                "OB": col_ob_009717,
                "Status": col_status_009717,
            }
            faltantes_009717 = [
                nome
                for nome, coluna in colunas_obrigatorias_009717.items()
                if coluna is None
            ]

            if faltantes_009717:
                st.error(
                    "Não foi possível montar a visão executiva. "
                    "Colunas não encontradas na aba FOI_PAGO: "
                    + ", ".join(faltantes_009717)
                )
            else:
                df_exec_009717 = pd.DataFrame(
                    {
                        "Fonte": df_009717[col_fonte_009717],
                        "Credor": df_009717[col_credor_009717],
                        "Natureza": df_009717[col_natureza_009717],
                        "Valor": converter_valor_monetario(
                            df_009717[col_valor_009717]
                        ),
                        "OB": df_009717[col_ob_009717],
                        "Status_Original": df_009717[col_status_009717],
                    }
                )

                def _status_executivo_009717(valor):
                    texto_status = str(valor or "").strip().upper()
                    if "NÃO PAGO" in texto_status or "NAO PAGO" in texto_status:
                        return "Não pago"
                    if "RETEN" in texto_status:
                        return "Retenção"
                    if "PAGO" in texto_status:
                        return "Pago"
                    return "Sem status"

                df_exec_009717["Status"] = (
                    df_exec_009717["Status_Original"]
                    .apply(_status_executivo_009717)
                )

                df_exec_009717["Fonte"] = (
                    df_exec_009717["Fonte"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
                df_exec_009717["Credor"] = (
                    df_exec_009717["Credor"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
                df_exec_009717["Natureza"] = (
                    df_exec_009717["Natureza"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
                df_exec_009717["OB"] = (
                    df_exec_009717["OB"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

                # ---------------------------------------------------------
                # FILTROS EXECUTIVOS — BARRA LATERAL RESERVADA
                # ---------------------------------------------------------
                status_disponiveis_009717 = [
                    status
                    for status in ["Pago", "Não pago", "Retenção", "Sem status"]
                    if status in df_exec_009717["Status"].unique()
                ]

                fontes_disponiveis_009717 = sorted(
                    [
                        valor
                        for valor in df_exec_009717["Fonte"].unique()
                        if valor
                    ]
                )

                credores_disponiveis_009717 = sorted(
                    [
                        valor
                        for valor in df_exec_009717["Credor"].unique()
                        if valor
                    ]
                )

                naturezas_disponiveis_009717 = sorted(
                    [
                        valor
                        for valor in df_exec_009717["Natureza"].unique()
                        if valor
                    ]
                )

                st.sidebar.markdown("#### 🔎 Filtros — Relatório 009717")

                status_sel_009717 = st.sidebar.multiselect(
                    "Status do pagamento",
                    options=status_disponiveis_009717,
                    default=[],
                    placeholder="Todos",
                    key="filtro_status_009717",
                )

                st.sidebar.divider()

                fonte_sel_009717 = st.sidebar.multiselect(
                    "Fonte",
                    options=fontes_disponiveis_009717,
                    default=[],
                    placeholder="Todas",
                    key="filtro_fonte_009717",
                )

                st.sidebar.divider()

                credor_sel_009717 = st.sidebar.multiselect(
                    "Credor",
                    options=credores_disponiveis_009717,
                    default=[],
                    placeholder="Todos",
                    key="filtro_credor_009717",
                )

                st.sidebar.divider()

                natureza_sel_009717 = st.sidebar.multiselect(
                    "Natureza",
                    options=naturezas_disponiveis_009717,
                    default=[],
                    placeholder="Todas",
                    key="filtro_natureza_009717",
                )

                st.sidebar.divider()
                st.sidebar.caption(
                    "Os cards e a tabela respondem aos filtros selecionados."
                )

                df_filtrado_009717 = df_exec_009717.copy()

                if status_sel_009717:
                    df_filtrado_009717 = df_filtrado_009717[
                        df_filtrado_009717["Status"].isin(status_sel_009717)
                    ]

                if fonte_sel_009717:
                    df_filtrado_009717 = df_filtrado_009717[
                        df_filtrado_009717["Fonte"].isin(fonte_sel_009717)
                    ]

                if credor_sel_009717:
                    df_filtrado_009717 = df_filtrado_009717[
                        df_filtrado_009717["Credor"].isin(credor_sel_009717)
                    ]

                if natureza_sel_009717:
                    df_filtrado_009717 = df_filtrado_009717[
                        df_filtrado_009717["Natureza"].isin(natureza_sel_009717)
                    ]

                # ---------------------------------------------------------
                # CARDS EXECUTIVOS
                # ---------------------------------------------------------
                # Base dos cards financeiros respeita Fonte/Credor/Natureza.
                # O filtro de Status fica refletido no card "Valor em exibição".
                df_base_cards_009717 = df_exec_009717.copy()

                if fonte_sel_009717:
                    df_base_cards_009717 = df_base_cards_009717[
                        df_base_cards_009717["Fonte"].isin(fonte_sel_009717)
                    ]
                if credor_sel_009717:
                    df_base_cards_009717 = df_base_cards_009717[
                        df_base_cards_009717["Credor"].isin(credor_sel_009717)
                    ]
                if natureza_sel_009717:
                    df_base_cards_009717 = df_base_cards_009717[
                        df_base_cards_009717["Natureza"].isin(natureza_sel_009717)
                    ]

                valor_total_pago_009717 = float(
                    df_base_cards_009717.loc[
                        df_base_cards_009717["Status"] == "Pago",
                        "Valor",
                    ].sum()
                )
                valor_total_nao_pago_009717 = float(
                    df_base_cards_009717.loc[
                        df_base_cards_009717["Status"] == "Não pago",
                        "Valor",
                    ].sum()
                )
                valor_filtrado_009717 = float(
                    df_filtrado_009717["Valor"].sum()
                )
                qtd_filtrada_009717 = int(len(df_filtrado_009717))

                card_1, card_2, card_3, card_4 = st.columns(4)

                with card_1:
                    st.markdown(
                        f"""
                        <div class='metric-card'>
                            <p style='color:#6c757d;font-size:11px;font-weight:bold;margin:0;'>
                                VALOR TOTAL PAGO
                            </p>
                            <h3 style='color:#028090;margin:5px 0;'>
                                {formatar_brl(valor_total_pago_009717)}
                            </h3>
                            <p style='color:#6c757d;font-size:11px;margin:0;'>
                                Pagamentos confirmados por OB
                            </p>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with card_2:
                    st.markdown(
                        f"""
                        <div class='metric-card'>
                            <p style='color:#6c757d;font-size:11px;font-weight:bold;margin:0;'>
                                VALOR NÃO PAGO
                            </p>
                            <h3 style='color:#9f2d2d;margin:5px 0;'>
                                {formatar_brl(valor_total_nao_pago_009717)}
                            </h3>
                            <p style='color:#6c757d;font-size:11px;margin:0;'>
                                Registros ainda sem OB
                            </p>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with card_3:
                    st.markdown(
                        f"""
                        <div class='metric-card'>
                            <p style='color:#6c757d;font-size:11px;font-weight:bold;margin:0;'>
                                VALOR EM EXIBIÇÃO
                            </p>
                            <h3 style='color:#005691;margin:5px 0;'>
                                {formatar_brl(valor_filtrado_009717)}
                            </h3>
                            <p style='color:#6c757d;font-size:11px;margin:0;'>
                                Conforme os filtros selecionados
                            </p>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with card_4:
                    st.markdown(
                        f"""
                        <div class='metric-card'>
                            <p style='color:#6c757d;font-size:11px;font-weight:bold;margin:0;'>
                                REGISTROS EM EXIBIÇÃO
                            </p>
                            <h3 style='color:#002b49;margin:5px 0;'>
                                {qtd_filtrada_009717}
                            </h3>
                            <p style='color:#6c757d;font-size:11px;margin:0;'>
                                Linhas visíveis na conferência
                            </p>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                st.markdown("<div style='height: 12px;'></div>", unsafe_allow_html=True)

                # ---------------------------------------------------------
                # EXPORTAÇÃO DO RELATÓRIO
                # O botão fica no topo, ao lado de Testar e Atualizar.
                # ---------------------------------------------------------
                col_pdf_1, col_pdf_2 = st.columns([1.05, 2.95])
                gerar_pdf_009717 = gerar_pdf_topo_009717

                if gerar_pdf_009717:
                    try:
                        with st.spinner(
                            "Gerando o PDF consolidado diretamente da aba RELATORIO_SIAFIN..."
                        ):
                            retorno_pdf_009717 = chamar_api_relatorio_009717(
                                url_api_009717,
                                "pdf",
                                timeout=120,
                            )

                        conteudo_pdf_base64 = retorno_pdf_009717.get(
                            "pdf_base64",
                            "",
                        )
                        if not conteudo_pdf_base64:
                            raise ValueError(
                                "O Apps Script não retornou o conteúdo do PDF."
                            )

                        st.session_state["pdf_relatorio_009717"] = (
                            base64.b64decode(conteudo_pdf_base64)
                        )
                        st.session_state["pdf_relatorio_009717_nome"] = (
                            retorno_pdf_009717.get("nome_arquivo")
                            or "Relatorio_009717.pdf"
                        )
                        st.success(
                            retorno_pdf_009717.get(
                                "mensagem",
                                "PDF consolidado gerado com sucesso.",
                            )
                        )
                    except Exception as erro:
                        st.session_state["pdf_relatorio_009717"] = None
                        st.error(
                            f"Falha ao gerar o PDF do Relatório 009717: {erro}"
                        )

                pdf_pronto_009717 = st.session_state.get(
                    "pdf_relatorio_009717"
                )

                # ---------------------------------------------------------
                # CABEÇALHO DA TABELA + DOWNLOAD COMPACTO
                # O download fica alinhado à direita, imediatamente acima
                # da tabela, evitando a aparência de botão solto na tela.
                # ---------------------------------------------------------
                col_titulo_tabela_009717, col_download_tabela_009717 = st.columns(
                    [3.75, 1.05],
                    vertical_alignment="bottom",
                )

                with col_titulo_tabela_009717:
                    st.markdown(
                        """
                        <div style="margin-top:14px; margin-bottom:8px;">
                            <span style="
                                color:#002b49;
                                font-size:17px;
                                font-weight:800;
                            ">Conferência dos pagamentos</span>
                            <div style="
                                color:#64748b;
                                font-size:12px;
                                margin-top:2px;
                            ">
                                Fonte, credor, natureza, valor, OB e situação do pagamento
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                if pdf_pronto_009717:
                    with col_download_tabela_009717:
                        st.download_button(
                            "⬇️ Baixar PDF consolidado",
                            data=pdf_pronto_009717,
                            file_name=st.session_state.get(
                                "pdf_relatorio_009717_nome",
                                "Relatorio_009717.pdf",
                            ),
                            mime="application/pdf",
                            use_container_width=True,
                            key="download_pdf_009717",
                        )

                # ---------------------------------------------------------
                # TABELA EXECUTIVA — SOMENTE AS COLUNAS NECESSÁRIAS
                # ---------------------------------------------------------

                tabela_009717 = df_filtrado_009717[
                    [
                        "Fonte",
                        "Credor",
                        "Natureza",
                        "Valor",
                        "OB",
                        "Status",
                    ]
                ].copy()

                if tabela_009717.empty:
                    st.info(
                        "Nenhum registro encontrado para os filtros selecionados."
                    )
                else:
                    # -----------------------------------------------------
                    # FORMATAÇÃO EXECUTIVA DA TABELA
                    # -----------------------------------------------------
                    def formatar_brl_tabela_009717(valor):
                        try:
                            numero = float(valor)
                        except (TypeError, ValueError):
                            numero = 0.0

                        texto_valor = f"{numero:,.2f}"
                        texto_valor = (
                            texto_valor
                            .replace(",", "X")
                            .replace(".", ",")
                            .replace("X", ".")
                        )
                        return f"R$ {texto_valor}"

                    def formatar_status_tabela_009717(status):
                        status_limpo = str(status or "").strip().lower()

                        if status_limpo == "pago":
                            return (
                                "<span class='status-009717 status-pago-009717'>"
                                "✅ Pago"
                                "</span>"
                            )

                        if status_limpo in {"não pago", "nao pago"}:
                            return (
                                "<span class='status-009717 status-nao-pago-009717'>"
                                "❌ Não pago"
                                "</span>"
                            )

                        if status_limpo in {"retenção", "retencao"}:
                            return (
                                "<span class='status-009717 status-retencao-009717'>"
                                "🟡 Retenção"
                                "</span>"
                            )

                        return (
                            "<span class='status-009717 status-neutro-009717'>"
                            f"{html.escape(str(status or 'Sem status'))}"
                            "</span>"
                        )

                    linhas_html_009717 = []

                    for _, linha_009717 in tabela_009717.iterrows():
                        fonte_html = html.escape(str(linha_009717["Fonte"] or ""))
                        credor_html = html.escape(str(linha_009717["Credor"] or ""))
                        natureza_html = html.escape(str(linha_009717["Natureza"] or ""))
                        ob_html = html.escape(str(linha_009717["OB"] or ""))
                        valor_html = formatar_brl_tabela_009717(
                            linha_009717["Valor"]
                        )
                        status_html = formatar_status_tabela_009717(
                            linha_009717["Status"]
                        )

                        linhas_html_009717.append(
                            f"""
                            <tr>
                                <td class="fonte-009717" title="{fonte_html}">
                                    {fonte_html}
                                </td>
                                <td class="credor-009717" title="{credor_html}">
                                    {credor_html}
                                </td>
                                <td class="natureza-009717">
                                    {natureza_html}
                                </td>
                                <td class="valor-009717">
                                    {valor_html}
                                </td>
                                <td class="ob-009717" title="{ob_html}">
                                    {ob_html}
                                </td>
                                <td class="status-cell-009717">
                                    {status_html}
                                </td>
                            </tr>
                            """
                        )

                    altura_tabela_009717 = min(
                        640,
                        56 + (len(tabela_009717) * 43),
                    )

                    html_tabela_009717 = textwrap.dedent(
                        f"""
                        <style>
                            .tabela-executiva-wrap-009717 {{
                                width: 100%;
                                max-height: {altura_tabela_009717}px;
                                overflow: auto;
                                border: 1px solid #d7e0e8;
                                border-radius: 10px;
                                background: #ffffff;
                                box-shadow: 0 3px 12px rgba(15, 40, 65, 0.05);
                            }}

                            .tabela-executiva-009717 {{
                                width: 100%;
                                border-collapse: separate;
                                border-spacing: 0;
                                table-layout: fixed;
                                font-family: inherit;
                                font-size: 12px;
                                color: #243746;
                            }}

                            .tabela-executiva-009717 thead th {{
                                position: sticky;
                                top: 0;
                                z-index: 2;
                                background: #087da0;
                                color: #ffffff;
                                font-size: 11px;
                                font-weight: 800;
                                letter-spacing: 0.35px;
                                text-transform: uppercase;
                                padding: 11px 12px;
                                border-right: 1px solid rgba(255,255,255,0.20);
                                border-bottom: 1px solid #066b89;
                                text-align: left;
                            }}

                            .tabela-executiva-009717 thead th:first-child {{
                                border-top-left-radius: 9px;
                            }}

                            .tabela-executiva-009717 thead th:last-child {{
                                border-top-right-radius: 9px;
                                border-right: none;
                            }}

                            .tabela-executiva-009717 tbody td {{
                                padding: 10px 12px;
                                border-bottom: 1px solid #e6ebef;
                                border-right: 1px solid #edf1f4;
                                background: #ffffff;
                                vertical-align: middle;
                                line-height: 1.25;
                            }}

                            .tabela-executiva-009717 tbody tr:nth-child(even) td {{
                                background: #f8fafc;
                            }}

                            .tabela-executiva-009717 tbody tr:hover td {{
                                background: #eef7fb;
                            }}

                            .tabela-executiva-009717 tbody td:last-child {{
                                border-right: none;
                            }}

                            .tabela-executiva-009717 .fonte-009717 {{
                                width: 18%;
                                white-space: nowrap;
                                overflow: hidden;
                                text-overflow: ellipsis;
                            }}

                            .tabela-executiva-009717 .credor-009717 {{
                                width: 24%;
                                font-weight: 600;
                                color: #18384f;
                                white-space: nowrap;
                                overflow: hidden;
                                text-overflow: ellipsis;
                            }}

                            .tabela-executiva-009717 .natureza-009717 {{
                                width: 10%;
                                text-align: center;
                                font-variant-numeric: tabular-nums;
                            }}

                            .tabela-executiva-009717 .valor-009717 {{
                                width: 14%;
                                text-align: right;
                                white-space: nowrap;
                                color: #0b5f7a;
                                font-weight: 800;
                                font-variant-numeric: tabular-nums;
                            }}

                            .tabela-executiva-009717 .ob-009717 {{
                                width: 24%;
                                white-space: nowrap;
                                overflow: hidden;
                                text-overflow: ellipsis;
                                font-size: 11px;
                                color: #374151;
                            }}

                            .tabela-executiva-009717 .status-cell-009717 {{
                                width: 10%;
                                text-align: center;
                                white-space: nowrap;
                            }}

                            .status-009717 {{
                                display: inline-flex;
                                align-items: center;
                                justify-content: center;
                                gap: 4px;
                                min-width: 92px;
                                padding: 5px 9px;
                                border-radius: 999px;
                                font-size: 11px;
                                font-weight: 800;
                            }}

                            .status-pago-009717 {{
                                color: #08743f;
                                background: #e8f7ef;
                                border: 1px solid #bde8cf;
                            }}

                            .status-nao-pago-009717 {{
                                color: #b42318;
                                background: #fff0ef;
                                border: 1px solid #f3c3bf;
                            }}

                            .status-retencao-009717 {{
                                color: #8a5a00;
                                background: #fff7df;
                                border: 1px solid #f1dea2;
                            }}

                            .status-neutro-009717 {{
                                color: #52606d;
                                background: #f3f5f7;
                                border: 1px solid #dbe1e5;
                            }}
                        </style>

                        <div class="tabela-executiva-wrap-009717">
                            <table class="tabela-executiva-009717">
                                <colgroup>
                                    <col style="width:18%">
                                    <col style="width:24%">
                                    <col style="width:10%">
                                    <col style="width:14%">
                                    <col style="width:24%">
                                    <col style="width:10%">
                                </colgroup>
                                <thead>
                                    <tr>
                                        <th>Fonte</th>
                                        <th>Credor</th>
                                        <th style="text-align:center;">Natureza</th>
                                        <th style="text-align:right;">Valor</th>
                                        <th>OB</th>
                                        <th style="text-align:center;">Status OB</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {''.join(linhas_html_009717)}
                                </tbody>
                            </table>
                        </div>
                        """
                    )

                    # Remove qualquer indentação restante, inclusive das linhas
                    # HTML geradas dinamicamente, para o Markdown não interpretar
                    # <tr>/<td> como bloco de código.
                    html_tabela_009717 = "\n".join(
                        linha.lstrip()
                        for linha in html_tabela_009717.splitlines()
                    )

                    st.markdown(
                        html_tabela_009717,
                        unsafe_allow_html=True,
                    )
        else:
            st.info(
                "Ainda não há dados do relatório para exibir. "
                "Clique em 'Atualizar Relatório'."
            )

elif st.session_state["tela_atual"] == "Planejar Priorização":
    st.markdown(
        "<h2 class='titulo-pagina'>🎯 Priorização da Semana — Exercício 2026</h2>",
        unsafe_allow_html=True,
    )

    base_nl_espelho = carregar_base_nl_espelho_planejamento()
    if base_nl_espelho.empty:
        st.warning("Não foi possível carregar a base de Liquidações para o planejamento.")
        st.stop()
    dados_planejamento = pd.DataFrame(
        {
            "Mês": base_nl_espelho["Competencia"],
            "Data_DT": base_nl_espelho["Data_DT"],
            "Fonte": base_nl_espelho["Fonte_Relacao"],
            "Grupo": base_nl_espelho["Grupo_Classificado"],
            "Objeto": base_nl_espelho["Objeto_Relacao"],
            "Status": base_nl_espelho["Status_Filtro"],
            "Tipo de NL": base_nl_espelho["Tipo_NL_Filtro"],
            "Credor": base_nl_espelho["Credor_NL"],
            "Número NL": base_nl_espelho["Numero_NL"],
            "Valor_Executado": base_nl_espelho["Valor_Total_Limpo"],
        }
    )

    # A pactuação deste fluxo é exclusiva da Fonte 500. O usuário informa um
    # valor mensal por grupo e organiza, com base nas NL, as prioridades das 4 semanas.
    def ordenar_mes_planejamento(valor):
        try:
            return datetime.datetime.strptime(str(valor), "%m/%Y")
        except ValueError:
            return datetime.datetime.max

    base_fonte_500 = dados_planejamento[
        dados_planejamento["Fonte"].astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False)
    ].copy()
    if base_fonte_500.empty:
        st.warning("Não há liquidações identificadas para a Fonte 500.")
        st.stop()

    # O planejamento deste exercício trabalha exclusivamente com competências de 2026.
    meses_500 = sorted(
        [mes for mes in base_fonte_500["Mês"].dropna().unique() if str(mes).endswith("/2026")],
        key=ordenar_mes_planejamento,
    )
    grupos_500 = sorted(base_fonte_500["Grupo"].dropna().astype(str).unique())
    tipos_nl_500 = sorted(base_fonte_500["Tipo de NL"].dropna().astype(str).unique())
    if not meses_500:
        st.warning("Não há competências de 2026 identificadas na Fonte 500.")
        st.stop()
    grupos_validos_planejamento = [
        grupo for grupo in st.session_state["mem_plan_grupos"] if grupo in grupos_500
    ]
    if "grupos_planejamento_nl" not in st.session_state:
        st.session_state["grupos_planejamento_nl"] = grupos_validos_planejamento
    tipos_validos_planejamento = [
        tipo for tipo in st.session_state["mem_plan_tipos_nl"] if tipo in tipos_nl_500
    ]
    if "tipos_nl_planejamento_nl" not in st.session_state:
        st.session_state["tipos_nl_planejamento_nl"] = tipos_validos_planejamento
    # O seletor abaixo controla somente o mês em que a programação será feita.
    # Para que os totais sejam idênticos aos da tela de NL, a disponibilidade
    # considera todo o histórico da Fonte 500, inclusive NL de competências
    # anteriores (por exemplo, RPP de 2025 liquidado no exercício atual).
    meses_aplicados = sorted(
        base_fonte_500["Mês"].dropna().astype(str).unique(),
        key=ordenar_mes_planejamento,
    )
    # A programação pode ser feita para o mês mais atual da consulta ou para
    # qualquer mês posterior do exercício, mesmo antes de existirem NL nele.
    ultimo_mes_fonte = max(ordenar_mes_planejamento(mes) for mes in meses_aplicados)
    meses_programacao = [
        f"{numero_mes:02d}/2026"
        for numero_mes in range(ultimo_mes_fonte.month, 13)
    ]
    if (
        "mes_planejamento_semana_nl" not in st.session_state
        or st.session_state["mes_planejamento_semana_nl"] not in meses_programacao
    ):
        mes_memoria = st.session_state["mem_plan_mes_programacao"]
        st.session_state["mes_planejamento_semana_nl"] = (
            mes_memoria if mes_memoria in meses_programacao else meses_programacao[0]
        )
    st.sidebar.markdown("#### 🎯 Planejamento semanal — Fonte 500")
    grupos_consulta = st.sidebar.multiselect(
        "Grupos",
        grupos_500,
        key="grupos_planejamento_nl",
        placeholder="Todos os grupos",
        on_change=sincronizar_filtro,
        args=("mem_plan_grupos", "grupos_planejamento_nl"),
    )
    tipos_nl_consulta = st.sidebar.multiselect(
        "Tipos de NL",
        tipos_nl_500,
        key="tipos_nl_planejamento_nl",
        placeholder="Todos os tipos de NL",
        on_change=sincronizar_filtro,
        args=("mem_plan_tipos_nl", "tipos_nl_planejamento_nl"),
    )
    mes_planejado = st.sidebar.selectbox(
        "Mês para programar as semanas",
        meses_programacao,
        key="mes_planejamento_semana_nl",
        on_change=sincronizar_filtro,
        args=("mem_plan_mes_programacao", "mes_planejamento_semana_nl"),
    )
    grupos_aplicados = grupos_consulta if grupos_consulta else grupos_500
    tipos_nl_aplicados = tipos_nl_consulta if tipos_nl_consulta else tipos_nl_500
    grupo_planejado = " + ".join(grupos_aplicados)

    # Resumo sem limitar pelo mês: é o mesmo recorte exibido na tela NL quando
    # Fonte 500 e o grupo selecionado estão ativos. O mês fica só para planejar semanas.
    base_resumo_nl = base_fonte_500[
        (base_fonte_500["Grupo"].astype(str).isin(grupos_aplicados))
        & (base_fonte_500["Tipo de NL"].astype(str).isin(tipos_nl_aplicados))
    ].copy()
    # A consulta de grupos filtra apenas o catálogo de objetos. O detalhamento
    # preserva todas as NLs da Fonte 500, inclusive RPP/DEA de exercícios
    # anteriores, para que uma troca GD3/GD4 não apague itens já programados.
    base_detalhamento_nl = base_fonte_500.copy()
    # Os objetos das semanas vêm do recorte atual da NL. O mês de programação
    # identifica a pactuação, inclusive quando for um mês futuro.
    base_selecionada = base_resumo_nl.copy()
    valor_total_resumo = float(base_resumo_nl["Valor_Executado"].sum())
    valor_gd3_mes_500 = float(
        base_resumo_nl.loc[base_resumo_nl["Grupo"].astype(str).str.upper() == "GD3", "Valor_Executado"].sum()
    )
    valor_gd4_mes_500 = float(
        base_resumo_nl.loc[base_resumo_nl["Grupo"].astype(str).str.upper() == "GD4", "Valor_Executado"].sum()
    )
    qtd_liquidacoes_resumo = int(len(base_resumo_nl))

    def quadro_resumo_planejamento(df, coluna, titulo):
        resumo = df.groupby(coluna, dropna=False)["Valor_Executado"].sum().sort_values(ascending=False)
        linhas = "".join(
            f"<tr><td style='padding:7px;'>{item}</td><td style='padding:7px; text-align:center;'>{formatar_brl(valor)}</td></tr>"
            for item, valor in resumo.items()
        )
        return f"""
        <div style='border:1px solid #cbd5e1; border-radius:7px; overflow:hidden; margin-bottom:12px;'>
          <table style='width:100%; border-collapse:collapse; font-size:13px;'>
            <thead><tr style='background:#dbe7f3; color:#002b49;'>
              <th style='padding:7px; text-align:left;'>{titulo}</th>
              <th style='padding:7px; text-align:center;'>SOMA DE VALOR</th>
            </tr></thead>
            <tbody>{linhas}
              <tr style='background:#e8f0f7; border-top:2px solid #2d6a9f; font-weight:700;'>
                <td style='padding:7px;'>TOTAL GERAL</td>
                <td style='padding:7px; text-align:center;'>{formatar_brl(resumo.sum())}</td>
              </tr>
            </tbody>
          </table>
        </div>"""

    st.markdown("#### Liquidações disponíveis para priorização")
    st.caption(
        f"Fonte 500 | Meses: {', '.join(meses_aplicados)} | Grupos: {', '.join(grupos_aplicados)}. "
        "Selecione no painel NL os mesmos filtros para conferir os valores."
    )
    card_qtd, card_total, card_gd3, card_gd4 = st.columns(4)
    with card_qtd:
        st.markdown(
            f"""<div class='metric-card'>
                <p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>QTD DE LIQUIDAÇÕES</p>
                <h3 style='color:#002b49; margin:5px 0;'>{f'{qtd_liquidacoes_resumo:,}'.replace(',', '.')}</h3>
                <p style='color:#28a745; font-size:11px; margin:0;'>📋 Documentos NL</p>
            </div>""",
            unsafe_allow_html=True,
        )
    with card_total:
        st.markdown(
            f"""<div class='metric-card'>
                <p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>VALOR TOTAL</p>
                <h3 style='color:#028090; margin:5px 0;'>{formatar_brl(valor_total_resumo)}</h3>
                <p style='color:#6c757d; font-size:11px; margin:0;'>Total Liquidado</p>
            </div>""",
            unsafe_allow_html=True,
        )
    with card_gd3:
        st.markdown(
            f"""<div class='metric-card'>
                <p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>VALOR TOTAL GD3</p>
                <h3 style='color:#f77f00; margin:5px 0;'>{formatar_brl(valor_gd3_mes_500)}</h3>
                <p style='color:#6c757d; font-size:11px; margin:0;'>Grupo GD3</p>
            </div>""",
            unsafe_allow_html=True,
        )
    with card_gd4:
        st.markdown(
            f"""<div class='metric-card'>
                <p style='color:#6c757d; font-size:11px; font-weight:bold; margin:0;'>VALOR TOTAL GD4</p>
                <h3 style='color:#2563eb; margin:5px 0;'>{formatar_brl(valor_gd4_mes_500)}</h3>
                <p style='color:#6c757d; font-size:11px; margin:0;'>Grupo GD4</p>
            </div>""",
            unsafe_allow_html=True,
        )

    # Respiro visual entre os indicadores e os gráficos de apoio.
    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

    # Painel de apoio à reunião: leitura visual do status e do tipo de NL,
    # usando o espaço antes ocupado por duas tabelas resumidas.
    coluna_resumos, coluna_historico = st.columns([1, 1.55], gap="large")
    with coluna_resumos:
        st.markdown(
            """
            <style>
            div[data-testid="stPlotlyChart"] {
                background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
                border: 1px solid #d9e3ed;
                border-radius: 12px;
                box-shadow: 0 8px 18px rgba(15, 42, 68, 0.12);
                box-sizing: border-box;
                overflow: hidden;
                padding: 10px 10px 6px;
                margin: 0 0 14px;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        resumo_tipo_grafico = (
            base_resumo_nl.groupby("Tipo de NL", dropna=False, as_index=False)["Valor_Executado"]
            .sum()
            .sort_values("Valor_Executado", ascending=False)
        )
        resumo_tipo_grafico["Tipo de NL"] = resumo_tipo_grafico["Tipo de NL"].fillna("Não informado")
        def rotulo_valor_compacto(valor):
            """Rótulo curto para não cortar valores no gráfico horizontal."""
            valor = float(valor or 0)
            if abs(valor) >= 1_000_000:
                return f"R$ {valor / 1_000_000:.1f} mi".replace(".", ",")
            if abs(valor) >= 1_000:
                return f"R$ {valor / 1_000:.0f} mil".replace(".", ",")
            return f"R$ {valor:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

        total_tipo_grafico = float(resumo_tipo_grafico["Valor_Executado"].sum())
        resumo_tipo_grafico["Percentual"] = (
            resumo_tipo_grafico["Valor_Executado"] / total_tipo_grafico * 100
            if total_tipo_grafico > 0
            else 0.0
        )
        resumo_tipo_grafico["Rótulo"] = resumo_tipo_grafico.apply(
            lambda linha: (
                f"{rotulo_valor_compacto(linha['Valor_Executado'])}  •  "
                f"{float(linha['Percentual']):.1f}%".replace(".", ",")
            ),
            axis=1,
        )

        grafico_tipo = px.bar(
            resumo_tipo_grafico,
            x="Valor_Executado",
            y="Tipo de NL",
            orientation="h",
            text="Rótulo",
            color="Tipo de NL",
            color_discrete_map={
                "Orçamentária": "#0b4f6c",
                "RPP": "#2a7f8e",
                "RPNP": "#8a6d3b",
                "Não informado": "#64748b",
            },
            custom_data=["Percentual"],
        )
        grafico_tipo.update_traces(
            textposition="outside",
            cliponaxis=False,
            width=0.48,
            marker_line_width=0,
            textfont=dict(size=12, color="#475569"),
            hovertemplate=(
                "<b>%{y}</b><br>"
                "Valor: R$ %{x:,.2f}<br>"
                "Participação: %{customdata[0]:.1f}%<extra></extra>"
            ),
        )
        grafico_tipo.update_layout(
            title={
                "text": "Valores por tipo de NL",
                "x": 0.0,
                "xanchor": "left",
                "font": {"size": 15, "color": "#002b49"},
            },
            margin=dict(l=18, r=145, t=48, b=12),
            height=220,
            showlegend=False,
            bargap=0.38,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            hoverlabel=dict(
                bgcolor="#ffffff",
                bordercolor="#d7e1ea",
                font_color="#0f2a44",
                font_size=12,
            ),
            xaxis=dict(
                visible=False,
                showgrid=False,
                zeroline=False,
                fixedrange=True,
            ),
            yaxis=dict(
                title=None,
                categoryorder="array",
                categoryarray=resumo_tipo_grafico["Tipo de NL"].tolist(),
                autorange="reversed",
                showgrid=False,
                ticks="",
                tickfont=dict(size=12, color="#64748b"),
                fixedrange=True,
            ),
            separators=",.",
        )
        st.plotly_chart(
            grafico_tipo,
            use_container_width=True,
            key="grafico_tipo_planejamento",
            config={
                "displayModeBar": False,
                "displaylogo": False,
                "responsive": True,
                "scrollZoom": False,
            },
        )

        st.markdown(
            "<div style='border-top:1px solid #dbe4ee; margin:12px 8px 14px;'></div>",
            unsafe_allow_html=True,
        )

        # O status é uma informação de conferência: a tabela é mais direta e
        # legível que o gráfico de pizza para a tomada de decisão.
        st.markdown("##### Status das liquidações")
        st.html(quadro_resumo_planejamento(base_resumo_nl, "Status", "STATUS"))

    with coluna_historico:
        st.markdown("##### Histórico de pagamentos consolidados — Fonte 500")
        ordem_meses_500 = ["Jan/2026", "Fev/2026", "Mar/2026", "Abr/2026", "Mai/2026", "Jun/2026", "Jul/2026", "Ago/2026"]
        historico_ob_500 = carregar_historico_ob_fonte_500()
        if historico_ob_500.empty:
            st.info("Ainda não há pagamentos consolidados da Fonte 500 na base de OB.")
        else:
            resumo_historico_500 = (
                historico_ob_500.groupby("Mês", as_index=False)
                .agg(**{"Qtd. Docs": ("Número", "count"), "Total Pago": ("Valor_Pago", "sum")})
                .set_index("Mês")
                .reindex(ordem_meses_500, fill_value=0.0)
                .reset_index()
            )
            linhas_historico = ""
            for _, linha in resumo_historico_500.iterrows():
                qtd_docs = f"{int(linha['Qtd. Docs']):,}".replace(",", ".")
                linhas_historico += (
                    f"<tr style='border-bottom:1px solid #f1f5f9;'>"
                    f"<td style='padding:8px 12px; text-align:left; color:#334155;'>{linha['Mês']}</td>"
                    f"<td style='padding:8px 12px; text-align:center; color:#334155;'>{qtd_docs}</td>"
                    f"<td style='padding:8px 12px; text-align:right; color:#0f172a; font-weight:600;'>{formatar_brl(linha['Total Pago'])}</td>"
                    f"</tr>"
                )
            total_documentos_500 = int(resumo_historico_500["Qtd. Docs"].sum())
            total_pago_500 = float(resumo_historico_500["Total Pago"].sum())
            # st.html evita que o Markdown acrescente um parágrafo vazio após a
            # tabela, mantendo a borda inferior encostada no Total Geral.
            st.html(
                f"""
                <div style='border:1px solid #e2e8f0; border-radius:8px; overflow:hidden; background:#fff; line-height:normal; margin:0; padding:0; display:block;'>
                  <table style='width:100%; height:auto !important; border-collapse:collapse; border-spacing:0; margin:0 !important; padding:0; display:table !important;'>
                    <thead><tr style='background:#f8fafc; border-bottom:1px solid #e2e8f0;'>
                      <th style='padding:10px 12px; color:#475569; font-size:11px; text-align:left;'>MÊS DE REFERÊNCIA</th>
                      <th style='padding:10px 12px; color:#475569; font-size:11px; text-align:center;'>QTD. DOCS</th>
                      <th style='padding:10px 12px; color:#475569; font-size:11px; text-align:right;'>TOTAL PAGO</th>
                    </tr></thead>
                    <tbody>{linhas_historico}</tbody>
                    <tfoot><tr style='background:#f8fafc; border-top:2px solid #002b49; font-weight:700;'>
                      <td style='padding:10px 12px; color:#002b49; border-bottom:0;'>📊 TOTAL GERAL</td>
                      <td style='padding:10px 12px; text-align:center; color:#002b49; border-bottom:0;'>{f'{total_documentos_500:,}'.replace(',', '.')}</td>
                      <td style='padding:10px 12px; text-align:right; color:#002b49; border-bottom:0;'>{formatar_brl(total_pago_500)}</td>
                    </tr></tfoot>
                  </table>
                </div>
                """,
            )

    objetos_disponiveis = (
        base_selecionada.groupby(["Grupo", "Objeto"], as_index=False)
        .agg(
            **{
                "Qtd. NL": ("Número NL", lambda serie: serie[serie.astype(str).str.strip() != ""].nunique()),
                "Competências": (
                    "Mês",
                    lambda serie: ", ".join(
                        sorted(
                            {str(valor) for valor in serie.dropna() if str(valor).strip()},
                            key=ordenar_mes_planejamento,
                        )
                    ),
                ),
                "Tipos de NL": (
                    "Tipo de NL",
                    lambda serie: ", ".join(
                        sorted({str(valor) for valor in serie.dropna() if str(valor).strip()})
                    ),
                ),
                "Valor_Executado": ("Valor_Executado", "sum"),
            }
        )
        .sort_values("Valor_Executado", ascending=False)
    )

    # A lista de objetos vem antes da programação: é a referência para a reunião.
    st.divider()
    st.markdown("### 1. Objetos disponíveis para programar")
    st.caption(
        f"Competência de programação: {mes_planejado} | Grupos: {grupo_planejado}. "
        "A disponibilidade inclui RPP/DEA e demais liquidações de exercícios anteriores da Fonte 500."
    )
    objetos_referencia = objetos_disponiveis.copy()
    objetos_referencia.insert(
        0,
        "Código",
        objetos_referencia["Objeto"].fillna("").astype(str).str.extract(r"^\s*([0-9]+)", expand=False).fillna(""),
    )
    objetos_referencia = objetos_referencia.rename(columns={"Valor_Executado": "Disponível na NL"})
    # Exibição fixa no padrão brasileiro: ponto para milhares e vírgula para centavos.
    objetos_referencia["Disponível na NL"] = objetos_referencia["Disponível na NL"].apply(formatar_brl)
    objetos_visualizacao = objetos_referencia[[
        "Grupo", "Código", "Objeto", "Competências", "Tipos de NL", "Qtd. NL", "Disponível na NL"
    ]].copy()
    objetos_visualizacao["Qtd. NL"] = objetos_visualizacao["Qtd. NL"].fillna(0).astype(int).astype(str)
    objetos_visualizacao = objetos_visualizacao.rename(
        columns={"Objeto": "Objeto da despesa"}
    )
    renderizar_tabela_priorizacao(
        objetos_visualizacao,
        colunas_centralizadas={"Grupo", "Código", "Qtd. NL", "Disponível na NL"},
        altura_maxima=360,
        template_colunas=["7%", "7%", "34%", "13%", "15%", "7%", "17%"],
    )

    chave_planejamento = "FONTE_500"
    chave_prioridades = f"prioridades_{mes_planejado}_{chave_planejamento}"

    # Sincronização com a planilha de priorização. A leitura deixa de acontecer
    # apenas uma vez por sessão: o painel verifica periodicamente se a planilha
    # foi alterada fora do Streamlit e reconstrói o mês atual quando necessário.
    if "nonce_atualizacao_priorizacao" not in st.session_state:
        st.session_state["nonce_atualizacao_priorizacao"] = 0

    col_sync1, col_sync2 = st.columns([0.24, 0.76])
    with col_sync1:
        atualizar_planilha_agora = st.button(
            "🔄 Atualizar da planilha",
            key="atualizar_priorizacao_planilha",
            use_container_width=True,
        )
        if atualizar_planilha_agora:
            # O clique precisa ser uma recarga FORÇADA, não apenas uma checagem.
            # Limpa o cache local e usa um nonce único para obter nova resposta.
            carregar_priorizacao_google_sincronizada.clear()
            st.session_state["nonce_atualizacao_priorizacao"] = int(
                datetime.datetime.now().timestamp() * 1000
            )

    retorno_google = carregar_priorizacao_google_sincronizada(
        st.session_state["nonce_atualizacao_priorizacao"]
    )
    assinatura_remota = assinatura_priorizacao_remota(retorno_google)
    primeira_sincronizacao = "assinatura_priorizacao_remota" not in st.session_state
    assinatura_anterior = st.session_state.get("assinatura_priorizacao_remota", "")
    houve_alteracao_remota = bool(
        assinatura_anterior and assinatura_remota and assinatura_remota != assinatura_anterior
    )

    st.session_state["priorizacao_google_carregada"] = True
    st.session_state["priorizacao_google_erro"] = retorno_google.get("erro", "")

    if retorno_google.get("ok") is not False:
        # Canonicaliza os campos recebidos. Isso evita que pequenas diferenças
        # no cabeçalho da planilha (Mes/Mês, Valor_Programado etc.) zerem o painel.
        programacao_google_api = canonicalizar_programacao_priorizacao(
            retorno_google.get("programacao", [])
        )
        pactuados_google_api = canonicalizar_pactuados_priorizacao(
            retorno_google.get("pactuados", [])
        )
        st.session_state["programacao_google_salva"] = programacao_google_api.copy()

        if pactuados_google_api.empty:
            st.session_state["pactuados_fonte500"] = pd.DataFrame(
                columns=["Mês", "Grupo", "Pactuado Mensal", "Pactuado Semana 1", "Pactuado Semana 2", "Pactuado Semana 3", "Pactuado Semana 4"]
            )
        else:
            pactuados_google = pactuados_google_api.rename(
                columns={
                    "Mes": "Mês",
                    "Valor_Pactuado": "Pactuado Mensal",
                    "Pactuado_Semana_1": "Pactuado Semana 1",
                    "Pactuado_Semana_2": "Pactuado Semana 2",
                    "Pactuado_Semana_3": "Pactuado Semana 3",
                    "Pactuado_Semana_4": "Pactuado Semana 4",
                }
            ).copy()
            # A aba é exclusiva da Fonte 500; se Grupo vier vazio, assume a chave
            # institucional usada pelo painel.
            pactuados_google.loc[
                pactuados_google["Grupo"].fillna("").astype(str).str.strip() == "", "Grupo"
            ] = chave_planejamento
            st.session_state["pactuados_fonte500"] = pactuados_google[
                ["Mês", "Grupo", "Pactuado Mensal", "Pactuado Semana 1", "Pactuado Semana 2", "Pactuado Semana 3", "Pactuado Semana 4"]
            ].copy()

        # Se a planilha mudou por fora do painel, invalida apenas o recorte do
        # mês atual. Ele será reconstruído logo abaixo com os dados mais novos.
        if houve_alteracao_remota or primeira_sincronizacao or atualizar_planilha_agora:
            # No clique manual, sempre descarta a cópia local do mês e reconstrói
            # as quatro semanas com exatamente o que a API acabou de retornar.
            st.session_state.pop(chave_prioridades, None)
            for semana_sync in ["Semana 1", "Semana 2", "Semana 3", "Semana 4"]:
                detalhe_sync = f"detalhe_nl_{chave_prioridades}_{semana_sync}"
                st.session_state.pop(detalhe_sync, None)
                st.session_state.pop(f"{detalhe_sync}_assinatura", None)
                st.session_state.pop(f"{detalhe_sync}_restaurado_planilha", None)
            # Remove estados de editores antigos desse mês para que checkboxes
            # de uma versão anterior não sobrescrevam o que veio da planilha.
            prefixo_editor = f"editor_detalhe_nl_{chave_prioridades}_"
            for chave_estado in list(st.session_state.keys()):
                if str(chave_estado).startswith(prefixo_editor):
                    st.session_state.pop(chave_estado, None)
            st.session_state.pop(f"pactuado_{mes_planejado}_{chave_planejamento}", None)
            st.session_state["priorizacao_planilha_atualizada_em"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        st.session_state["assinatura_priorizacao_remota"] = assinatura_remota

    with col_sync2:
        if st.session_state.get("priorizacao_google_erro"):
            st.warning("Não foi possível atualizar a priorização da planilha vinculada.")
        elif atualizar_planilha_agora:
            qtd_programacao_api = len(retorno_google.get("programacao", []) or [])
            qtd_pactuados_api = len(retorno_google.get("pactuados", []) or [])
            prog_diag = canonicalizar_programacao_priorizacao(retorno_google.get("programacao", []))
            pact_diag = canonicalizar_pactuados_priorizacao(retorno_google.get("pactuados", []))
            mes_diag = normalizar_mes_priorizacao(mes_planejado)
            prog_mes_diag = prog_diag[
                (prog_diag["Mes"] == mes_diag)
                & (prog_diag["Fonte"].astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False))
            ].copy()
            pact_mes_diag = pact_diag[
                (pact_diag["Mes"] == mes_diag)
                & (pact_diag["Fonte"].astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False))
            ].copy()
            valor_pact_diag = float(pact_mes_diag["Valor_Pactuado"].sum()) if not pact_mes_diag.empty else 0.0
            total_prog_diag = float(prog_mes_diag["Valor_Programado"].sum()) if not prog_mes_diag.empty else 0.0
            semanas_diag = ", ".join(
                f"{semana}: {len(grupo)} NL / {formatar_brl(grupo['Valor_Programado'].sum())}"
                for semana, grupo in prog_mes_diag.groupby("Semana", sort=True)
            ) or "nenhuma semana encontrada"
            st.success(
                f"Sincronização concluída para {mes_diag}: API total {qtd_programacao_api} programação / "
                f"{qtd_pactuados_api} pactuado(s). No mês selecionado: {len(prog_mes_diag)} NL, "
                f"{formatar_brl(total_prog_diag)} programado e {formatar_brl(valor_pact_diag)} pactuado. "
                f"{semanas_diag}."
            )
        elif houve_alteracao_remota:
            st.success(
                "Alterações feitas diretamente na planilha foram reconhecidas e aplicadas ao painel."
            )
        else:
            st.caption(
                "A priorização é sincronizada com a planilha a cada nova interação (cache de até 20 s). "
                "Use o botão ao lado para forçar a leitura imediata."
            )

    if "pactuados_fonte500" not in st.session_state:
        st.session_state["pactuados_fonte500"] = pd.DataFrame(
            columns=["Mês", "Grupo", "Pactuado Mensal", "Pactuado Semana 1", "Pactuado Semana 2", "Pactuado Semana 3", "Pactuado Semana 4"]
        )

    mes_planejado_normalizado = normalizar_mes_priorizacao(mes_planejado)
    colunas_pactuado_estado = [
        "Mês", "Grupo", "Pactuado Mensal",
        "Pactuado Semana 1", "Pactuado Semana 2", "Pactuado Semana 3", "Pactuado Semana 4",
    ]
    for coluna in colunas_pactuado_estado:
        if coluna not in st.session_state["pactuados_fonte500"].columns:
            st.session_state["pactuados_fonte500"][coluna] = 0.0 if "Pactuado" in coluna else ""

    tabela_pactuados_estado = st.session_state["pactuados_fonte500"].copy()
    tabela_pactuados_estado["Mês"] = tabela_pactuados_estado["Mês"].apply(normalizar_mes_priorizacao)
    tabela_pactuados_estado["Grupo"] = (
        tabela_pactuados_estado["Grupo"].fillna("").astype(str).str.strip().str.upper()
    )
    chave_pactuado = (
        (tabela_pactuados_estado["Mês"] == mes_planejado_normalizado)
        & (tabela_pactuados_estado["Grupo"].isin(["", chave_planejamento.upper()]))
    )
    registro_pactuado = (
        tabela_pactuados_estado.loc[chave_pactuado].iloc[0]
        if chave_pactuado.any() else None
    )
    pactuado_atual = (
        float(registro_pactuado.get("Pactuado Mensal", 0.0))
        if registro_pactuado is not None
        else (54_000_000.0 if mes_planejado == "09/2026" else 0.0)
    )
    # O pactuado semanal não é mais digitado manualmente.
    # Ele será calculado automaticamente a partir do que já está salvo na
    # programação de cada semana (histórico) e do que o usuário adicionar
    # na tela. Assim, trocar de módulo não faz os valores desaparecerem.
    st.markdown("### 2. Defina o teto do mês")
    entrada_pactuado, descricao = st.columns([1, 2])
    with entrada_pactuado:
        pactuado_mensal = st.number_input(
            "Pactuado mensal — Fonte 500",
            min_value=0.0, value=pactuado_atual, step=1000.0, format="%.2f",
            key=f"pactuado_{mes_planejado}_{chave_planejamento}",
        )
    with descricao:
        st.caption(
            "Informe somente o teto mensal. O pactuado de cada semana é calculado "
            "automaticamente pelos objetos/NLs que já estão registrados na semana "
            "ou que forem adicionados durante o planejamento."
        )

    # Atualiza apenas o teto mensal neste ponto. Os quatro pactuados semanais
    # são preenchidos mais abaixo, depois que as semanas forem reconstruídas.
    valores_registro_pactuado = {
        "Mês": mes_planejado_normalizado,
        "Grupo": chave_planejamento,
        "Pactuado Mensal": float(pactuado_mensal),
    }
    if chave_pactuado.any():
        indice_real = tabela_pactuados_estado.loc[chave_pactuado].index[0]
        for coluna, valor in valores_registro_pactuado.items():
            st.session_state["pactuados_fonte500"].loc[indice_real, coluna] = valor
    else:
        novo_registro = {
            "Mês": mes_planejado_normalizado,
            "Grupo": chave_planejamento,
            "Pactuado Mensal": float(pactuado_mensal),
            "Pactuado Semana 1": 0.0,
            "Pactuado Semana 2": 0.0,
            "Pactuado Semana 3": 0.0,
            "Pactuado Semana 4": 0.0,
        }
        st.session_state["pactuados_fonte500"] = pd.concat(
            [st.session_state["pactuados_fonte500"], pd.DataFrame([novo_registro])],
            ignore_index=True,
        )

    if chave_prioridades not in st.session_state:
        semanas_padrao = ["Semana 1", "Semana 2", "Semana 3", "Semana 4"]
        st.session_state[chave_prioridades] = {
            semana: pd.DataFrame(columns=["Objeto", "Valor Prioridade", "Observação"])
            for semana in semanas_padrao
        }

        # Reconstrói a tela a partir das NLs detalhadas salvas. Cada NL salva
        # continua marcada para pagamento; as demais do objeto ficam excluídas.
        # Isso permite reabrir o sistema sem duplicar valores.
        programacao_google = st.session_state.get("programacao_google_salva", pd.DataFrame()).copy()
        if isinstance(programacao_google, pd.DataFrame) and not programacao_google.empty:
            # A resposta já foi canonicalizada na sincronização; renomeia apenas
            # para os títulos usados internamente na interface.
            programacao_google = programacao_google.rename(
                columns={
                    "Mes": "Mês de programação",
                    "Numero_NL": "Número NL",
                    "Tipo_NL": "Tipo de NL",
                    "Valor_Programado": "Valor programado",
                    "Observacao": "Observação",
                }
            )
            for coluna in [
                "Mês de programação", "Semana", "Objeto", "Número NL", "Credor",
                "Valor programado", "Grupo", "Tipo de NL", "Observação",
            ]:
                if coluna not in programacao_google.columns:
                    programacao_google[coluna] = 0.0 if coluna == "Valor programado" else ""
            fonte_programacao = (
                programacao_google["Fonte"]
                if "Fonte" in programacao_google.columns
                else pd.Series("500", index=programacao_google.index)
            )
            programacao_google["__Mes_Normalizado"] = programacao_google["Mês de programação"].apply(
                normalizar_mes_priorizacao
            )
            programacao_google["__Semana_Normalizada"] = programacao_google["Semana"].apply(
                normalizar_semana_priorizacao
            )
            programacao_google["__Numero_NL_Normalizado"] = programacao_google["Número NL"].apply(
                normalizar_numero_nl_priorizacao
            )
            registros_mes = programacao_google[
                (programacao_google["__Mes_Normalizado"] == normalizar_mes_priorizacao(mes_planejado))
                & (fonte_programacao.astype(str).str.contains("500", na=False))
            ].copy()
            for semana in semanas_padrao:
                registros_semana = registros_mes[
                    registros_mes["__Semana_Normalizada"] == semana
                ].copy()
                if registros_semana.empty:
                    continue
                linhas_restauradas = []
                for objeto, grupo_objeto in registros_semana.groupby("Objeto", dropna=False, sort=False):
                    valor_restaurado = converter_valor_monetario(
                        grupo_objeto.get("Valor programado", pd.Series(dtype=float))
                    ).fillna(0.0).sum()
                    observacao_restaurada = str(
                        grupo_objeto.get("Observação", pd.Series([""])).fillna("").iloc[0]
                    )
                    linhas_restauradas.append({
                        "Objeto": str(objeto),
                        "Valor Prioridade": float(valor_restaurado),
                        "Observação": observacao_restaurada,
                    })
                st.session_state[chave_prioridades][semana] = pd.DataFrame(linhas_restauradas)

                objetos_restaurados = [linha["Objeto"] for linha in linhas_restauradas]

                # A programação salva na planilha é a fonte de verdade do histórico.
                # Não dependemos de a NL ainda existir na base atual para restaurar
                # Semana 1/2 (ou qualquer semana já registrada).
                detalhe_remoto = registros_semana[[
                    "Objeto", "Número NL", "Credor", "Valor programado", "Grupo", "Tipo de NL"
                ]].copy()
                detalhe_remoto = detalhe_remoto.rename(columns={"Valor programado": "Valor da NL"})
                detalhe_remoto["Valor da NL"] = converter_valor_monetario(detalhe_remoto["Valor da NL"])
                detalhe_remoto["Excluir da semana"] = False
                detalhe_remoto["__Numero_NL_Normalizado"] = detalhe_remoto["Número NL"].apply(
                    normalizar_numero_nl_priorizacao
                )

                # Acrescenta as NLs atualmente disponíveis para os mesmos objetos.
                # As que não estavam salvas naquela semana entram marcadas como
                # excluídas, preservando exatamente o histórico registrado.
                detalhe_atual = base_detalhamento_nl[
                    base_detalhamento_nl["Objeto"].astype(str).isin(objetos_restaurados)
                ][["Objeto", "Número NL", "Credor", "Valor_Executado", "Grupo", "Tipo de NL"]].copy()
                detalhe_atual = detalhe_atual.rename(columns={"Valor_Executado": "Valor da NL"})
                detalhe_atual["__Numero_NL_Normalizado"] = detalhe_atual["Número NL"].apply(
                    normalizar_numero_nl_priorizacao
                )
                chaves_remotas = set(
                    detalhe_remoto["Objeto"].astype(str) + "::" + detalhe_remoto["__Numero_NL_Normalizado"].astype(str)
                )
                chave_atual = detalhe_atual["Objeto"].astype(str) + "::" + detalhe_atual["__Numero_NL_Normalizado"].astype(str)
                detalhe_atual = detalhe_atual.loc[~chave_atual.isin(chaves_remotas)].copy()
                detalhe_atual["Excluir da semana"] = True

                detalhe_restaurado = pd.concat([detalhe_remoto, detalhe_atual], ignore_index=True)
                detalhe_restaurado["Valor da NL"] = detalhe_restaurado["Valor da NL"].apply(formatar_brl)
                detalhe_chave_restaurada = f"detalhe_nl_{chave_prioridades}_{semana}"
                st.session_state[detalhe_chave_restaurada] = detalhe_restaurado[[
                    "Objeto", "Número NL", "Credor", "Valor da NL", "Excluir da semana", "Grupo", "Tipo de NL"
                ]].copy()

                # A assinatura acompanha a base atual, mas o estado acima mantém
                # também as NLs históricas que já não existem mais nessa base.
                assinatura_base_atual = "|".join(sorted(
                    detalhe_atual["Objeto"].astype(str) + "::" + detalhe_atual["Número NL"].astype(str)
                ))
                assinatura_base_salva = "|".join(sorted(
                    base_detalhamento_nl[
                        base_detalhamento_nl["Objeto"].astype(str).isin(objetos_restaurados)
                    ]["Objeto"].astype(str) + "::" +
                    base_detalhamento_nl[
                        base_detalhamento_nl["Objeto"].astype(str).isin(objetos_restaurados)
                    ]["Número NL"].astype(str)
                ))
                st.session_state[f"{detalhe_chave_restaurada}_assinatura"] = assinatura_base_salva
                st.session_state[f"{detalhe_chave_restaurada}_restaurado_planilha"] = True

    # Um mesmo objeto pode aparecer em mais de um grupo na referência acima.
    # Para a priorização, o limite é a soma das NLs dos grupos selecionados.
    limites_objetos = (
        base_selecionada.groupby("Objeto", as_index=False)["Valor_Executado"].sum()
    )
    limite_por_objeto = limites_objetos.set_index("Objeto")["Valor_Executado"].to_dict()
    opcoes_objeto = limites_objetos.sort_values("Objeto")["Objeto"].tolist()

    def limpar_semana_planejamento(chave, semana):
        st.session_state[chave][semana] = pd.DataFrame(
            columns=["Objeto", "Valor Prioridade", "Observação"]
        )
        st.session_state.pop(f"editor_{chave}_{semana}", None)

    st.markdown("### 3. Monte as prioridades semanais")
    st.caption("Escolha os objetos e distribua o valor de pagamento. Você pode limpar uma semana inteira quando precisar refazer o cenário.")
    abas_semanais = st.tabs(["Semana 1", "Semana 2", "Semana 3", "Semana 4"])
    prioridades_semanais = []
    for aba, semana in zip(abas_semanais, ["Semana 1", "Semana 2", "Semana 3", "Semana 4"]):
        with aba:
            st.caption("Selecione os objetos prioritários e informe quanto será programado para pagamento nesta semana.")
            # Fluxo vertical: primeiro o objeto, depois seus credores e NLs.
            col_grade = st.container()
            col_detalhes = st.container()
            with col_grade:
                st.button(
                    "🗑️ Limpar semana",
                    key=f"limpar_{chave_prioridades}_{semana}",
                    on_click=limpar_semana_planejamento,
                    args=(chave_prioridades, semana),
                )
                escolha_col, adicionar_col = st.columns([5, 1])
                objeto_novo = escolha_col.selectbox(
                    "Adicionar objeto à semana",
                    opcoes_objeto,
                    index=None,
                    placeholder="Pesquise e selecione um objeto",
                    key=f"novo_objeto_{chave_prioridades}_{semana}",
                )
                if adicionar_col.button(
                    "Adicionar",
                    key=f"adicionar_objeto_{chave_prioridades}_{semana}",
                    use_container_width=True,
                ):
                    if not objeto_novo:
                        st.warning("Selecione um objeto antes de adicioná-lo à semana.")
                    else:
                        objetos_atuais = (
                            st.session_state[chave_prioridades][semana]["Objeto"]
                            .fillna("").astype(str).str.strip().tolist()
                        )
                        if objeto_novo in objetos_atuais:
                            st.info("Esse objeto já está incluído nesta semana.")
                        else:
                            valor_inicial = float(limite_por_objeto.get(objeto_novo, 0.0))
                            nova_linha = pd.DataFrame([{
                                "Objeto": objeto_novo,
                                "Valor Prioridade": valor_inicial,
                                "Observação": "",
                                "Disponível na NL": formatar_brl(valor_inicial),
                            }])
                            st.session_state[chave_prioridades][semana] = pd.concat(
                                [st.session_state[chave_prioridades][semana], nova_linha],
                                ignore_index=True,
                            )
                            st.rerun()
                tabela_semana = st.session_state[chave_prioridades][semana].copy()
                tabela_semana = tabela_semana.drop(columns=["Código"], errors="ignore")
                objetos_escolhidos = (
                    tabela_semana.get("Objeto", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
                )
                objetos_escolhidos = objetos_escolhidos[objetos_escolhidos != ""].unique().tolist()
                detalhe_chave = f"detalhe_nl_{chave_prioridades}_{semana}"

                # A grade abaixo substitui o data_editor plano: cada objeto é uma linha expansível
                # e traz, dentro dele, os credores e as NLs que poderão ser excluídas.
                if not objetos_escolhidos:
                    prioridade_editada = pd.DataFrame(columns=["Objeto", "Valor Prioridade", "Observação", "Disponível na NL"])
                    st.info("Adicione um objeto acima para montar a semana.")
                else:
                    detalhamento_nl = base_detalhamento_nl[
                        base_detalhamento_nl["Objeto"].isin(objetos_escolhidos)
                    ][["Objeto", "Número NL", "Credor", "Valor_Executado", "Grupo", "Tipo de NL"]].copy()
                    detalhamento_nl = detalhamento_nl.rename(columns={"Valor_Executado": "Valor da NL"})
                    detalhamento_nl = detalhamento_nl[
                        detalhamento_nl["Número NL"].fillna("").astype(str).str.strip() != ""
                    ].copy()
                    assinatura_detalhe = "|".join(sorted(
                        detalhamento_nl["Objeto"].astype(str) + "::" + detalhamento_nl["Número NL"].astype(str)
                    ))
                    detalhe_anterior = st.session_state.get(detalhe_chave)
                    restaurado_da_planilha = bool(
                        st.session_state.get(f"{detalhe_chave}_restaurado_planilha", False)
                    )
                    precisa_recriar_detalhe = (
                        detalhe_anterior is None
                        or "Objeto" not in detalhe_anterior.columns
                        or (
                            not restaurado_da_planilha
                            and st.session_state.get(f"{detalhe_chave}_assinatura") != assinatura_detalhe
                        )
                    )
                    if precisa_recriar_detalhe:
                        detalhe_anterior = detalhamento_nl[[
                            "Objeto", "Número NL", "Credor", "Valor da NL", "Grupo", "Tipo de NL"
                        ]].copy()
                        detalhe_anterior["Valor da NL"] = detalhe_anterior["Valor da NL"].apply(formatar_brl)
                        detalhe_anterior["Excluir da semana"] = False
                        st.session_state[detalhe_chave] = detalhe_anterior
                        st.session_state[f"{detalhe_chave}_assinatura"] = assinatura_detalhe

                    # Renderizar todas as NLs de todos os objetos a cada clique
                    # deixava a tela lenta. O cenário continua completo, mas o
                    # detalhamento pesado é aberto para um objeto por vez.
                    st.markdown("##### Objetos incluídos na semana")
                    chave_objeto_detalhe = f"objeto_detalhe_{chave_prioridades}_{semana}"
                    if st.session_state.get(chave_objeto_detalhe) not in objetos_escolhidos:
                        st.session_state[chave_objeto_detalhe] = objetos_escolhidos[0]
                    if len(objetos_escolhidos) == 1:
                        # Com um único objeto, a linha expansível abaixo já é
                        # suficiente; não repetimos o nome em um seletor.
                        objeto_detalhado = objetos_escolhidos[0]
                    else:
                        objeto_detalhado = st.selectbox(
                            "Detalhar objeto",
                            objetos_escolhidos,
                            key=chave_objeto_detalhe,
                        )

                    # O data_editor informa as alterações da caixa de exclusão no
                    # início do rerun. Aplicamos essas alterações antes de montar
                    # os títulos dos objetos, para que "A PROGRAMAR" já venha
                    # subtraído na mesma interação do usuário.
                    for indice_objeto, objeto_estado in enumerate([objeto_detalhado]):
                        indice_objeto = objetos_escolhidos.index(objeto_estado)
                        detalhes_estado = detalhe_anterior[
                            detalhe_anterior["Objeto"].astype(str) == str(objeto_estado)
                        ]
                        for indice_credor, (_, grupo_estado) in enumerate(
                            detalhes_estado.groupby("Credor", dropna=False, sort=False)
                        ):
                            chave_editor = f"editor_{detalhe_chave}_{indice_objeto}_{indice_credor}"
                            estado_editor = st.session_state.get(chave_editor, {})
                            linhas_alteradas = (
                                estado_editor.get("edited_rows", {})
                                if isinstance(estado_editor, dict) else {}
                            )
                            for indice_linha, alteracoes in linhas_alteradas.items():
                                if "Excluir da semana" not in alteracoes:
                                    continue
                                try:
                                    numero_nl = str(grupo_estado.iloc[int(indice_linha)]["Número NL"])
                                except (IndexError, KeyError, TypeError, ValueError):
                                    continue
                                mascara = (
                                    (detalhe_anterior["Objeto"].astype(str) == str(objeto_estado))
                                    & (detalhe_anterior["Número NL"].astype(str) == numero_nl)
                                )
                                detalhe_anterior.loc[mascara, "Excluir da semana"] = bool(
                                    alteracoes["Excluir da semana"]
                                )
                    st.session_state[detalhe_chave] = detalhe_anterior.copy()

                    linhas_prioridade = []
                    grupos_editados = []
                    for indice_objeto, linha in tabela_semana.reset_index(drop=True).iterrows():
                        objeto = str(linha.get("Objeto", "")).strip()
                        if not objeto:
                            continue
                        detalhes_objeto = detalhe_anterior[
                            detalhe_anterior["Objeto"].astype(str) == objeto
                        ].copy()
                        linhas_mantidas_objeto = detalhes_objeto.loc[
                            ~detalhes_objeto["Excluir da semana"].fillna(False)
                        ].copy()
                        total_objeto = float(
                            converter_valor_monetario(linhas_mantidas_objeto["Valor da NL"]).sum()
                        )
                        chave_objeto = re.sub(r"[^A-Za-z0-9]+", "_", objeto).strip("_")[:70]
                        with st.expander(
                            f"{objeto}   |   A PROGRAMAR: {formatar_brl(total_objeto)}",
                            expanded=False,
                        ):
                            cabecalho_objeto = st.columns([2, 2, 1])
                            cabecalho_objeto[0].markdown(f"**Disponível na NL:** {formatar_brl(total_objeto)}")
                            observacao = cabecalho_objeto[1].text_input(
                                "Observação",
                                value=str(linha.get("Observação", "") or ""),
                                key=f"obs_{chave_prioridades}_{semana}_{chave_objeto}",
                            )
                            remover = cabecalho_objeto[2].checkbox(
                                "Remover objeto",
                                value=False,
                                key=f"remover_{chave_prioridades}_{semana}_{chave_objeto}",
                            )
                            if remover:
                                continue
                            if objeto != objeto_detalhado:
                                st.caption("Selecione este objeto no campo acima para abrir os credores e as NLs.")
                            elif detalhes_objeto.empty:
                                st.warning("Não há NL detalhada para este objeto na fonte consultada.")
                            else:
                                for indice_credor, (credor, grupo_credor) in enumerate(
                                    detalhes_objeto.groupby("Credor", dropna=False, sort=False)
                                ):
                                    grupo_credor = grupo_credor.reset_index(drop=True)
                                    linhas_mantidas_credor = grupo_credor.loc[
                                        ~grupo_credor["Excluir da semana"].fillna(False)
                                    ].copy()
                                    total_credor = float(
                                        converter_valor_monetario(linhas_mantidas_credor["Valor da NL"]).sum()
                                    )
                                    nome_credor = str(credor).strip() if pd.notna(credor) else "Credor não informado"
                                    with st.expander(
                                        f"{nome_credor}   •   VALOR TOTAL: {formatar_brl(total_credor)}",
                                        expanded=False,
                                    ):
                                        grupo_editado = st.data_editor(
                                            grupo_credor[["Número NL", "Valor da NL", "Excluir da semana"]],
                                            column_config={
                                                "Número NL": st.column_config.TextColumn("Número NL", disabled=True, width="medium"),
                                                "Valor da NL": st.column_config.TextColumn("Valor da NL", disabled=True, width="medium"),
                                                "Excluir da semana": st.column_config.CheckboxColumn("Excluir", width="small"),
                                            },
                                            hide_index=True,
                                            use_container_width=True,
                                            height=min(48 + (len(grupo_credor) * 36), 260),
                                            key=f"editor_{detalhe_chave}_{indice_objeto}_{indice_credor}",
                                        )
                                    grupo_editado["Objeto"] = objeto
                                    grupo_editado["Credor"] = credor
                                    # As alterações do editor são aplicadas no
                                    # próximo rerun pelo estado do widget acima.
                            linhas_prioridade.append({
                                "Objeto": objeto,
                                "Valor Prioridade": total_objeto,
                                "Observação": observacao,
                                "Disponível na NL": formatar_brl(total_objeto),
                            })

                    prioridade_editada = pd.DataFrame(linhas_prioridade)
                    st.session_state[chave_prioridades][semana] = prioridade_editada.reset_index(drop=True).copy()
            prioridade_editada["Semana"] = semana
            st.session_state[chave_prioridades][semana] = prioridade_editada.drop(columns="Semana").reset_index(drop=True)
            prioridades_semanais.append(prioridade_editada)

    prioridades = pd.concat(prioridades_semanais, ignore_index=True)
    prioridades["Valor Prioridade"] = pd.to_numeric(prioridades["Valor Prioridade"], errors="coerce").fillna(0.0)
    prioridades = prioridades[prioridades["Objeto"].fillna("").astype(str).str.strip() != ""]
    # Validação de disponibilidade: registros já existentes na planilha são
    # histórico e não devem ser comparados novamente com a NL disponível hoje.
    programacao_historica = st.session_state.get("programacao_google_salva", pd.DataFrame()).copy()
    pares_historicos = set()
    if isinstance(programacao_historica, pd.DataFrame) and not programacao_historica.empty:
        hist = canonicalizar_programacao_priorizacao(programacao_historica.to_dict(orient="records"))
        hist = hist[(hist["Mes"] == normalizar_mes_priorizacao(mes_planejado)) & hist["Fonte"].astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False)]
        pares_historicos = set(zip(hist["Semana"].astype(str), hist["Objeto"].astype(str)))

    prioridades_validacao = prioridades.copy()
    if not prioridades_validacao.empty:
        prioridades_validacao["_historico"] = prioridades_validacao.apply(
            lambda linha: (str(linha.get("Semana", "")), str(linha.get("Objeto", ""))) in pares_historicos, axis=1
        )
        prioridades_validacao = prioridades_validacao[~prioridades_validacao["_historico"]].copy()
    priorizado_por_objeto = prioridades_validacao.groupby("Objeto", as_index=False)["Valor Prioridade"].sum() if not prioridades_validacao.empty else pd.DataFrame(columns=["Objeto", "Valor Prioridade"])
    priorizado_por_objeto["Limite NL"] = priorizado_por_objeto["Objeto"].map(limite_por_objeto).fillna(0.0)
    priorizado_por_objeto["Excedente"] = (priorizado_por_objeto["Valor Prioridade"] - priorizado_por_objeto["Limite NL"]).clip(lower=0)

    total_priorizado = prioridades["Valor Prioridade"].sum()
    total_nl = objetos_disponiveis["Valor_Executado"].sum()
    saldo_pactuado = pactuado_mensal - total_priorizado
    saldo_nl = total_nl - total_priorizado

    st.markdown("### 4. Confira o cenário antes de fechar")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("PACTUADO DO MÊS", formatar_brl(pactuado_mensal))
    k2.metric("NL DISPONÍVEL", formatar_brl(total_nl))
    k3.metric("JÁ PRIORIZADO", formatar_brl(total_priorizado))
    k4.metric("SALDO DO PACTUADO", formatar_brl(saldo_pactuado))

    if pactuado_mensal > 0:
        percentual_planejado = max(0.0, min(total_priorizado / pactuado_mensal, 1.0))
        percentual_visual = percentual_planejado * 100
        st.markdown(
            f"""
            <div class='progresso-priorizacao'>
                <span class='progresso-priorizacao-texto'>
                    Planejado: <strong>{formatar_brl(total_priorizado)}</strong>
                    de <strong>{formatar_brl(pactuado_mensal)}</strong>
                    ({percentual_planejado:.0%})
                </span>
                <div class='progresso-priorizacao-trilha'>
                    <div class='progresso-priorizacao-preenchimento' style='width: {percentual_visual:.2f}%;'></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.info("Defina o pactuado mensal para acompanhar o percentual já distribuído entre as semanas.")

    if not priorizado_por_objeto.empty and (priorizado_por_objeto["Excedente"] > 0).any():
        st.warning(
            "Há uma NOVA priorização acima do valor disponível atualmente na NL. "
            "O histórico já salvo nas semanas anteriores não entra nessa validação."
        )

    resumo_semanal = (
        prioridades.groupby("Semana", as_index=False)["Valor Prioridade"].sum()
        .set_index("Semana")
        .reindex(["Semana 1", "Semana 2", "Semana 3", "Semana 4"], fill_value=0.0)
        .reset_index()
    )

    # PACTUADO SEMANAL AUTOMÁTICO:
    # o que está registrado em cada semana é o pactuado daquela semana.
    # Isso vale tanto para o histórico restaurado da planilha (Semanas 1/2,
    # por exemplo) quanto para os objetos que o usuário adicionar agora.
    pactuados_semanais = {
        str(linha["Semana"]): float(linha["Valor Prioridade"])
        for _, linha in resumo_semanal.iterrows()
    }
    resumo_semanal["Meta semanal"] = resumo_semanal["Semana"].map(pactuados_semanais).fillna(0.0)

    # Persiste também os totais derivados. Eles não dependem de widgets, por
    # isso permanecem corretos quando o usuário troca entre OB, NL e Priorização.
    mascara_pactuado_estado = (
        st.session_state["pactuados_fonte500"]["Mês"].apply(normalizar_mes_priorizacao) == mes_planejado_normalizado
    ) & (
        st.session_state["pactuados_fonte500"]["Grupo"].fillna("").astype(str).str.upper().isin(["", chave_planejamento.upper()])
    )
    if mascara_pactuado_estado.any():
        indice_pactuado_estado = st.session_state["pactuados_fonte500"].loc[mascara_pactuado_estado].index[0]
        for numero in range(1, 5):
            st.session_state["pactuados_fonte500"].loc[
                indice_pactuado_estado, f"Pactuado Semana {numero}"
            ] = float(pactuados_semanais.get(f"Semana {numero}", 0.0))

    # O saldo exibido por semana representa o saldo ACUMULADO do teto mensal
    # depois do que foi programado até aquela semana. Assim a gestão consegue
    # acompanhar quanto ainda resta do mês após Semana 1, Semana 2, etc.
    # O pactuado semanal permanece independente e continua sendo usado para
    # medir o cumprimento da meta específica de cada semana.
    resumo_semanal["Programado acumulado"] = resumo_semanal["Valor Prioridade"].cumsum()
    resumo_semanal["Saldo da semana"] = pactuado_mensal - resumo_semanal["Programado acumulado"]

    resumo_semanal["Percentual da meta"] = np.where(
        resumo_semanal["Meta semanal"] > 0,
        (resumo_semanal["Valor Prioridade"] / resumo_semanal["Meta semanal"] * 100),
        0.0,
    )
    resumo_semanal["Percentual exibido"] = resumo_semanal["Percentual da meta"].clip(upper=100)
    resumo_semanal["Rótulo do gráfico"] = resumo_semanal.apply(
        lambda linha: f"{formatar_brl(linha['Valor Prioridade'])}  ({linha['Percentual da meta']:.0f}%)",
        axis=1,
    )
    resumo_semanal["Situação"] = np.select(
        [
            resumo_semanal["Percentual da meta"] >= 100,
            resumo_semanal["Percentual da meta"] >= 75,
        ],
        ["Meta atingida", "Em andamento"],
        default="A priorizar",
    )
    st.markdown("### 5. Resumo consolidado para envio")
    st.caption("Conferência final por semana. O pactuado semanal é calculado automaticamente a partir dos objetos/NLs registrados em cada semana.")

    # A saída é operacional: uma linha para cada NL que será paga, e não um
    # total consolidado por objeto. Assim o arquivo pode seguir direto para a
    # conferência/execução dos pagamentos.
    linhas_programacao = []
    for _, prioridade in prioridades.iterrows():
        semana_programada = str(prioridade.get("Semana", ""))
        objeto_programado = str(prioridade.get("Objeto", "")).strip()
        if not objeto_programado:
            continue
        detalhe_chave = f"detalhe_nl_{chave_prioridades}_{semana_programada}"
        detalhe_salvo = st.session_state.get(detalhe_chave, pd.DataFrame())
        if isinstance(detalhe_salvo, pd.DataFrame) and not detalhe_salvo.empty:
            nls_programadas = detalhe_salvo.loc[
                (detalhe_salvo["Objeto"].astype(str) == objeto_programado)
                & ~detalhe_salvo["Excluir da semana"].fillna(False)
            ].copy()
        else:
            # Fallback para uma semana recém-montada que ainda não criou estado
            # detalhado. Mantém o comportamento anterior para novos cenários.
            nls_programadas = base_detalhamento_nl[
                base_detalhamento_nl["Objeto"].astype(str) == objeto_programado
            ].copy()
            if "Valor_Executado" in nls_programadas.columns:
                nls_programadas["Valor da NL"] = nls_programadas["Valor_Executado"]

        for _, nl in nls_programadas.iterrows():
            valor_nl = converter_valor_monetario(
                pd.Series([nl.get("Valor da NL", nl.get("Valor_Executado", 0.0))])
            ).iloc[0]
            valor_nl = float(valor_nl) if pd.notna(valor_nl) and np.isfinite(valor_nl) else 0.0
            linhas_programacao.append({
                "Mês de programação": mes_planejado,
                "Fonte": "500",
                "Grupo": nl.get("Grupo", ""),
                "Tipo de NL": nl.get("Tipo de NL", ""),
                "Semana": semana_programada,
                "Objeto": objeto_programado,
                "Número NL": nl.get("Número NL", ""),
                "Credor": nl.get("Credor", ""),
                "Valor programado": valor_nl,
                "Observação": prioridade.get("Observação", ""),
            })
    colunas_envio = [
        "Mês de programação", "Fonte", "Grupo", "Tipo de NL", "Semana", "Objeto",
        "Número NL", "Credor", "Valor programado", "Observação",
    ]
    programacao_consolidada = pd.DataFrame(linhas_programacao, columns=colunas_envio)
    if not programacao_consolidada.empty:
        programacao_consolidada = programacao_consolidada.sort_values(
            ["Semana", "Grupo", "Tipo de NL", "Objeto", "Credor", "Número NL"]
        ).reset_index(drop=True)

    # A planilha é gravada somente por solicitação. Fazer um POST a cada clique
    # de priorização bloqueava a tela e criava filas no Apps Script.
    pactuados_api = st.session_state["pactuados_fonte500"].rename(
        columns={
            "Mês": "Mes", "Pactuado Mensal": "Valor_Pactuado",
            "Pactuado Semana 1": "Pactuado_Semana_1",
            "Pactuado Semana 2": "Pactuado_Semana_2",
            "Pactuado Semana 3": "Pactuado_Semana_3",
            "Pactuado Semana 4": "Pactuado_Semana_4",
        }
    ).copy()
    pactuados_api["Fonte"] = "500"
    pactuados_api["Atualizado_Em"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pactuados_api = pactuados_api.reindex(columns=[
        "Mes", "Fonte", "Grupo", "Valor_Pactuado",
        "Pactuado_Semana_1", "Pactuado_Semana_2", "Pactuado_Semana_3", "Pactuado_Semana_4",
        "Atualizado_Em",
    ])
    pactuado_mes_api = pactuados_api[
        (pactuados_api["Mes"].apply(normalizar_mes_priorizacao) == normalizar_mes_priorizacao(mes_planejado))
        & (pactuados_api["Grupo"].fillna("").astype(str).str.upper().isin(["", chave_planejamento.upper()]))
    ].head(1).copy()

    programacao_atual_api = programacao_consolidada.rename(
        columns={
            "Mês de programação": "Mes",
            "Tipo de NL": "Tipo_NL",
            "Número NL": "Numero_NL",
            "Valor programado": "Valor_Programado",
            "Observação": "Observacao",
        }
    ).copy()
    programacao_atual_api["Atualizado_Em"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    colunas_api_programacao = [
        "Mes", "Fonte", "Grupo", "Tipo_NL", "Semana", "Objeto", "Numero_NL",
        "Credor", "Valor_Programado", "Observacao", "Atualizado_Em",
    ]
    programacao_atual_api = programacao_atual_api.reindex(columns=colunas_api_programacao)

    programacao_remota = st.session_state.get("programacao_google_salva", pd.DataFrame()).copy()
    programacao_remota = programacao_remota.rename(
        columns={
            "Mês de programação": "Mes",
            "Tipo de NL": "Tipo_NL",
            "Número NL": "Numero_NL",
            "Valor programado": "Valor_Programado",
            "Observação": "Observacao",
        }
    )
    for coluna in colunas_api_programacao:
        if coluna not in programacao_remota.columns:
            programacao_remota[coluna] = "" if coluna != "Valor_Programado" else 0.0
    programacao_remota = programacao_remota[colunas_api_programacao].copy()
    substituir_recorte = (
        (programacao_remota["Mes"].astype(str) == mes_planejado)
        & (programacao_remota["Grupo"].astype(str).isin([str(grupo) for grupo in grupos_aplicados]))
    )
    programacao_para_salvar = pd.concat(
        [programacao_remota.loc[~substituir_recorte], programacao_atual_api],
        ignore_index=True,
    )
    # Assinaturas separadas permitem salvar somente o que realmente mudou.
    colunas_ordem_assinatura = ["Semana", "Grupo", "Tipo_NL", "Objeto", "Numero_NL", "Credor", "Valor_Programado"]
    atual_assinatura_df = programacao_atual_api.drop(columns=["Atualizado_Em"], errors="ignore").copy()
    if not atual_assinatura_df.empty:
        atual_assinatura_df = atual_assinatura_df.sort_values(
            [c for c in colunas_ordem_assinatura if c in atual_assinatura_df.columns], kind="stable"
        ).reset_index(drop=True)
    assinatura_programacao_atual = json.dumps(
        _registros_json(atual_assinatura_df), ensure_ascii=False, sort_keys=True, default=str,
    )
    remoto_mes_assinatura = programacao_remota[
        (programacao_remota["Mes"].apply(normalizar_mes_priorizacao) == normalizar_mes_priorizacao(mes_planejado))
        & programacao_remota["Fonte"].astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False)
    ].copy()
    remoto_assinatura_df = remoto_mes_assinatura.drop(columns=["Atualizado_Em"], errors="ignore").copy()
    if not remoto_assinatura_df.empty:
        remoto_assinatura_df = remoto_assinatura_df.sort_values(
            [c for c in colunas_ordem_assinatura if c in remoto_assinatura_df.columns], kind="stable"
        ).reset_index(drop=True)
    assinatura_programacao_remota = json.dumps(
        _registros_json(remoto_assinatura_df), ensure_ascii=False, sort_keys=True, default=str,
    )
    assinatura_pactuado_atual = json.dumps(
        _registros_json(pactuado_mes_api.drop(columns=["Atualizado_Em"], errors="ignore")),
        ensure_ascii=False, sort_keys=True, default=str,
    )

    # O último pactuado efetivamente lido da API é guardado para detectar se só
    # o teto/distribuição semanal mudou.
    pactuado_remoto_canon = canonicalizar_pactuados_priorizacao(retorno_google.get("pactuados", []))
    pactuado_remoto_mes = pactuado_remoto_canon[
        (pactuado_remoto_canon["Mes"] == normalizar_mes_priorizacao(mes_planejado))
        & pactuado_remoto_canon["Fonte"].astype(str).str.contains(r"(?<!\d)500(?!\d)", regex=True, na=False)
    ].head(1).copy()
    assinatura_pactuado_remoto = json.dumps(
        _registros_json(pactuado_remoto_mes.drop(columns=["Atualizado_Em"], errors="ignore")),
        ensure_ascii=False, sort_keys=True, default=str,
    )
    mudou_programacao = assinatura_programacao_atual != assinatura_programacao_remota
    mudou_pactuado = assinatura_pactuado_atual != assinatura_pactuado_remoto
    assinatura_persistencia = json.dumps({
        "pactuado": assinatura_pactuado_atual,
        "programacao": assinatura_programacao_atual,
    }, sort_keys=True)

    salvar_coluna, situacao_coluna = st.columns([0.28, 0.72])
    with salvar_coluna:
        salvar_agora = st.button(
            "💾 Salvar priorização", key="salvar_priorizacao_google", type="primary", use_container_width=True,
        )
    if salvar_agora:
        with st.spinner("Gravando somente as alterações deste mês..."):
            try:
                if mudou_programacao:
                    salvar_mes_priorizacao_google(
                        mes_planejado, "500", pactuado_mes_api, programacao_atual_api
                    )
                    mensagem_salvamento = "Pactuado e programação do mês salvos."
                elif mudou_pactuado:
                    salvar_pactuado_google(pactuado_mes_api)
                    mensagem_salvamento = "Pactuado mensal/semanal atualizado sem regravar a programação."
                else:
                    mensagem_salvamento = "Nenhuma alteração pendente para gravar."

                st.session_state["assinatura_priorizacao_google"] = assinatura_persistencia
                # Mantém o estado local com o recorte atual; a próxima atualização
                # manual confirmará a versão persistida no Google Sheets.
                st.session_state["programacao_google_salva"] = programacao_para_salvar.copy()
                st.session_state["nonce_atualizacao_priorizacao"] = st.session_state.get("nonce_atualizacao_priorizacao", 0) + 1
                carregar_priorizacao_google_sincronizada.clear()
                st.session_state["priorizacao_google_erro"] = ""
                st.session_state["priorizacao_google_salva_em"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                st.session_state["priorizacao_google_mensagem"] = mensagem_salvamento
            except Exception as erro:
                st.session_state["priorizacao_google_erro"] = str(erro)

    with situacao_coluna:
        if st.session_state.get("priorizacao_google_erro"):
            st.warning(
                "Não foi possível gravar na planilha vinculada: "
                f"{st.session_state['priorizacao_google_erro']}"
            )
        elif st.session_state.get("assinatura_priorizacao_google") == assinatura_persistencia:
            st.success(
                st.session_state.get("priorizacao_google_mensagem", "Priorização salva na planilha vinculada.")
                + (f" Em {st.session_state.get('priorizacao_google_salva_em')}." if st.session_state.get("priorizacao_google_salva_em") else "")
            )
        elif mudou_programacao or mudou_pactuado:
            partes = []
            if mudou_pactuado:
                partes.append("pactuado")
            if mudou_programacao:
                partes.append("programação")
            st.info("Alterações pendentes em " + " e ".join(partes) + ". Clique em ‘Salvar priorização’.")
        else:
            st.success("Os dados exibidos estão sincronizados com a planilha.")

    # Resumo executivo: o pactuado mensal continua sendo usado internamente
    # para calcular o saldo, mas não é repetido como coluna na grade. Em seu
    # lugar mostramos a quantidade de NLs efetivamente programadas por semana.
    if programacao_consolidada.empty:
        qtd_nl_por_semana = pd.Series(dtype="int64")
        qtd_nl_total = 0
    else:
        programacao_qtd = programacao_consolidada.copy()
        programacao_qtd["Número NL"] = programacao_qtd["Número NL"].fillna("").astype(str).str.strip()
        programacao_qtd = programacao_qtd[programacao_qtd["Número NL"] != ""]
        qtd_nl_por_semana = programacao_qtd.groupby("Semana")["Número NL"].nunique()
        qtd_nl_total = int(programacao_qtd["Número NL"].nunique())

    resumo_envio = resumo_semanal[["Semana", "Valor Prioridade", "Saldo da semana"]].copy()
    resumo_envio["Qtd. NL"] = resumo_envio["Semana"].map(qtd_nl_por_semana).fillna(0).astype(int)
    resumo_envio = resumo_envio[["Semana", "Qtd. NL", "Valor Prioridade", "Saldo da semana"]].rename(
        columns={"Valor Prioridade": "Valor programado"}
    )

    total_resumo = pd.DataFrame([{
        "Semana": "TOTAL GERAL",
        "Qtd. NL": qtd_nl_total,
        "Valor programado": total_priorizado,
        "Saldo da semana": saldo_pactuado,
    }])
    resumo_envio = pd.concat([resumo_envio, total_resumo], ignore_index=True)

    col_resumo_envio, col_exportar = st.columns([1.5, 0.5])
    with col_resumo_envio:
        resumo_visivel = resumo_envio.copy()
        for coluna_moeda in ["Valor programado", "Saldo da semana"]:
            resumo_visivel[coluna_moeda] = resumo_visivel[coluna_moeda].apply(formatar_brl)
        resumo_visivel["Qtd. NL"] = resumo_visivel["Qtd. NL"].fillna(0).astype(int).astype(str)
        resumo_visivel = resumo_visivel.rename(columns={"Saldo da semana": "Saldo"})
        renderizar_tabela_priorizacao(
            resumo_visivel,
            colunas_centralizadas={"Semana", "Qtd. NL", "Valor programado", "Saldo"},
        )
    with col_exportar:
        st.metric("TOTAL PROGRAMADO", formatar_brl(total_priorizado))
        st.metric("SALDO DO PACTUADO", formatar_brl(saldo_pactuado))

    st.markdown("#### Detalhamento consolidado por semana")
    programacao_visivel = programacao_consolidada.copy()
    if "Valor programado" in programacao_visivel.columns:
        programacao_visivel["Valor programado"] = programacao_visivel["Valor programado"].apply(formatar_brl)
    colunas_visualizacao = [
        "Semana", "Grupo", "Tipo de NL", "Objeto", "Número NL", "Credor",
        "Valor programado", "Observação",
    ]
    programacao_visivel = programacao_visivel.reindex(columns=colunas_visualizacao)
    renderizar_tabela_priorizacao(
        programacao_visivel,
        colunas_centralizadas={"Semana", "Grupo", "Tipo de NL", "Número NL", "Valor programado"},
        altura_maxima=360,
    )

    arquivo_programacao = io.BytesIO()
    with pd.ExcelWriter(arquivo_programacao, engine="xlsxwriter") as writer:
        programacao_consolidada.to_excel(writer, sheet_name="Programação Semanal", startrow=2, index=False)
        resumo_envio.to_excel(writer, sheet_name="Resumo por Semana", startrow=2, index=False)
        workbook = writer.book
        formato_titulo = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#002B49", "font_size": 14, "align": "center"})
        formato_cabecalho = workbook.add_format({"bold": True, "font_color": "#FFFFFF", "bg_color": "#315B85", "align": "center", "valign": "vcenter"})
        formato_moeda = workbook.add_format({"num_format": "R$ #,##0.00", "align": "right"})
        formato_qtd = workbook.add_format({"num_format": "0", "align": "center"})
        formato_total = workbook.add_format({"bold": True, "bg_color": "#E8F0F7", "top": 2, "top_color": "#002B49", "num_format": "R$ #,##0.00", "align": "right"})
        formato_total_qtd = workbook.add_format({"bold": True, "bg_color": "#E8F0F7", "top": 2, "top_color": "#002B49", "num_format": "0", "align": "center"})
        for nome_aba, titulo, dados_exportar in [
            ("Programação Semanal", "Programação Semanal — Fonte 500", programacao_consolidada),
            ("Resumo por Semana", "Resumo consolidado por semana — Fonte 500", resumo_envio),
        ]:
            aba_exportar = writer.sheets[nome_aba]
            ultima_coluna = len(dados_exportar.columns) - 1
            aba_exportar.hide_gridlines(2)
            aba_exportar.merge_range(0, 0, 0, ultima_coluna, titulo, formato_titulo)
            aba_exportar.set_row(0, 24)
            aba_exportar.freeze_panes(3, 0)
            aba_exportar.autofilter(2, 0, len(dados_exportar) + 2, ultima_coluna)
            for coluna, cabecalho in enumerate(dados_exportar.columns):
                aba_exportar.write(2, coluna, cabecalho, formato_cabecalho)
            aba_exportar.set_column(0, ultima_coluna, 18)
        aba_programacao = writer.sheets["Programação Semanal"]
        aba_programacao.set_column("A:A", 18)
        aba_programacao.set_column("B:C", 12)
        aba_programacao.set_column("D:D", 12)
        aba_programacao.set_column("E:E", 10)
        aba_programacao.set_column("F:F", 55)
        aba_programacao.set_column("G:G", 18)
        aba_programacao.set_column("H:H", 50)
        aba_programacao.set_column("I:I", 18, formato_moeda)
        aba_programacao.set_column("J:J", 35)
        aba_resumo = writer.sheets["Resumo por Semana"]
        aba_resumo.set_column("A:A", 20)
        aba_resumo.set_column("B:B", 12, formato_qtd)
        aba_resumo.set_column("C:D", 20, formato_moeda)
        linha_total = len(resumo_envio) + 2
        for coluna in range(len(resumo_envio.columns)):
            valor = resumo_envio.iloc[-1, coluna]
            if coluna == 0:
                aba_resumo.write(linha_total, coluna, valor, workbook.add_format({"bold": True, "bg_color": "#E8F0F7", "top": 2, "top_color": "#002B49"}))
            else:
                # O XlsxWriter não aceita NaN/inf em write_number. Se algum
                # campo do resumo estiver vazio, preservamos a célula vazia
                # em vez de interromper a geração do arquivo.
                numero = pd.to_numeric(pd.Series([valor]), errors="coerce").iloc[0]
                formato_coluna = formato_total_qtd if resumo_envio.columns[coluna] == "Qtd. NL" else formato_total
                if pd.notna(numero) and np.isfinite(numero):
                    aba_resumo.write_number(linha_total, coluna, float(numero), formato_coluna)
                else:
                    aba_resumo.write_blank(linha_total, coluna, None, formato_coluna)
    arquivo_programacao.seek(0)
    st.download_button(
        "📥 Baixar resumo consolidado (.xlsx)",
        arquivo_programacao.getvalue(),
        file_name=f"programacao_semanal_fonte500_{mes_planejado.replace('/', '-')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.stop()

    def ordenar_mes(valor):
        try:
            return datetime.datetime.strptime(str(valor), "%m/%Y")
        except ValueError:
            return datetime.datetime.max

    meses_planejamento = sorted(dados_planejamento["Mês"].dropna().unique(), key=ordenar_mes)
    fontes_planejamento = sorted(dados_planejamento["Fonte"].dropna().astype(str).unique())
    grupos_planejamento = sorted(dados_planejamento["Grupo"].dropna().astype(str).unique())
    objetos_planejamento = sorted(dados_planejamento["Objeto"].dropna().astype(str).unique())

    st.sidebar.markdown("#### 🎯 Filtros — Planejamento NL")
    filtro_mes_plan = st.sidebar.multiselect("Mês de referência", meses_planejamento, key="plan_meses")
    filtro_fonte_plan = st.sidebar.multiselect("Fonte de recurso", fontes_planejamento, key="plan_fontes")
    filtro_grupo_plan = st.sidebar.multiselect("Grupo", grupos_planejamento, key="plan_grupos")
    filtro_objeto_plan = st.sidebar.multiselect("Objeto da despesa", objetos_planejamento, key="plan_objetos")

    realizado = dados_planejamento.copy()
    if filtro_mes_plan:
        realizado = realizado[realizado["Mês"].isin(filtro_mes_plan)]
    if filtro_fonte_plan:
        realizado = realizado[realizado["Fonte"].isin(filtro_fonte_plan)]
    if filtro_grupo_plan:
        realizado = realizado[realizado["Grupo"].isin(filtro_grupo_plan)]
    if filtro_objeto_plan:
        realizado = realizado[realizado["Objeto"].isin(filtro_objeto_plan)]

    modelo_pactuado = (
        dados_planejamento.groupby(["Mês", "Fonte", "Grupo", "Objeto"], as_index=False)["Valor_Executado"]
        .sum()
        .drop(columns="Valor_Executado")
    )
    modelo_pactuado["Pactuado Mensal"] = 0.0
    modelo_pactuado["Meta Semanal"] = 0.0

    arquivo_modelo = modelo_pactuado.head(0).to_csv(index=False).encode("utf-8-sig")
    col_importar, col_modelo = st.columns([2, 1])
    with col_importar:
        arquivo_pactuado = st.file_uploader(
            "Importar pactuado (.xlsx ou .csv)",
            type=["xlsx", "csv"],
            help="Colunas esperadas: Mês, Fonte, Grupo, Objeto, Pactuado Mensal e Meta Semanal.",
        )
    with col_modelo:
        st.download_button(
            "⬇️ Baixar modelo de pactuado",
            arquivo_modelo,
            file_name="modelo_pactuado_nl.csv",
            mime="text/csv",
            use_container_width=True,
        )

    plano_importado = modelo_pactuado.copy()
    if arquivo_pactuado is not None:
        try:
            if arquivo_pactuado.name.lower().endswith(".csv"):
                plano_importado = pd.read_csv(arquivo_pactuado, sep=None, engine="python")
            else:
                plano_importado = pd.read_excel(arquivo_pactuado)
            plano_importado.columns = [str(c).strip() for c in plano_importado.columns]
            obrigatorias = ["Mês", "Fonte", "Grupo", "Objeto", "Pactuado Mensal", "Meta Semanal"]
            faltantes = [c for c in obrigatorias if c not in plano_importado.columns]
            if faltantes:
                raise ValueError("Colunas ausentes: " + ", ".join(faltantes))
            plano_importado = plano_importado[obrigatorias].copy()
        except Exception as erro:
            st.error(f"Não foi possível ler o arquivo de pactuado: {erro}")
            plano_importado = modelo_pactuado.copy()

    identificador_arquivo = (
        f"{arquivo_pactuado.name}:{len(arquivo_pactuado.getvalue())}:{hash(arquivo_pactuado.getvalue())}"
        if arquivo_pactuado is not None
        else None
    )
    if (
        "plano_nl_editavel" not in st.session_state
        or (
            identificador_arquivo is not None
            and st.session_state.get("arquivo_pactuado_atual") != identificador_arquivo
        )
    ):
        st.session_state["plano_nl_editavel"] = plano_importado
        st.session_state["arquivo_pactuado_atual"] = identificador_arquivo

    with st.expander("✏️ Cadastrar ou ajustar pactuados", expanded=arquivo_pactuado is None):
        st.caption("Use uma linha por Mês, Fonte, Grupo e Objeto. Depois, baixe a configuração para guardá-la e reutilizá-la.")
        plano_editado = st.data_editor(
            st.session_state["plano_nl_editavel"],
            column_config={
                "Pactuado Mensal": st.column_config.NumberColumn("Pactuado Mensal", min_value=0.0, format="R$ %.2f"),
                "Meta Semanal": st.column_config.NumberColumn("Meta Semanal", min_value=0.0, format="R$ %.2f"),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="editor_pactuado_nl",
        )
        st.session_state["plano_nl_editavel"] = plano_editado
        st.download_button(
            "💾 Baixar configuração de pactuados",
            plano_editado.to_csv(index=False).encode("utf-8-sig"),
            file_name="pactuado_nl_configurado.csv",
            mime="text/csv",
        )

    plano = plano_editado.copy()
    for coluna in ["Pactuado Mensal", "Meta Semanal"]:
        plano[coluna] = pd.to_numeric(plano[coluna], errors="coerce").fillna(0.0)
    for coluna in ["Mês", "Fonte", "Grupo", "Objeto"]:
        plano[coluna] = plano[coluna].fillna("").astype(str).str.strip()

    executado = realizado.groupby(["Mês", "Fonte", "Grupo", "Objeto"], as_index=False)["Valor_Executado"].sum()
    comparativo = executado.merge(plano, on=["Mês", "Fonte", "Grupo", "Objeto"], how="left")
    comparativo[["Pactuado Mensal", "Meta Semanal"]] = comparativo[["Pactuado Mensal", "Meta Semanal"]].fillna(0.0)
    comparativo["Saldo a Executar"] = comparativo["Pactuado Mensal"] - comparativo["Valor_Executado"]
    comparativo["% Execução"] = np.where(
        comparativo["Pactuado Mensal"] > 0,
        comparativo["Valor_Executado"] / comparativo["Pactuado Mensal"],
        np.nan,
    )
    comparativo["Situação"] = np.select(
        [
            comparativo["Pactuado Mensal"] <= 0,
            comparativo["% Execução"] >= 1,
            comparativo["% Execução"] >= 0.75,
        ],
        ["Sem pactuado", "Meta atingida", "Em atenção"],
        default="Risco de não executar",
    )

    pactuado_total = comparativo["Pactuado Mensal"].sum()
    executado_total = comparativo["Valor_Executado"].sum()
    saldo_total = pactuado_total - executado_total
    percentual_total = executado_total / pactuado_total if pactuado_total else 0.0
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("PACTUADO", formatar_brl(pactuado_total))
    k2.metric("EXECUTADO (NL)", formatar_brl(executado_total))
    k3.metric("SALDO A EXECUTAR", formatar_brl(saldo_total))
    k4.metric("EXECUÇÃO DO PACTUADO", f"{percentual_total:.1%}")

    resumo_mes = comparativo.groupby("Mês", as_index=False).agg(
        Pactuado=("Pactuado Mensal", "sum"), Executado=("Valor_Executado", "sum"), Meta_Semanal=("Meta Semanal", "sum")
    )
    resumo_mes["ordem"] = resumo_mes["Mês"].map(ordenar_mes)
    resumo_mes = resumo_mes.sort_values("ordem").drop(columns="ordem")
    col_grafico, col_resumo = st.columns([1.2, 0.8])
    with col_grafico:
        st.markdown("#### Pactuado × Executado por mês")
        grafico_mes = px.bar(
            resumo_mes.melt(id_vars="Mês", value_vars=["Pactuado", "Executado"], var_name="Indicador", value_name="Valor"),
            x="Mês", y="Valor", color="Indicador", barmode="group",
            color_discrete_map={"Pactuado": "#f77f00", "Executado": "#028090"},
            text_auto=".3s",
        )
        grafico_mes.update_layout(yaxis_tickprefix="R$ ", legend_title_text="", margin=dict(l=10, r=10, t=25, b=10))
        st.plotly_chart(grafico_mes, use_container_width=True)
    with col_resumo:
        st.markdown("#### Meta semanal consolidada")
        meta_semana = comparativo["Meta Semanal"].sum()
        st.metric("META SEMANAL", formatar_brl(meta_semana))
        st.caption("Cadastre a meta semanal no pactuado para acompanhar o ritmo de execução nas reuniões semanais.")
        st.dataframe(
            resumo_mes[["Mês", "Pactuado", "Executado", "Meta_Semanal"]],
            column_config={
                "Pactuado": st.column_config.NumberColumn(format="R$ %.2f"),
                "Executado": st.column_config.NumberColumn(format="R$ %.2f"),
                "Meta_Semanal": st.column_config.NumberColumn("Meta semanal", format="R$ %.2f"),
            },
            hide_index=True,
            use_container_width=True,
        )

    st.markdown("#### Objetos que exigem acompanhamento")
    criticos = comparativo.groupby("Objeto", as_index=False).agg(
        Pactuado=("Pactuado Mensal", "sum"), Executado=("Valor_Executado", "sum"), Saldo=("Saldo a Executar", "sum")
    )
    criticos["% Execução"] = np.where(criticos["Pactuado"] > 0, criticos["Executado"] / criticos["Pactuado"], np.nan)
    criticos["Situação"] = np.select(
        [criticos["Pactuado"] <= 0, criticos["% Execução"] >= 1, criticos["% Execução"] >= 0.75],
        ["Sem pactuado", "Meta atingida", "Em atenção"], default="Risco de não executar",
    )
    criticos = criticos.sort_values(["Situação", "Saldo"], ascending=[True, False])
    st.dataframe(
        criticos,
        column_config={
            "Pactuado": st.column_config.NumberColumn(format="R$ %.2f"),
            "Executado": st.column_config.NumberColumn(format="R$ %.2f"),
            "Saldo": st.column_config.NumberColumn("Saldo a executar", format="R$ %.2f"),
            "% Execução": st.column_config.NumberColumn(format="%.1f%%"),
        },
        hide_index=True,
        use_container_width=True,
        height=420,
    )
